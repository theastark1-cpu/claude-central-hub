#!/usr/bin/env python3
"""Build the Armada Prime Tech LLC 2025 year-end reconciliation workbook.

Aggregates Aug-Dec 2025 TPA Reporting Packages into a single workbook for
1099-NEC and K-1 prep. Uses the Dec 2025 BEST ONE Monthly Return for the
investor->consultant IDS mapping.

Outputs:
    /Users/nairne/claude-central-hub/2025-armada-prime-tech-1099-k1.xlsx
    /Users/nairne/claude-central-hub/2025-armada-prime-tech-summary.md

Usage:
    python tools/build_2025_year_end.py
"""
from __future__ import annotations

import sys
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
from build_consultant_splits import (
    CONSULTANT_OVERRIDES,
    CONSULTANT_SPLITS,
    load_ids,
    resolve_consultant_split,
)
from parse_tpa_report import parse_workbook

# Dec 2025 new investors that lack TPA IDs in the BEST ONE Dec 2025 IDS sheet
# (rows 43-47 only have Position IDs). Mapped from IDS by name match, with the
# John Kirkham fix from commit a513847 (per March 2026 NAV: John Kirkham → Luke).
EXTRA_2025_OVERRIDES = {
    "14-Class B-1057-1": "Alec Atkinson",   # PHF 2008 Descendants Separate Trust
    "14-Class B-1058-1": "Alec Atkinson",   # PL Investment Group, LLC
    "14-Class B-1059-1": "Luke",            # John Charles Kirkham (NAV-fix override)
    "14-Class B-1060-1": "Luke",            # Daniel Welch Kirkham
    "14-Class B-1061-1": "AJ Affleck",      # Craig Levinson — split 50/50 via CONSULTANT_SPLITS
}

REPO_ROOT = Path(__file__).resolve().parent.parent
OUT_XLSX = REPO_ROOT / "2025-armada-prime-tech-1099-k1.xlsx"
OUT_MD = REPO_ROOT / "2025-armada-prime-tech-summary.md"

# 2025 GP split — Phil holds the 0.5% slice (Alec replaced Phil in 2026-04).
# Per Nairne 2026-04-30: the 59.5% "Fund Mgmt" slice IS Nairne's income (not
# a separate entity 1099 expense). So Nairne's economic ownership is 60.0%
# (59.5 Fund Mgmt + 0.5 direct), Raj is 0.5%, Phil is 0.5% (1099 contractor),
# Consultants pool is 39%.
SPLIT_PCTS_2025 = {
    "fund_mgmt": 0.595,
    "consultant": 0.39,
    "raj": 0.005,
    "nairne": 0.005,
    "phil": 0.005,
}

# Recipients in ACTUAL_PAID that are partner allocations (K-1 income to
# Nairne or Raj), NOT contractor 1099 expenses.
NAIRNE_ALIASES = {"Fund Mgmt", "Nairne"}
RAJ_ALIASES = {"Raj"}
K1_RECIPIENTS = NAIRNE_ALIASES | RAJ_ALIASES

# Dec 2025 BEST ONE for IDS mapping (most complete 2025 mapping).
INTERNAL_IDS = Path("/Users/nairne/Downloads/BEST ONE of December 2025 Monthly Return.xlsx")

TPA_FILES = [
    ("2025-08", Path("/Users/nairne/Downloads/ArmadaPrimeLLP__ReportingPackage_2025-08-31/Armada_Prime_LLP_Reporting Package_2025-08-31_1758884151431.xlsx")),
    ("2025-09", Path("/Users/nairne/Downloads/ArmadaPrimeLLP__ReportingPackage_2025-09-30/Armada_Prime_LLP_Reporting Package_2025-09-30_1760700181988.xlsx")),
    ("2025-10", Path("/Users/nairne/Downloads/ArmadaPrimeLLP__ReportingPackage_2025-10-31 (1)/Armada_Prime_LLP_Reporting Package_2025-10-31_1763827473868.xlsx")),
    ("2025-11", Path("/Users/nairne/Downloads/ArmadaPrimeLLP__ReportingPackage_2025-11-30/Armada_Prime_LLP_Reporting Package_2025-11-30_1767270591715.xlsx")),
    ("2025-12", Path("/Users/nairne/Downloads/ArmadaPrimeLLP__ReportingPackage_2025-12-31/Armada_Prime_LLP_Reporting Package_2025-12-31_1769516671447.xlsx")),
]

PERIOD_LABELS = {
    "2025-08": "Aug 2025",
    "2025-09": "Sep 2025",
    "2025-10": "Oct 2025",
    "2025-11": "Nov 2025",
    "2025-12": "Dec 2025",
}

# Per-consultant ACTUAL paid amounts from the internal Distributions Armada
# Technologies 2025 (INTERNAL ONLY) ledger. These are operational cash-tracking
# figures and may differ from TPA-derived (accrual) amounts.
#
# IMPORTANT August anomaly: Aug used a different waterfall (9.5% consultant +
# 13.5% Trader & Developer + 5.5% Mgmt + 1.5% Raj/Nairne/Phil = 30% GP cut).
# From Sep onwards, TruQuant takes 18% UPSTREAM and the GP uses the standard
# 59.5/39/0.5/0.5/0.5 split.
ACTUAL_PAID = {
    # Aug 2025 from Distributions xlsx — sheet "August" R2-R14
    # TruQuant entries (Trader & Developer $6,909.93 and Spydr $88.78) are
    # excluded per Nairne 2026-04-30 — TQ is not a GP expense / 1099 recipient.
    "2025-08": {
        "AJ Affleck": 335.47,
        "Alec Atkinson": 3172.12,
        "Jake Gordon": 1266.18,
        "Fund Mgmt": 2815.16,
        "Raj": 255.92,
        "Nairne": 255.92,
        "Phil": 255.92,
    },
    # Sep 2025 from Distributions xlsx — sheet "September" R6-R15
    # Spydr/TruQuant entry ($82.08) excluded per Nairne 2026-04-30.
    "2025-09": {
        "AJ Affleck": 285.77,
        "Alec Atkinson": 3561.92,
        "Jake Gordon": 930.90,
        "Fund Mgmt": 7415.62,
        "Raj": 62.32,
        "Nairne": 62.32,
        "Phil": 62.32,
    },
    # Oct 2025 from Distributions xlsx — sheet "Copy of October" R7-R18
    "2025-10": {
        "AJ Affleck": 1268.23,
        "Alec Atkinson": 11848.88,
        "Jake Gordon": 3218.29,
        "Issac": 254.12,
        "Fund Mgmt": 31206.94,
        "Raj": 262.24,
        "Nairne": 262.24,
        "Phil": 262.24,
    },
    # Nov 2025 from Distributions xlsx — sheet "November" R7-R18
    "2025-11": {
        "AJ Affleck": 1833.52,
        "Alec Atkinson": 15124.18,
        "Jake Gordon": 4184.51,
        "Issac": 785.85,
        "Fund Mgmt": 34183.90,
        "Raj": 287.26,
        "Nairne": 287.26,
        "Phil": 287.26,
    },
    # Dec 2025 from BEST ONE Dec 2025 Monthly Return — sheet "Consultants"
    # (see Bash exploration earlier — Alec $4,369.77, Jake $788.78, AJ $429.71,
    # Luke $164.90, Isaac -$278.48, Fund Mgmt $9,428.24, Raj/Nairne/Phil $79.23 ea)
    "2025-12": {
        "AJ Affleck": 429.71,
        "Alec Atkinson": 4369.77,
        "Jake Gordon": 788.78,
        "Luke": 164.90,
        "Issac": -278.48,
        "Fund Mgmt": 9428.24,
        "Raj": 79.23,
        "Nairne": 79.23,
        "Phil": 79.23,
    },
}

# Per-month formula GROSS (what each person was owed by the % split).
# For Aug 2025: special waterfall (5.5% Mgmt, 9.5% Consultant, 13.5% T&D excluded,
#   0.5% × 3 principals — applied to TRUE GROSS = Fund Total Income).
# For Sep-Dec: standard 59.5% Fund Mgmt + 39% Consultant + 0.5% × 3 (applied to GP cut).
# These are the GROSS amounts before any cost netting. Net Paid = ACTUAL_PAID values.
ACTUAL_GROSS = {
    # In all months, the GROSS formula amount equals the Net actually paid for
    # contractors (because op expenses are paid out of GP retained, NOT
    # individually allocated to each consultant). The user confirmed payouts
    # are net of "weighted costs" — meaning the cost netting was already applied
    # within the Distributions ledger calc that produces the formula amounts.
    # So we present Gross == Net here. Where they differ, see the Distributions
    # Ledger original tabs (cumulative settlement balances).
    "2025-08": {
        "AJ Affleck": 335.47,
        "Alec Atkinson": 3172.12,
        "Jake Gordon": 1266.18,
        "Fund Mgmt": 2815.16,
        "Raj": 255.92,
        "Nairne": 255.92,
        "Phil": 255.92,
    },
    "2025-09": {
        "AJ Affleck": 285.77,
        "Alec Atkinson": 3561.92,
        "Jake Gordon": 930.90,
        "Fund Mgmt": 7415.62,
        "Raj": 62.32,
        "Nairne": 62.32,
        "Phil": 62.32,
    },
    "2025-10": {
        "AJ Affleck": 1268.23,
        "Alec Atkinson": 11848.88,
        "Jake Gordon": 3218.29,
        "Issac": 254.12,
        "Fund Mgmt": 31206.94,
        "Raj": 262.24,
        "Nairne": 262.24,
        "Phil": 262.24,
    },
    "2025-11": {
        "AJ Affleck": 1833.52,
        "Alec Atkinson": 15124.18,
        "Jake Gordon": 4184.51,
        "Issac": 785.85,
        "Fund Mgmt": 34183.90,
        "Raj": 287.26,
        "Nairne": 287.26,
        "Phil": 287.26,
    },
    "2025-12": {
        "AJ Affleck": 429.71,
        "Alec Atkinson": 4369.77,
        "Jake Gordon": 788.78,
        "Luke": 164.90,
        "Issac": -278.48,
        "Fund Mgmt": 9428.24,
        "Raj": 79.23,
        "Nairne": 79.23,
        "Phil": 79.23,
    },
}

# Operating expenses paid by Armada Prime Tech LLC, from each month's
# Distributions ledger Costs section + Dec from BEST ONE Costs sheet.
GP_OP_EXPENSES = {
    "2025-08": [
        ("506c SPV Loan", 4275.00),
        ("PVD", 6000.00),
    ],
    "2025-09": [],  # Sep R20 Total = 0
    "2025-10": [
        ("506c SPV Loan", 25000.00),
        ("PVD", 6000.00),
        ("Website", 7500.00),
        ("Ad Spend", 5000.00),
    ],
    "2025-11": [
        ("Chris", 7500.00),
        ("Alpha Verification", 2250.00),
        ("Formidium (TPA)", 600.00),
    ],
    "2025-12": [
        ("Chris", 4000.00),
        ("TPA (Formidium)", 600.00),
        ("Insurance", 18000.00),
        ("TPA (second line)", 4500.00),
    ],
}


# Styling
HDR_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
TOTAL_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
TOTAL_FONT = Font(bold=True, size=11)
THIN = Side(border_style="thin", color="999999")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MONEY = "$#,##0.00"
PCT = "0.00%"
INT = "#,##0"


def style_header_row(ws, row: int, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BOX


def _categorize(vendor: str) -> str:
    v = vendor.lower()
    if "tpa" in v or "formidium" in v: return "Admin / TPA Fee"
    if "chris" in v: return "Payroll/Contractor"
    if "insurance" in v: return "Insurance"
    if "spv" in v or "loan" in v: return "Loan/SPV"
    if "pvd" in v: return "PVD (verify)"
    if "website" in v: return "Marketing/Web"
    if "ad spend" in v or "ads" in v: return "Marketing/Ads"
    if "alpha" in v: return "Verification/Compliance"
    if "badtwin" in v: return "Marketing"
    return "Other"


def autosize(ws, min_w: int = 10, max_w: int = 40) -> None:
    for col in ws.columns:
        try:
            col_letter = get_column_letter(col[0].column)
        except (AttributeError, TypeError):
            continue
        width = min_w
        for cell in col:
            try:
                if cell.value is not None:
                    width = max(width, min(max_w, len(str(cell.value)) + 2))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = width


# ---------------------------------------------------------------------------
# Aggregate
# ---------------------------------------------------------------------------

def aggregate() -> dict:
    """Read each TPA file, attribute perf_fees per consultant, aggregate."""
    ids = load_ids(INTERNAL_IDS)
    # Apply 2025-specific overrides for investors that lack TPA IDs in IDS
    for tpa_id, cons in EXTRA_2025_OVERRIDES.items():
        ids[tpa_id] = {"name": None, "consultant": cons, "position_id": None}

    # Per-period results
    by_month = {}
    # Per-consultant year totals (39% pool)
    consultant_year = defaultdict(lambda: {
        "total_perf_fee_pool": 0.0,  # sum of investor perf_fees attributed to consultant
        "consultant_cut": 0.0,        # 39% of above
        "investors": defaultdict(lambda: {"perf_fee": 0.0, "consultant_cut": 0.0, "name": ""}),
    })
    unmapped = []  # investors not in IDS

    for period, path in TPA_FILES:
        if not path.exists():
            print(f"WARNING: {path} not found, skipping {period}", file=sys.stderr)
            continue
        rec = parse_workbook(str(path))
        total_perf_fee = 0.0
        consultant_pool = defaultdict(float)  # consultant -> sum of perf_fees attributed
        for inv in rec["investors"]:
            tpa_id = inv.get("investor_no")
            perf_fee = inv.get("perf_fee", 0) or 0
            total_perf_fee += perf_fee
            mapping = ids.get(tpa_id)
            if not mapping:
                unmapped.append({"period": period, "tpa_id": tpa_id, "name": inv.get("name"), "perf_fee": perf_fee})
                continue
            default_consultant = mapping.get("consultant")
            if not default_consultant:
                unmapped.append({"period": period, "tpa_id": tpa_id, "name": inv.get("name"), "perf_fee": perf_fee})
                continue
            # NOTE: TruQuant-tagged investors (e.g. Asif Moeez) are kept in the
            # consultant pool here so no perf-fee $ leaks. The user should decide
            # at tax-prep time whether TruQuant gets a 1099 or whether those $
            # should be reclassed as GP retained income.
            # Apply per-investor consultant split (e.g. Craig Levinson 50/50 AJ/Raj-Split)
            shares = resolve_consultant_split(
                tpa_id,
                inv.get("ending_balance", 0),
                period,
                default_consultant,
            )
            for cons_name, share in shares:
                share_perf = perf_fee * share
                consultant_pool[cons_name] += share_perf
                consultant_year[cons_name]["total_perf_fee_pool"] += share_perf
                consultant_year[cons_name]["consultant_cut"] += share_perf * SPLIT_PCTS_2025["consultant"]
                inv_entry = consultant_year[cons_name]["investors"][tpa_id]
                inv_entry["perf_fee"] += share_perf
                inv_entry["consultant_cut"] += share_perf * SPLIT_PCTS_2025["consultant"]
                inv_entry["name"] = inv.get("name") or mapping.get("name") or tpa_id

        # Period roll-up
        by_month[period] = {
            "label": PERIOD_LABELS[period],
            "fund_gross_income": rec["income_statement"]["total_income"],
            "fund_total_expense": -rec["income_statement"]["total_expense"],  # positive number
            "fund_net_income": rec["income_statement"]["net_income"],
            "perf_fees_crystallized": total_perf_fee,
            "fund_mgmt": total_perf_fee * SPLIT_PCTS_2025["fund_mgmt"],
            "consultant_pool_total": total_perf_fee * SPLIT_PCTS_2025["consultant"],
            "raj": total_perf_fee * SPLIT_PCTS_2025["raj"],
            "nairne": total_perf_fee * SPLIT_PCTS_2025["nairne"],
            "phil": total_perf_fee * SPLIT_PCTS_2025["phil"],
            "consultant_pool_by_consultant": dict(consultant_pool),
            "investor_count": len(rec["investors"]),
            "as_of": rec["as_of"],
            "operating_expenses": rec.get("operating_expenses", {}),
        }

    return {
        "by_month": by_month,
        "consultant_year": dict(consultant_year),
        "unmapped": unmapped,
        "ids": ids,
    }


# ---------------------------------------------------------------------------
# Workbook
# ---------------------------------------------------------------------------

def _build_month_tab(wb, period: str, by_month: dict, op_expense_year_total: float,
                     total_perf_fees_year: float) -> None:
    """Create a 'MMM 2025 Detail' tab showing capital in, gross/net per recipient,
    op expenses, and reconciliation for a single month."""
    label = PERIOD_LABELS[period]
    short = label.replace(" 2025", "")  # "Aug", "Sep", etc.
    ws = wb.create_sheet(f"{short} 2025 Detail")

    m = by_month.get(period, {})
    perf_fees = m.get("perf_fees_crystallized", 0)
    fund_gross = m.get("fund_gross_income", 0)
    paid = ACTUAL_PAID.get(period, {})
    gross = ACTUAL_GROSS.get(period, {})
    op_items = GP_OP_EXPENSES.get(period, [])
    op_total = sum(a for _, a in op_items)

    # Header
    ws["A1"] = f"{label} — GP P&L Detail (Armada Prime Tech LLC)"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:E1")

    is_aug = period == "2025-08"
    if is_aug:
        ws["A2"] = "AUGUST SPECIAL WATERFALL: 9.5% Consultant + 13.5% Trader & Developer (TQ-excluded) + 5.5% Mgmt + 0.5%×3 principals = 30% of true gross. TruQuant moved upstream from Sep onwards."
        ws["A2"].font = Font(italic=True, color="C00000")
    else:
        ws["A2"] = "Standard waterfall: 59.5% Fund Mgmt (Nairne K-1) + 39% Consultant + 0.5%×3 principals = 30% of fund net."
        ws["A2"].font = Font(italic=True, color="666666")
    ws.merge_cells("A2:E2")

    r = 4
    ws.cell(row=r, column=1, value="CAPITAL IN").font = Font(bold=True, size=12)
    r += 1
    ws.cell(row=r, column=1, value="  Fund Total Income (Armada Prime LLP, 82% post-TQ)")
    ws.cell(row=r, column=2, value=fund_gross).number_format = MONEY
    r += 1
    ws.cell(row=r, column=1, value="  GP Cut: Performance Fees Crystallized (TPA — 30% of fund net)")
    ws.cell(row=r, column=2, value=perf_fees).number_format = MONEY
    ws.cell(row=r, column=2).font = Font(bold=True)

    # K-1 partners section
    r += 2
    ws.cell(row=r, column=1, value="PAID OUT — K-1 PARTNERS (Nairne 60% + Raj 0.5%)").font = Font(bold=True, size=12)
    r += 1
    ws.cell(row=r, column=1, value="Recipient")
    ws.cell(row=r, column=2, value="Tax Type")
    ws.cell(row=r, column=3, value="Gross Slice")
    ws.cell(row=r, column=4, value="Net Paid")
    ws.cell(row=r, column=5, value="Notes")
    style_header_row(ws, r, 5)

    k1_gross_total = 0.0
    k1_net_total = 0.0
    fm_label = "Fund Mgmt 5.5% slice" if is_aug else "Fund Mgmt 59.5% slice"
    k1_items = [
        ("Nairne — " + fm_label, "Fund Mgmt", "K-1", "Nairne's share of GP cut"),
        ("Nairne — direct 0.5%", "Nairne", "K-1", "Direct member slice"),
        ("Raj Duggal — direct 0.5%", "Raj", "K-1", "Direct member slice"),
    ]
    for display, key, ttype, note in k1_items:
        r += 1
        g = gross.get(key, 0)
        n = paid.get(key, 0)
        ws.cell(row=r, column=1, value=display)
        ws.cell(row=r, column=2, value=ttype)
        ws.cell(row=r, column=3, value=g).number_format = MONEY
        ws.cell(row=r, column=4, value=n).number_format = MONEY
        ws.cell(row=r, column=5, value=note)
        for c in range(1, 6):
            ws.cell(row=r, column=c).fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
        k1_gross_total += g
        k1_net_total += n
    r += 1
    ws.cell(row=r, column=1, value="K-1 Subtotal")
    ws.cell(row=r, column=3, value=k1_gross_total).number_format = MONEY
    ws.cell(row=r, column=4, value=k1_net_total).number_format = MONEY
    for c in range(1, 6):
        ws.cell(row=r, column=c).fill = TOTAL_FILL
        ws.cell(row=r, column=c).font = Font(bold=True)

    # 1099 contractors section
    r += 2
    ws.cell(row=r, column=1, value="PAID OUT — 1099 CONTRACTORS").font = Font(bold=True, size=12)
    r += 1
    ws.cell(row=r, column=1, value="Recipient")
    ws.cell(row=r, column=2, value="Tax Type")
    ws.cell(row=r, column=3, value="Gross Slice")
    ws.cell(row=r, column=4, value="Net Paid")
    ws.cell(row=r, column=5, value="Notes")
    style_header_row(ws, r, 5)

    contractor_keys = sorted(
        [k for k in (set(gross.keys()) | set(paid.keys())) if k not in K1_RECIPIENTS],
        key=lambda k: -paid.get(k, gross.get(k, 0))
    )
    contractor_notes = {
        "Phil": "0.5% direct slice (1099, not K-1)",
        "Alec Atkinson": f"39% × his investors' perf fees" if not is_aug else "9.5% × his investors' gross profit",
        "Jake Gordon": f"39% × his investors' perf fees" if not is_aug else "9.5% × his investors' gross profit",
        "AJ Affleck": f"39% × her investors' perf fees" if not is_aug else "9.5% × her investors' gross profit",
        "Issac": "39% × his investors' perf fees",
        "Luke": "39% × his investors' perf fees",
    }
    c1099_gross_total = 0.0
    c1099_net_total = 0.0
    for key in contractor_keys:
        r += 1
        g = gross.get(key, 0)
        n = paid.get(key, 0)
        ws.cell(row=r, column=1, value=key)
        ws.cell(row=r, column=2, value="1099")
        ws.cell(row=r, column=3, value=g).number_format = MONEY
        ws.cell(row=r, column=4, value=n).number_format = MONEY
        ws.cell(row=r, column=5, value=contractor_notes.get(key, ""))
        c1099_gross_total += g
        c1099_net_total += n
        if n < 0:
            ws.cell(row=r, column=4).font = Font(bold=True, color="C00000")
    r += 1
    ws.cell(row=r, column=1, value="1099 Subtotal")
    ws.cell(row=r, column=3, value=c1099_gross_total).number_format = MONEY
    ws.cell(row=r, column=4, value=c1099_net_total).number_format = MONEY
    for c in range(1, 6):
        ws.cell(row=r, column=c).fill = TOTAL_FILL
        ws.cell(row=r, column=c).font = Font(bold=True)

    # Total cash to people
    r += 2
    ws.cell(row=r, column=1, value="TOTAL CASH PAID TO PEOPLE (K-1 + 1099)").font = Font(bold=True)
    cash_total = k1_net_total + c1099_net_total
    ws.cell(row=r, column=4, value=cash_total).number_format = MONEY
    ws.cell(row=r, column=4).font = Font(bold=True)

    # Op expenses
    r += 2
    ws.cell(row=r, column=1, value="OPERATING EXPENSES (paid by GP entity to vendors)").font = Font(bold=True, size=12)
    r += 1
    ws.cell(row=r, column=1, value="Vendor")
    ws.cell(row=r, column=2, value="Category")
    ws.cell(row=r, column=3, value="Amount")
    ws.cell(row=r, column=4, value="")
    ws.cell(row=r, column=5, value="Notes")
    style_header_row(ws, r, 5)
    if op_items:
        for vendor, amount in op_items:
            r += 1
            ws.cell(row=r, column=1, value=vendor)
            ws.cell(row=r, column=2, value=_categorize(vendor))
            ws.cell(row=r, column=3, value=amount).number_format = MONEY
            note = ""
            if "SPV" in vendor or "Loan" in vendor:
                note = "RECLASS: likely balance sheet, not P&L"
            elif "Insurance" in vendor and amount > 5000:
                note = "RECLASS: likely annual D&O, pro-rate"
            elif "TPA" in vendor or "Formidium" in vendor:
                note = "VERIFY GP-paid vs fund-paid"
            ws.cell(row=r, column=5, value=note)
    else:
        r += 1
        ws.cell(row=r, column=1, value="(no GP-paid op expenses recorded for this month)")
        ws.cell(row=r, column=1).font = Font(italic=True, color="999999")
    r += 1
    ws.cell(row=r, column=1, value="Op Expenses Subtotal")
    ws.cell(row=r, column=3, value=op_total).number_format = MONEY
    for c in range(1, 6):
        ws.cell(row=r, column=c).fill = TOTAL_FILL
        ws.cell(row=r, column=c).font = Font(bold=True)

    # Reconciliation section
    r += 2
    ws.cell(row=r, column=1, value="RECONCILIATION").font = Font(bold=True, size=12)
    r += 1
    ws.cell(row=r, column=1, value="Capital In (TPA Perf Fees Crystallized)")
    ws.cell(row=r, column=4, value=perf_fees).number_format = MONEY
    r += 1
    ws.cell(row=r, column=1, value="Less: Cash to people (K-1 + 1099)")
    ws.cell(row=r, column=4, value=-cash_total).number_format = MONEY
    r += 1
    ws.cell(row=r, column=1, value="Less: Op expenses to vendors")
    ws.cell(row=r, column=4, value=-op_total).number_format = MONEY
    r += 1
    residual = perf_fees - cash_total - op_total
    ws.cell(row=r, column=1, value="= Residual / (Deficit) for the month").font = Font(bold=True)
    ws.cell(row=r, column=4, value=residual).number_format = MONEY
    ws.cell(row=r, column=4).font = Font(bold=True, color="C00000" if residual < 0 else "006400")
    if abs(perf_fees - cash_total - op_total) > 1:
        for c in range(1, 6):
            ws.cell(row=r, column=c).fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    if is_aug:
        r += 2
        ws.cell(row=r, column=1, value="AUG NOTES:").font = Font(bold=True, color="C00000")
        r += 1
        notes = [
            "TruQuant payments excluded: $6,909.93 'Trader & Developer' (13.5%) + $88.78 Spydr — total $6,998.71. These are upstream of the GP entity per Nairne 2026-04-30.",
            "If you ADD BACK the $6,998.71 TQ-excluded amounts, total cash flow ties to the full $15,355.51 GP cut: $8,356.69 (people) + $6,998.71 (TQ) = $15,355.40.",
            "Op expenses ($10,275 SPV+PVD) are funded out of GP retained earnings, NOT from people's slices.",
        ]
        for note in notes:
            ws.cell(row=r, column=1, value=note).alignment = Alignment(wrap_text=True, vertical="top")
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
            ws.row_dimensions[r].height = 30
            r += 1

    # Column widths
    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 50


def build_workbook(agg: dict) -> None:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    by_month = agg["by_month"]
    cons_year = agg["consultant_year"]
    unmapped = agg["unmapped"]

    # Year totals (TPA-derived = what was OWED / accrual basis)
    total_perf_fees = sum(m["perf_fees_crystallized"] for m in by_month.values())
    total_fund_gross = sum(m["fund_gross_income"] for m in by_month.values())
    total_fund_net = sum(m["fund_net_income"] for m in by_month.values())
    total_fund_mgmt = total_perf_fees * SPLIT_PCTS_2025["fund_mgmt"]
    total_consultant_pool = total_perf_fees * SPLIT_PCTS_2025["consultant"]
    total_raj = total_perf_fees * SPLIT_PCTS_2025["raj"]
    total_nairne = total_perf_fees * SPLIT_PCTS_2025["nairne"]
    total_phil = total_perf_fees * SPLIT_PCTS_2025["phil"]

    # ACTUAL paid year totals (from Distributions ledger = cash basis)
    actual_year_totals = defaultdict(float)
    actual_month_totals = {}
    for period, payouts in ACTUAL_PAID.items():
        actual_month_totals[period] = sum(payouts.values())
        for recipient, amount in payouts.items():
            actual_year_totals[recipient] += amount

    # GP operating expense year total
    op_expense_year_total = sum(
        amount for items in GP_OP_EXPENSES.values() for _, amount in items
    )

    # ---------------- Tab 1: Summary ----------------
    ws = wb.create_sheet("Summary")
    ws["A1"] = "Armada Prime Tech LLC — 2025 Year-End P&L"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:D1")
    ws["A2"] = "Period: Aug 1, 2025 – Dec 31, 2025 (entity formed at Armada Prime relaunch)"
    ws["A2"].font = Font(italic=True, color="666666")
    ws.merge_cells("A2:D2")
    ws["A3"] = "Sources: TPA Reporting Packages (gross income); Distributions Armada Tech 2025 ledger (actual paid); BEST ONE Dec 2025 Monthly Return (Dec breakdown)"
    ws["A3"].font = Font(italic=True, color="666666")
    ws.merge_cells("A3:D3")
    ws["A4"] = "TruQuant payments excluded entirely (per Nairne 2026-04-30) — TQ is not a GP expense / 1099 recipient. Aug TQ-tagged ~$6,999 stays in retained earnings → flows to K-1 net income."
    ws["A4"].font = Font(italic=True, color="C00000")
    ws.merge_cells("A4:D4")
    ws.row_dimensions[4].height = 30

    r = 5
    ws.cell(row=r, column=1, value="Line Item").font = Font(bold=True)
    ws.cell(row=r, column=2, value="Amount").font = Font(bold=True)
    ws.cell(row=r, column=3, value="Note").font = Font(bold=True)
    style_header_row(ws, r, 3)

    actual_total_paid = sum(actual_year_totals.values())
    # K-1 partner allocations (Nairne owns 60% = Fund Mgmt 59.5 + direct 0.5; Raj owns 0.5%)
    nairne_total = sum(actual_year_totals.get(a, 0.0) for a in NAIRNE_ALIASES)
    actual_raj = actual_year_totals.get("Raj", 0.0)
    actual_phil = actual_year_totals.get("Phil", 0.0)
    # Contractor 1099 expenses = everyone except K-1 partners
    actual_consultant_total = sum(
        v for k, v in actual_year_totals.items()
        if k not in K1_RECIPIENTS and k != "Phil"
    )
    actual_1099_total = actual_consultant_total + actual_phil
    # Partnership net income = gross - 1099 expenses - op expenses
    # Partner allocations are NOT expenses; they're how net income is divided.
    net_income_actual = total_perf_fees - actual_1099_total - op_expense_year_total

    rows = [
        ("GROSS INCOME", None, ""),
        ("  Performance Fees from Armada Prime LLP", total_perf_fees, "TPA Performance Fees Crystallized, Aug-Dec 2025"),
        ("  TOTAL GROSS INCOME", total_perf_fees, "= Armada Prime Tech LLC's gross 2025 receipts"),
        ("", None, ""),
        ("LESS: 1099 EXPENSES (contractor payments — partners excluded)", None, ""),
        ("  Consultant pool (Alec, Jake, AJ, Issac, Luke)", actual_consultant_total, "See 1099 Summary tab for per-consultant breakdown"),
        ("  Phil (0.5% GP fixed slice)", actual_phil, "Phil held the 0.5% slice in 2025 as a 1099 contractor"),
        ("  TOTAL 1099 EXPENSES", actual_1099_total, ""),
        ("", None, ""),
        ("LESS: GP-PAID OPERATING EXPENSES", None, ""),
        ("  Vendor expenses (Chris, Insurance, PVD, Website, etc.)", op_expense_year_total, "See GP Expenses tab. Includes items likely needing reclass (SPV loans → balance sheet, Insurance pro-rata)."),
        ("", None, ""),
        ("PARTNERSHIP NET INCOME (allocated to K-1 partners)", net_income_actual, "Gross - 1099 expenses - Op expenses. This is what flows to K-1s."),
        ("", None, ""),
        ("K-1 ALLOCATION (per ownership: Nairne 60% / Raj 0.5%)", None, ""),
        ("  Nairne — Fund Mgmt 59.5% slice received as cash", actual_year_totals.get("Fund Mgmt", 0), "Reported as K-1 income (partner allocation, not expense)"),
        ("  Nairne — direct 0.5% slice received as cash", actual_year_totals.get("Nairne", 0), "Reported as K-1 income"),
        ("  Nairne — TOTAL cash received (60% of perf fees)", nairne_total, "= Fund Mgmt + direct slice"),
        ("  Raj — direct 0.5% slice received as cash", actual_raj, "Reported as K-1 income"),
        ("", None, ""),
        ("(K-1 net income vs cash distributions reconcile on accountant's K-1 forms)", None, "Cash received above is what each partner got operationally. The K-1 'net income' line above is the partnership's taxable income — accountant allocates it per the operating agreement."),
    ]
    for label, val, note in rows:
        r += 1
        ws.cell(row=r, column=1, value=label)
        if val is not None:
            cell = ws.cell(row=r, column=2, value=val)
            cell.number_format = MONEY
        ws.cell(row=r, column=3, value=note)
        if label.startswith("  TOTAL") or label.startswith("NET INCOME") or label.startswith("GROSS INCOME") or label.startswith("LESS:") or label.startswith("K-1 ALLOCATION"):
            for c in range(1, 4):
                ws.cell(row=r, column=c).font = Font(bold=True)
            if label.startswith("  TOTAL") or label.startswith("NET INCOME"):
                for c in range(1, 4):
                    ws.cell(row=r, column=c).fill = TOTAL_FILL

    # Cross-check (cash basis sum should approximately equal gross perf fees)
    r += 2
    ws.cell(row=r, column=1, value="CROSS-CHECK").font = Font(italic=True, bold=True)
    r += 1
    cash_total = actual_1099_total + nairne_total + actual_raj
    ws.cell(row=r, column=1, value=f"Total cash out (1099s ${actual_1099_total:,.2f} + Nairne K-1 ${nairne_total:,.2f} + Raj K-1 ${actual_raj:,.2f})")
    ws.cell(row=r, column=2, value=cash_total).number_format = MONEY
    ws.cell(row=r, column=3, value=f"vs Gross ${total_perf_fees:,.2f} — delta ${total_perf_fees - cash_total:,.2f} (= retained / cash-vs-accrual / TQ exclusion)").font = Font(italic=True)
    r += 1
    ws.cell(row=r, column=1, value="Operating expenses + retained should equal the delta above (with cash-vs-accrual smoothing).").font = Font(italic=True, color="666666")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)

    autosize(ws, max_w=70)
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 70

    # ---------------- Tab 2: 1099 Summary ----------------
    ws = wb.create_sheet("1099 Summary")
    ws["A1"] = "Armada Prime Tech LLC — 2025 1099-NEC Recipients"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:F1")
    ws["A2"] = "All amounts cover Aug 1, 2025 – Dec 31, 2025. Tax-classification: per Nairne 2026-04-30 — Raj/Nairne are K-1 members; everyone else is 1099."
    ws["A2"].font = Font(italic=True, color="666666")
    ws.merge_cells("A2:F2")

    r = 4
    headers = ["Recipient", "2025 Total ($)", "Type", "Source of $", "EIN/SSN (fill in)", "Address (fill in)"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=r, column=i, value=h)
    style_header_row(ws, r, len(headers))

    # Build 1099 list: contractor recipients only (Fund Mgmt is K-1 to Nairne, NOT a 1099)
    one099_rows = []
    one099_rows.append(("Phil (last name TBD)", actual_year_totals.get("Phil", 0), "Individual 1099", "0.5% GP fixed slice (held by Phil all 2025)"))
    # Consultants (excludes K-1 partners and Phil who's already added)
    consultant_names = [k for k in actual_year_totals if k not in K1_RECIPIENTS and k != "Phil"]
    for cons_name in sorted(consultant_names, key=lambda k: -actual_year_totals[k]):
        one099_rows.append((cons_name, actual_year_totals[cons_name], "Individual 1099", "From Distributions ledger (cash actually paid)"))

    total_1099 = 0.0
    for label, amount, ttype, src in one099_rows:
        r += 1
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=amount).number_format = MONEY
        ws.cell(row=r, column=3, value=ttype)
        ws.cell(row=r, column=4, value=src)
        ws.cell(row=r, column=5, value="")
        ws.cell(row=r, column=6, value="")
        total_1099 += amount

    r += 1
    ws.cell(row=r, column=1, value="TOTAL 1099 PAYMENTS")
    ws.cell(row=r, column=2, value=total_1099).number_format = MONEY
    for c in range(1, 7):
        ws.cell(row=r, column=c).fill = TOTAL_FILL
        ws.cell(row=r, column=c).font = TOTAL_FONT

    r += 2
    raj_actual = actual_year_totals.get("Raj", 0)
    nairne_total_check = sum(actual_year_totals.get(a, 0) for a in NAIRNE_ALIASES)
    ws.cell(row=r, column=1, value=f"K-1 partner cash distributions (NOT on 1099s): Nairne ${nairne_total_check:,.2f} + Raj ${raj_actual:,.2f} = ${nairne_total_check + raj_actual:,.2f}").font = Font(italic=True)
    grand = total_1099 + nairne_total_check + raj_actual
    ws.cell(row=r+1, column=1, value=f"Cross-check: 1099 total (${total_1099:,.2f}) + K-1 cash (${nairne_total_check + raj_actual:,.2f}) = ${grand:,.2f}; Gross perf fees per TPA = ${total_perf_fees:,.2f}; delta = ${total_perf_fees - grand:,.2f} (cash-vs-accrual timing + Aug TQ-T&D excluded)").font = Font(italic=True)

    autosize(ws)
    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["D"].width = 50
    ws.column_dimensions["F"].width = 30

    # ---------------- Tab 3: K-1 Summary ----------------
    ws = wb.create_sheet("K-1 Summary")
    ws["A1"] = "Armada Prime Tech LLC — 2025 K-1 Members"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:E1")
    ws["A2"] = "Members per Nairne 2026-04-30: Nairne (60% — Fund Mgmt 59.5% + direct 0.5%) and Raj Duggal (0.5%). Net income allocated per the operating agreement."
    ws["A2"].font = Font(italic=True, color="666666")
    ws.merge_cells("A2:E2")

    r = 4
    headers = ["Member", "Ownership %", "Cash Distributions ($)", "Allocated Share of Net Income ($)", "SSN/EIN + Address (fill in)"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=r, column=i, value=h)
    style_header_row(ws, r, len(headers))

    raj_actual = actual_year_totals.get("Raj", 0)
    nairne_total = sum(actual_year_totals.get(a, 0) for a in NAIRNE_ALIASES)
    contractor_1099 = sum(v for k, v in actual_year_totals.items() if k not in K1_RECIPIENTS)
    # Partnership net income (what gets allocated on K-1 schedule K)
    net_income = total_perf_fees - contractor_1099 - op_expense_year_total
    # Per ownership: Nairne 60/60.5 = 99.17%, Raj 0.5/60.5 = 0.83% (placeholder; real op agreement governs)
    nairne_ownership = 60.0 / 60.5
    raj_ownership = 0.5 / 60.5

    for member, owner_pct, cash in [
        ("Nairne", nairne_ownership, nairne_total),
        ("Raj Duggal", raj_ownership, raj_actual),
    ]:
        r += 1
        ws.cell(row=r, column=1, value=member)
        ws.cell(row=r, column=2, value=f"{owner_pct*100:.2f}%")
        ws.cell(row=r, column=3, value=cash).number_format = MONEY
        allocated = net_income * owner_pct
        ws.cell(row=r, column=4, value=allocated).number_format = MONEY
        ws.cell(row=r, column=5, value="")
        if allocated < 0:
            ws.cell(row=r, column=4).font = Font(bold=True, color="C00000")

    r += 2
    ws.cell(row=r, column=1, value="K-1 ARITHMETIC").font = Font(bold=True)
    r += 1
    arith_lines = [
        f"GP Gross Income (TPA Perf Fees Crystallized): ${total_perf_fees:,.2f}",
        f"Less: 1099 contractor expenses (Alec/Jake/AJ/Phil/Issac/Luke): ${contractor_1099:,.2f}",
        f"Less: GP-paid operating expenses: ${op_expense_year_total:,.2f}",
        f"= Partnership NET INCOME (allocated to K-1s): ${net_income:,.2f}",
        f"Nairne allocated share ({nairne_ownership*100:.2f}%): ${net_income * nairne_ownership:,.2f}",
        f"Raj allocated share ({raj_ownership*100:.2f}%): ${net_income * raj_ownership:,.2f}",
    ]
    for line in arith_lines:
        ws.cell(row=r, column=1, value=line)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="NOTES").font = Font(bold=True, color="C00000")
    notes = [
        "Fund Mgmt 59.5% slice is Nairne's K-1 income (per Nairne 2026-04-30 — it's not a separate entity 1099).",
        "Cash distributions and allocated K-1 net income are different things in partnership tax — Schedule K-1 reports both. The accountant will compute capital account changes.",
        "Operating expenses include items that may be reclassed by the accountant (506c SPV Loans → balance sheet; Insurance $18k may need annual proration).",
        "Cash-basis 1099s (above) may differ from accrual-basis (TPA-derived). Confirm GP's tax basis with accountant.",
        "TruQuant payments are excluded per Nairne 2026-04-30 (TQ is not a GP expense).",
        "Ownership % shown (Nairne 99.17%/Raj 0.83%) is derived from the 60/0.5 split. The actual LLC operating agreement governs — confirm with the accountant before filing.",
    ]
    r += 1
    for note in notes:
        ws.cell(row=r, column=1, value=note)
        ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        ws.row_dimensions[r].height = 30
        r += 1

    autosize(ws)
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 35

    # ---------------- Tab 4: Monthly Detail ----------------
    ws = wb.create_sheet("Monthly Detail")
    ws["A1"] = "Per-Month TPA Roll-Up (Aug-Dec 2025)"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:I1")

    r = 3
    headers = ["Month", "Investor Count", "Fund Gross Income (82% cut)", "GP Perf Fees Crystallized", "Fund Mgmt 59.5%", "Consultant Pool 39%", "Phil 0.5%", "Raj 0.5%", "Nairne 0.5%"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=r, column=i, value=h)
    style_header_row(ws, r, len(headers))

    for period in ["2025-08", "2025-09", "2025-10", "2025-11", "2025-12"]:
        if period not in by_month:
            continue
        m = by_month[period]
        r += 1
        ws.cell(row=r, column=1, value=m["label"])
        ws.cell(row=r, column=2, value=m["investor_count"]).number_format = INT
        ws.cell(row=r, column=3, value=m["fund_gross_income"]).number_format = MONEY
        ws.cell(row=r, column=4, value=m["perf_fees_crystallized"]).number_format = MONEY
        ws.cell(row=r, column=5, value=m["fund_mgmt"]).number_format = MONEY
        ws.cell(row=r, column=6, value=m["consultant_pool_total"]).number_format = MONEY
        ws.cell(row=r, column=7, value=m["phil"]).number_format = MONEY
        ws.cell(row=r, column=8, value=m["raj"]).number_format = MONEY
        ws.cell(row=r, column=9, value=m["nairne"]).number_format = MONEY

    r += 1
    ws.cell(row=r, column=1, value="2025 TOTAL")
    ws.cell(row=r, column=3, value=total_fund_gross).number_format = MONEY
    ws.cell(row=r, column=4, value=total_perf_fees).number_format = MONEY
    ws.cell(row=r, column=5, value=total_fund_mgmt).number_format = MONEY
    ws.cell(row=r, column=6, value=total_consultant_pool).number_format = MONEY
    ws.cell(row=r, column=7, value=total_phil).number_format = MONEY
    ws.cell(row=r, column=8, value=total_raj).number_format = MONEY
    ws.cell(row=r, column=9, value=total_nairne).number_format = MONEY
    for c in range(1, 10):
        ws.cell(row=r, column=c).fill = TOTAL_FILL
        ws.cell(row=r, column=c).font = TOTAL_FONT

    autosize(ws)

    # ---------------- Tab 5: Per-Consultant Monthly ----------------
    ws = wb.create_sheet("Per-Consultant Monthly")
    ws["A1"] = "Per-Consultant 39% Pool Allocation by Month"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:G1")
    ws["A2"] = "Each cell = consultant_cut (39% × sum of their investors' perf_fees that month). Used for 1099 cross-check vs per-consultant tracker xlsx files."
    ws["A2"].font = Font(italic=True, color="666666")
    ws.merge_cells("A2:G2")

    r = 4
    months = ["2025-08", "2025-09", "2025-10", "2025-11", "2025-12"]
    headers = ["Consultant"] + [PERIOD_LABELS[p] for p in months] + ["2025 Total"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=r, column=i, value=h)
    style_header_row(ws, r, len(headers))

    consultants_sorted = sorted(cons_year.items(), key=lambda x: -x[1]["consultant_cut"])
    col_totals = [0.0] * len(months)
    for cons_name, data in consultants_sorted:
        r += 1
        ws.cell(row=r, column=1, value=cons_name)
        row_total = 0.0
        for j, period in enumerate(months):
            m = by_month.get(period, {})
            pool_amt = m.get("consultant_pool_by_consultant", {}).get(cons_name, 0.0)
            cut = pool_amt * SPLIT_PCTS_2025["consultant"]
            cell = ws.cell(row=r, column=2 + j, value=cut)
            cell.number_format = MONEY
            row_total += cut
            col_totals[j] += cut
        ws.cell(row=r, column=2 + len(months), value=row_total).number_format = MONEY
        ws.cell(row=r, column=2 + len(months)).font = Font(bold=True)

    r += 1
    ws.cell(row=r, column=1, value="TOTAL")
    grand = 0.0
    for j, t in enumerate(col_totals):
        ws.cell(row=r, column=2 + j, value=t).number_format = MONEY
        grand += t
    ws.cell(row=r, column=2 + len(months), value=grand).number_format = MONEY
    for c in range(1, 2 + len(months) + 1):
        ws.cell(row=r, column=c).fill = TOTAL_FILL
        ws.cell(row=r, column=c).font = TOTAL_FONT

    autosize(ws)
    ws.column_dimensions["A"].width = 28

    # ---------------- Tab 6: GP Expenses ----------------
    ws = wb.create_sheet("GP Expenses")
    ws["A1"] = "GP-Paid Operating Expenses (Aug-Dec 2025)"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:E1")
    ws["A2"] = "Only expenses paid by Armada Prime Tech LLC. Source: 'Costs' section of each month's Distributions Armada Tech 2025 ledger; Dec from BEST ONE Costs sheet."
    ws["A2"].font = Font(italic=True, color="666666")
    ws.merge_cells("A2:E2")

    r = 4
    headers = ["Month", "Vendor / Line Item", "Amount", "Category", "Source"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=r, column=i, value=h)
    style_header_row(ws, r, len(headers))

    # All expenses from GP_OP_EXPENSES dict
    grand_total = 0.0
    for period, items in GP_OP_EXPENSES.items():
        if not items:
            r += 1
            ws.cell(row=r, column=1, value=PERIOD_LABELS[period])
            ws.cell(row=r, column=2, value="(no GP-paid op expenses recorded)")
            ws.cell(row=r, column=2).font = Font(italic=True, color="999999")
            continue
        period_total = 0.0
        for vendor, amount in items:
            r += 1
            ws.cell(row=r, column=1, value=PERIOD_LABELS[period])
            ws.cell(row=r, column=2, value=vendor)
            ws.cell(row=r, column=3, value=amount).number_format = MONEY
            ws.cell(row=r, column=4, value=_categorize(vendor))
            ws.cell(row=r, column=5, value=("BEST ONE Dec Costs" if period == "2025-12" else f"Distributions ledger — {PERIOD_LABELS[period]} sheet"))
            period_total += amount
            grand_total += amount
        r += 1
        ws.cell(row=r, column=1, value=f"{PERIOD_LABELS[period]} subtotal")
        ws.cell(row=r, column=3, value=period_total).number_format = MONEY
        for c in range(1, 6):
            ws.cell(row=r, column=c).fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
            ws.cell(row=r, column=c).font = Font(bold=True)

    r += 2
    ws.cell(row=r, column=1, value="2025 TOTAL GP-PAID OPERATING EXPENSES")
    ws.cell(row=r, column=3, value=grand_total).number_format = MONEY
    for c in range(1, 6):
        ws.cell(row=r, column=c).fill = TOTAL_FILL
        ws.cell(row=r, column=c).font = TOTAL_FONT

    r += 2
    ws.cell(row=r, column=1, value="Verify with user before sending to accountant").font = Font(bold=True, color="C00000")
    r += 1
    notes = [
        "1. Insurance ($18k Dec): is this annual or already monthly-allocated? If annual, reduce to ~$1.5k for Dec.",
        "2. TPA fee duplication: Dec has TWO TPA lines ($600 + $4,500). The $600 also appears as fund-level admin in TPA's books. Confirm only the GP-paid portion goes here.",
        "3. 506c SPV Loan ($4,275 Aug + $25,000 Oct): is this an expense or a loan repayment? Expense for 1099/K-1 vs balance-sheet item.",
        "4. PVD ($6k Aug + $6k Oct): vendor identification needed for 1099 (if individual) or just expense (if entity).",
        "5. Badtwin / Ad Spend / Website (Oct): vendor names + 1099 obligations need confirmation.",
    ]
    for note in notes:
        ws.cell(row=r, column=1, value=note)
        ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        r += 1

    autosize(ws)
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 50

    # ---------------- Tab 6b: Distributions Ledger (Actual Paid) ----------------
    ws = wb.create_sheet("Distributions Ledger")
    ws["A1"] = "Per-Recipient ACTUAL Paid Per Month (Distributions Armada Tech 2025 Ledger)"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:H1")
    ws["A2"] = "Cash-basis amounts the GP actually paid out. Aug uses a different waterfall (9.5% consultant + 13.5% Trader & Developer + 5.5% Mgmt) — TruQuant moved upstream from Sep onwards."
    ws["A2"].font = Font(italic=True, color="666666")
    ws.merge_cells("A2:H2")

    r = 4
    months = ["2025-08", "2025-09", "2025-10", "2025-11", "2025-12"]
    headers = ["Recipient"] + [PERIOD_LABELS[p] for p in months] + ["2025 Total"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=r, column=i, value=h)
    style_header_row(ws, r, len(headers))

    # All recipients sorted by year total — annotate K-1 partners
    all_recipients = sorted(actual_year_totals.keys(), key=lambda k: -actual_year_totals[k])
    col_totals = [0.0] * len(months)
    for recipient in all_recipients:
        r += 1
        label = recipient
        if recipient == "Fund Mgmt":
            label = "Fund Mgmt 59.5% (= Nairne K-1)"
        elif recipient in K1_RECIPIENTS:
            label = f"{recipient} (K-1 partner)"
        ws.cell(row=r, column=1, value=label)
        row_total = 0.0
        for j, period in enumerate(months):
            val = ACTUAL_PAID.get(period, {}).get(recipient, 0.0)
            cell = ws.cell(row=r, column=2 + j, value=val)
            cell.number_format = MONEY
            row_total += val
            col_totals[j] += val
        cell = ws.cell(row=r, column=2 + len(months), value=row_total)
        cell.number_format = MONEY
        cell.font = Font(bold=True)
        if recipient in K1_RECIPIENTS:
            for c in range(1, 2 + len(months) + 1):
                ws.cell(row=r, column=c).fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")

    r += 1
    ws.cell(row=r, column=1, value="MONTH TOTAL (cash actually paid out)")
    for j, t in enumerate(col_totals):
        ws.cell(row=r, column=2 + j, value=t).number_format = MONEY
    ws.cell(row=r, column=2 + len(months), value=sum(col_totals)).number_format = MONEY
    for c in range(1, 2 + len(months) + 1):
        ws.cell(row=r, column=c).fill = TOTAL_FILL
        ws.cell(row=r, column=c).font = TOTAL_FONT

    r += 1
    ws.cell(row=r, column=1, value="TPA Perf Fees Crystallized (what was OWED)")
    for j, period in enumerate(months):
        val = by_month.get(period, {}).get("perf_fees_crystallized", 0)
        ws.cell(row=r, column=2 + j, value=val).number_format = MONEY
    ws.cell(row=r, column=2 + len(months), value=total_perf_fees).number_format = MONEY
    for c in range(1, 2 + len(months) + 1):
        ws.cell(row=r, column=c).font = Font(italic=True)

    r += 1
    ws.cell(row=r, column=1, value="DELTA (Owed - Paid; positive = unpaid liability)")
    for j, period in enumerate(months):
        owed = by_month.get(period, {}).get("perf_fees_crystallized", 0)
        delta = owed - col_totals[j]
        cell = ws.cell(row=r, column=2 + j, value=delta)
        cell.number_format = MONEY
        if abs(delta) > 50:
            cell.font = Font(bold=True, color="C00000")
    total_delta = total_perf_fees - sum(col_totals)
    cell = ws.cell(row=r, column=2 + len(months), value=total_delta)
    cell.number_format = MONEY
    cell.font = Font(bold=True, color="C00000" if abs(total_delta) > 50 else "000000")

    autosize(ws)
    ws.column_dimensions["A"].width = 35

    # ---------------- Tab 6c: Monthly Op Expenses (vendor × month matrix) ----------------
    ws = wb.create_sheet("Monthly Op Expenses")
    ws["A1"] = "Operating Expenses by Vendor & Month (Aug-Dec 2025)"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:H1")
    ws["A2"] = "Each vendor on its own row. Source: Costs sections of Distributions ledger (Aug-Nov) + BEST ONE Dec Costs."
    ws["A2"].font = Font(italic=True, color="666666")
    ws.merge_cells("A2:H2")

    r = 4
    months = ["2025-08", "2025-09", "2025-10", "2025-11", "2025-12"]
    headers = ["Vendor / Line Item", "Category"] + [PERIOD_LABELS[p] for p in months] + ["2025 Total", "Notes"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=r, column=i, value=h)
    style_header_row(ws, r, len(headers))

    # Build vendor × month matrix
    vendor_matrix = {}  # vendor -> {period -> amount}
    for period, items in GP_OP_EXPENSES.items():
        for vendor, amount in items:
            vendor_matrix.setdefault(vendor, {})[period] = amount

    NOTES_BY_VENDOR = {
        "506c SPV Loan": "RECLASS: likely balance sheet item (loan/capital), not P&L expense",
        "Insurance": "RECLASS: likely annual D&O — pro-rate to ~$7,500 for Aug-Dec",
        "TPA (Formidium)": "VERIFY: $600/mo also in fund-level admin (TPA books). May double-count.",
        "TPA (second line)": "VERIFY: second TPA line in Dec — separate billing or duplicate?",
        "Formidium (TPA)": "VERIFY: same as 'TPA (Formidium)' — TPA fee. Confirm GP-paid vs fund-paid.",
        "PVD": "Vendor identification needed for 1099 obligation",
        "Website": "Marketing/web build",
        "Ad Spend": "Marketing — vendor breakdown needed",
        "Chris": "Payroll/contractor — likely 1099 (already on consultant 1099 list?)",
        "Alpha Verification": "Compliance/verification service",
    }

    # Sort: keep vendors that span multiple months grouped
    vendors_sorted = sorted(vendor_matrix.keys(), key=lambda v: (-sum(vendor_matrix[v].values()), v))
    col_totals = {p: 0.0 for p in months}
    for vendor in vendors_sorted:
        r += 1
        ws.cell(row=r, column=1, value=vendor)
        ws.cell(row=r, column=2, value=_categorize(vendor))
        row_total = 0.0
        for j, period in enumerate(months):
            val = vendor_matrix[vendor].get(period, 0.0)
            cell = ws.cell(row=r, column=3 + j, value=val if val else None)
            if val:
                cell.number_format = MONEY
                row_total += val
                col_totals[period] += val
        ws.cell(row=r, column=3 + len(months), value=row_total).number_format = MONEY
        ws.cell(row=r, column=3 + len(months)).font = Font(bold=True)
        ws.cell(row=r, column=4 + len(months), value=NOTES_BY_VENDOR.get(vendor, ""))
        ws.cell(row=r, column=4 + len(months)).alignment = Alignment(wrap_text=True)

    r += 1
    ws.cell(row=r, column=1, value="MONTHLY TOTAL")
    grand_total = 0.0
    for j, period in enumerate(months):
        ws.cell(row=r, column=3 + j, value=col_totals[period]).number_format = MONEY
        grand_total += col_totals[period]
    ws.cell(row=r, column=3 + len(months), value=grand_total).number_format = MONEY
    for c in range(1, 4 + len(months) + 1):
        ws.cell(row=r, column=c).fill = TOTAL_FILL
        ws.cell(row=r, column=c).font = TOTAL_FONT

    # Adjusted-after-reclass row
    r += 2
    ws.cell(row=r, column=1, value="Likely-adjusted (after accountant reclass)").font = Font(bold=True, color="006400")
    r += 1
    ws.cell(row=r, column=1, value="  Less: 506c SPV Loans (balance sheet)")
    ws.cell(row=r, column=3 + len(months), value=-29275).number_format = MONEY
    r += 1
    ws.cell(row=r, column=1, value="  Less: Insurance proration ($18k → $7,500 for 5 mo)")
    ws.cell(row=r, column=3 + len(months), value=-10500).number_format = MONEY
    r += 1
    ws.cell(row=r, column=1, value="ADJUSTED 2025 OP EXPENSES")
    adjusted = grand_total - 29275 - 10500
    ws.cell(row=r, column=3 + len(months), value=adjusted).number_format = MONEY
    for c in range(1, 4 + len(months) + 1):
        ws.cell(row=r, column=c).fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
        ws.cell(row=r, column=c).font = Font(bold=True)

    autosize(ws)
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 22
    for j in range(len(months)):
        ws.column_dimensions[get_column_letter(3 + j)].width = 13
    ws.column_dimensions[get_column_letter(3 + len(months))].width = 14
    ws.column_dimensions[get_column_letter(4 + len(months))].width = 60

    # ---------------- Tab 6d: Monthly Per-Person (with running YTD) ----------------
    ws = wb.create_sheet("Monthly Per-Person")
    ws["A1"] = "Per-Person Cash Distributions by Month, with Running YTD"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:M1")
    ws["A2"] = "Each person: top row = monthly amount paid; bottom row = running YTD total. K-1 partners (Nairne, Raj) shaded blue. Per Nairne 2026-04-30: payouts shown ARE net (after costs + weighted costs). TruQuant excluded."
    ws["A2"].font = Font(italic=True, color="666666")
    ws.merge_cells("A2:M2")

    r = 4
    # Header: Person | Type | Aug | YTD | Sep | YTD | Oct | YTD | Nov | YTD | Dec | YTD | YE Total
    headers = ["Person / Recipient", "Tax Type"]
    for p in months:
        headers.append(PERIOD_LABELS[p])
        headers.append("YTD")
    headers.append("YE 2025 Total")
    for i, h in enumerate(headers, 1):
        ws.cell(row=r, column=i, value=h)
    style_header_row(ws, r, len(headers))

    # Recipients ordered: K-1 first, then 1099 by amount
    k1_recipients = [
        ("Nairne — Fund Mgmt 59.5% (Aug: 5.5%)", "Fund Mgmt", "K-1"),
        ("Nairne — direct 0.5%", "Nairne", "K-1"),
        ("Raj Duggal — direct 0.5%", "Raj", "K-1"),
    ]
    contractor_recipients = sorted(
        [(k, k, "1099") for k in actual_year_totals if k not in K1_RECIPIENTS],
        key=lambda x: -actual_year_totals[x[1]]
    )
    all_recipients = k1_recipients + contractor_recipients

    # Sub-totals for K-1 vs 1099
    k1_monthly_totals = {p: 0.0 for p in months}
    contractor_monthly_totals = {p: 0.0 for p in months}

    for label, key, ttype in all_recipients:
        r += 1
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=ttype)
        if ttype == "K-1":
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
        ytd = 0.0
        ye_total = 0.0
        for j, period in enumerate(months):
            val = ACTUAL_PAID.get(period, {}).get(key, 0.0)
            ytd += val
            ye_total += val
            if ttype == "K-1":
                k1_monthly_totals[period] += val
            else:
                contractor_monthly_totals[period] += val
            month_col = 3 + j * 2
            ytd_col = 4 + j * 2
            cell = ws.cell(row=r, column=month_col, value=val if val else None)
            if val:
                cell.number_format = MONEY
            cell_ytd = ws.cell(row=r, column=ytd_col, value=ytd)
            cell_ytd.number_format = MONEY
            cell_ytd.font = Font(italic=True, color="666666")
        ws.cell(row=r, column=len(headers), value=ye_total).number_format = MONEY
        ws.cell(row=r, column=len(headers)).font = Font(bold=True)

    # Sub-total row: K-1 total per month
    r += 1
    ws.cell(row=r, column=1, value="K-1 SUBTOTAL (Nairne + Raj)")
    ws.cell(row=r, column=2, value="K-1")
    k1_ytd = 0.0
    k1_year = 0.0
    for j, period in enumerate(months):
        v = k1_monthly_totals[period]
        k1_ytd += v
        k1_year += v
        ws.cell(row=r, column=3 + j*2, value=v).number_format = MONEY
        ws.cell(row=r, column=4 + j*2, value=k1_ytd).number_format = MONEY
    ws.cell(row=r, column=len(headers), value=k1_year).number_format = MONEY
    for c in range(1, len(headers) + 1):
        ws.cell(row=r, column=c).fill = TOTAL_FILL
        ws.cell(row=r, column=c).font = TOTAL_FONT

    # Sub-total row: 1099 total per month
    r += 1
    ws.cell(row=r, column=1, value="1099 SUBTOTAL (Phil + consultants)")
    ws.cell(row=r, column=2, value="1099")
    c1099_ytd = 0.0
    c1099_year = 0.0
    for j, period in enumerate(months):
        v = contractor_monthly_totals[period]
        c1099_ytd += v
        c1099_year += v
        ws.cell(row=r, column=3 + j*2, value=v).number_format = MONEY
        ws.cell(row=r, column=4 + j*2, value=c1099_ytd).number_format = MONEY
    ws.cell(row=r, column=len(headers), value=c1099_year).number_format = MONEY
    for c in range(1, len(headers) + 1):
        ws.cell(row=r, column=c).fill = TOTAL_FILL
        ws.cell(row=r, column=c).font = TOTAL_FONT

    # Grand total
    r += 1
    ws.cell(row=r, column=1, value="GRAND TOTAL (K-1 + 1099)")
    ws.cell(row=r, column=2, value="")
    grand_ytd = 0.0
    grand_year = 0.0
    for j, period in enumerate(months):
        v = k1_monthly_totals[period] + contractor_monthly_totals[period]
        grand_ytd += v
        grand_year += v
        ws.cell(row=r, column=3 + j*2, value=v).number_format = MONEY
        ws.cell(row=r, column=4 + j*2, value=grand_ytd).number_format = MONEY
    ws.cell(row=r, column=len(headers), value=grand_year).number_format = MONEY
    for c in range(1, len(headers) + 1):
        ws.cell(row=r, column=c).fill = PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")
        ws.cell(row=r, column=c).font = Font(bold=True)

    # TPA gross row for context
    r += 1
    ws.cell(row=r, column=1, value="TPA Perf Fees Crystallized (Gross GP Income)")
    for j, period in enumerate(months):
        v = by_month.get(period, {}).get("perf_fees_crystallized", 0)
        ws.cell(row=r, column=3 + j*2, value=v).number_format = MONEY
    ws.cell(row=r, column=len(headers), value=total_perf_fees).number_format = MONEY
    for c in range(1, len(headers) + 1):
        ws.cell(row=r, column=c).font = Font(italic=True)

    # Op expenses row
    r += 1
    ws.cell(row=r, column=1, value="Op Expenses (paid from GP retained)")
    for j, period in enumerate(months):
        items = GP_OP_EXPENSES.get(period, [])
        v = sum(a for _, a in items)
        ws.cell(row=r, column=3 + j*2, value=v if v else None)
        if v:
            ws.cell(row=r, column=3 + j*2).number_format = MONEY
    ws.cell(row=r, column=len(headers), value=op_expense_year_total).number_format = MONEY
    for c in range(1, len(headers) + 1):
        ws.cell(row=r, column=c).font = Font(italic=True)

    # Notes section
    r += 2
    ws.cell(row=r, column=1, value="NOTES").font = Font(bold=True, color="C00000")
    notes = [
        "Payouts shown are NET — already after weighted costs (per Nairne 2026-04-30). The Distributions ledger's separate 'Gross vs Net' table tracks cumulative settlement balances (carries forward unpaid amounts) and a layer of Coinbase/wire/Crypto fees (~2.8%) plus per-person 'Expense' allocations (Nairne Expense, Alec Expense, etc.) — those are nested inside the Net amounts you see here.",
        "TruQuant payments are excluded entirely (per Nairne 2026-04-30). The Aug TruQuant amounts ($6,909.93 Trader & Developer + $88.78 Spydr) and Sep Spydr $82.08 are NOT in any of the rows above.",
        "Aug 2025 used a different waterfall (Fund Mgmt was 5.5% not 59.5%). From Sep onwards: standard 59.5% Mgmt + 39% Consultant + 0.5%×3 principals.",
        "Issac shows -$278.48 in Dec — clawback against prior month overpayment.",
        "Op Expenses row at the bottom is the GP-paid vendor expenses (not allocated to individuals on this tab — see Monthly Op Expenses tab for vendor breakdown).",
    ]
    r += 1
    for note in notes:
        ws.cell(row=r, column=1, value=note).alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(headers))
        ws.row_dimensions[r].height = 35
        r += 1

    autosize(ws)
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 9
    for j in range(len(months)):
        ws.column_dimensions[get_column_letter(3 + j*2)].width = 12
        ws.column_dimensions[get_column_letter(4 + j*2)].width = 12
    ws.column_dimensions[get_column_letter(len(headers))].width = 14

    # ---------------- Tabs 6e-6i: Per-Month Detail tabs ----------------
    for period in ["2025-08", "2025-09", "2025-10", "2025-11", "2025-12"]:
        _build_month_tab(wb, period, by_month, op_expense_year_total, total_perf_fees)

    # ---------------- Tab 7: Reconciliation ----------------
    ws = wb.create_sheet("Reconciliation")
    ws["A1"] = "Reconciliation: TPA-derived vs Internal Sources"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:E1")

    r = 3
    headers = ["Item", "TPA Source ($)", "Internal Source ($)", "Delta ($)", "Notes"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=r, column=i, value=h)
    style_header_row(ws, r, len(headers))

    rows = [
        ("Dec 2025 Perf Fees Crystallized", by_month["2025-12"]["perf_fees_crystallized"], 15140.61, "BEST ONE Dec Consultants tab total"),
        ("Sum 1099 + 0.5% direct slices = Gross Perf", total_fund_mgmt + total_consultant_pool + total_phil + total_raj + total_nairne, total_perf_fees, "Internal arithmetic check"),
    ]
    for label, tpa, internal, note in rows:
        r += 1
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=tpa).number_format = MONEY
        ws.cell(row=r, column=3, value=internal).number_format = MONEY
        delta = (tpa or 0) - (internal or 0)
        ws.cell(row=r, column=4, value=delta).number_format = MONEY
        if abs(delta) > 1:
            ws.cell(row=r, column=4).font = Font(bold=True, color="C00000")
        ws.cell(row=r, column=5, value=note)

    r += 2
    ws.cell(row=r, column=1, value="Unmapped investors (no consultant attribution)").font = Font(bold=True)
    if unmapped:
        r += 1
        for h in ["Period", "TPA ID", "Name", "Perf Fee"]:
            ws.cell(row=r, column={"Period": 1, "TPA ID": 2, "Name": 3, "Perf Fee": 4}[h], value=h).font = Font(bold=True)
        for u in unmapped:
            r += 1
            ws.cell(row=r, column=1, value=u["period"])
            ws.cell(row=r, column=2, value=u["tpa_id"])
            ws.cell(row=r, column=3, value=u["name"])
            ws.cell(row=r, column=4, value=u["perf_fee"]).number_format = MONEY
    else:
        r += 1
        ws.cell(row=r, column=1, value="None — every investor mapped successfully.")

    autosize(ws)
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["E"].width = 45

    # Save
    wb.save(OUT_XLSX)
    print(f"Wrote {OUT_XLSX}")


# ---------------------------------------------------------------------------
# Markdown summary
# ---------------------------------------------------------------------------

def build_markdown(agg: dict) -> None:
    by_month = agg["by_month"]
    cons_year = agg["consultant_year"]
    unmapped = agg["unmapped"]

    total_perf_fees = sum(m["perf_fees_crystallized"] for m in by_month.values())
    total_fund_mgmt = total_perf_fees * SPLIT_PCTS_2025["fund_mgmt"]
    total_consultant_pool = total_perf_fees * SPLIT_PCTS_2025["consultant"]
    total_raj = total_perf_fees * SPLIT_PCTS_2025["raj"]
    total_nairne = total_perf_fees * SPLIT_PCTS_2025["nairne"]
    total_phil = total_perf_fees * SPLIT_PCTS_2025["phil"]

    # Actual paid (cash basis from Distributions ledger)
    actual_year_totals = defaultdict(float)
    for period, payouts in ACTUAL_PAID.items():
        for recipient, amount in payouts.items():
            actual_year_totals[recipient] += amount
    actual_total = sum(actual_year_totals.values())
    op_expense_year = sum(amount for items in GP_OP_EXPENSES.values() for _, amount in items)

    lines = []
    lines.append("# Armada Prime Tech LLC — 2025 Year-End Reconciliation")
    lines.append("")
    lines.append(f"**Tax year:** 2025  \n**Period covered:** Aug 1, 2025 – Dec 31, 2025 (entity formed at Armada Prime relaunch)  \n**Source of truth:** TPA Reporting Packages (Performance Fees Crystallized line)")
    lines.append("")
    nairne_total = sum(actual_year_totals.get(a, 0) for a in NAIRNE_ALIASES)
    raj_actual = actual_year_totals.get("Raj", 0)
    contractor_1099 = sum(v for k, v in actual_year_totals.items() if k not in K1_RECIPIENTS)
    net_income = total_perf_fees - contractor_1099 - op_expense_year
    nairne_ownership = 60.0 / 60.5
    raj_ownership = 0.5 / 60.5

    lines.append("## Bottom Line")
    lines.append("")
    lines.append(f"- **Gross income (perf fees from Armada Prime LLP):** **${total_perf_fees:,.2f}**")
    lines.append(f"- **1099 contractor expenses (cash paid per Distributions Ledger):** **${contractor_1099:,.2f}**")
    consultant_breakdown = sorted(((k, v) for k, v in actual_year_totals.items() if k not in K1_RECIPIENTS and k != 'Phil'), key=lambda x: -x[1])
    for name, amt in consultant_breakdown:
        lines.append(f"  - {name}: ${amt:,.2f}")
    lines.append(f"  - Phil (0.5% GP fixed): ${actual_year_totals.get('Phil', 0):,.2f}")
    lines.append(f"- **GP-paid operating expenses (Aug-Dec 2025):** ${op_expense_year:,.2f} *(see GP Expenses tab; verify Insurance + TPA double-count + SPV loans reclass)*")
    lines.append(f"- **Partnership net income (allocated to K-1 partners):** **${net_income:,.2f}**")
    lines.append("")
    lines.append("**K-1 partner allocations** (Nairne 60% ownership = Fund Mgmt 59.5% + direct 0.5%; Raj 0.5% ownership):")
    lines.append("")
    lines.append(f"- **Nairne** ({nairne_ownership*100:.2f}% ownership):")
    lines.append(f"  - Cash distributions: ${nairne_total:,.2f} (Fund Mgmt ${actual_year_totals.get('Fund Mgmt', 0):,.2f} + direct 0.5% ${actual_year_totals.get('Nairne', 0):,.2f})")
    lines.append(f"  - Allocated share of net income: ${net_income * nairne_ownership:,.2f}")
    lines.append(f"- **Raj Duggal** ({raj_ownership*100:.2f}% ownership):")
    lines.append(f"  - Cash distributions: ${raj_actual:,.2f}")
    lines.append(f"  - Allocated share of net income: ${net_income * raj_ownership:,.2f}")
    lines.append("")
    lines.append("## Monthly Roll-Up")
    lines.append("")
    lines.append("| Month | Investors | Fund Gross Income | GP Perf Fees Crystallized | Fund Mgmt 59.5% | Consultant Pool 39% | Phil 0.5% | Raj 0.5% | Nairne 0.5% |")
    lines.append("|---|---:|---:|---:|---:|---:|---:|---:|---:|")
    for period in ["2025-08", "2025-09", "2025-10", "2025-11", "2025-12"]:
        if period not in by_month:
            continue
        m = by_month[period]
        lines.append(f"| {m['label']} | {m['investor_count']} | ${m['fund_gross_income']:,.2f} | ${m['perf_fees_crystallized']:,.2f} | ${m['fund_mgmt']:,.2f} | ${m['consultant_pool_total']:,.2f} | ${m['phil']:,.2f} | ${m['raj']:,.2f} | ${m['nairne']:,.2f} |")
    lines.append(f"| **2025 TOTAL** | — | **${sum(m['fund_gross_income'] for m in by_month.values()):,.2f}** | **${total_perf_fees:,.2f}** | **${total_fund_mgmt:,.2f}** | **${total_consultant_pool:,.2f}** | **${total_phil:,.2f}** | **${total_raj:,.2f}** | **${total_nairne:,.2f}** |")
    lines.append("")
    lines.append("## Per-Recipient ACTUAL Paid (Distributions Ledger, Cash Basis)")
    lines.append("")
    lines.append("These are the amounts actually disbursed per the internal `Distributions Armada Tech 2025` ledger. **Use these for cash-basis 1099s.** For accrual basis, see the TPA-derived per-consultant breakdown in the workbook's `Per-Consultant Monthly` tab.")
    lines.append("")
    lines.append("| Recipient | Aug | Sep | Oct | Nov | Dec | 2025 Total |")
    lines.append("|---|---:|---:|---:|---:|---:|---:|")
    months = ["2025-08", "2025-09", "2025-10", "2025-11", "2025-12"]
    for recipient in sorted(actual_year_totals.keys(), key=lambda k: -actual_year_totals[k]):
        cells = [f"${ACTUAL_PAID.get(p, {}).get(recipient, 0):,.2f}" for p in months]
        lines.append(f"| {recipient} | {' | '.join(cells)} | **${actual_year_totals[recipient]:,.2f}** |")
    month_totals = [sum(ACTUAL_PAID.get(p, {}).values()) for p in months]
    lines.append(f"| **MONTH TOTAL** | {' | '.join(f'${t:,.2f}' for t in month_totals)} | **${sum(month_totals):,.2f}** |")
    tpa_cells = [f"${by_month[p]['perf_fees_crystallized']:,.2f}" for p in months]
    lines.append(f"| TPA Perf Fees (Owed) | {' | '.join(tpa_cells)} | ${total_perf_fees:,.2f} |")
    lines.append("")
    lines.append("### August Anomaly — DIFFERENT WATERFALL")
    lines.append("")
    lines.append("In August 2025, the GP cut (30% of fund net) was split as: 9.5% Consultant + 13.5% Trader & Developer (TruQuant) + 5.5% Mgmt + 0.5% × 3 (Raj/Nairne/Phil) = 30% of true gross. From September onwards, TruQuant's 18% moved upstream of the GP entity, and the GP cut adopted the standard 59.5/39/0.5/0.5/0.5 split. **TruQuant payments are excluded from this reconciliation entirely** (per Nairne 2026-04-30) — TQ is not a GP expense / 1099 recipient. The Aug TQ-tagged amounts (~$6,999) effectively flow to GP retained income → K-1 net income.")
    lines.append("")
    lines.append("## GP-Paid Operating Expenses (Aug-Dec 2025)")
    lines.append("")
    lines.append("| Month | Total GP-Paid Op Expenses |")
    lines.append("|---|---:|")
    for period in months:
        items = GP_OP_EXPENSES.get(period, [])
        total = sum(a for _, a in items)
        lines.append(f"| {PERIOD_LABELS[period]} | ${total:,.2f} |")
    lines.append(f"| **2025 Total** | **${op_expense_year:,.2f}** |")
    lines.append("")
    lines.append("Itemized breakdown in the workbook's `GP Expenses` and `Monthly Op Expenses` tabs (vendor × month matrix in the latter).")
    lines.append("")
    lines.append("## On Gross vs Net (Distributions Ledger Mechanics)")
    lines.append("")
    lines.append("Per Nairne 2026-04-30: the per-person amounts in this reconciliation are **already NET** — after weighted costs and per-person expense allocations. The `Monthly Per-Person` tab in the workbook shows monthly + running YTD per recipient.")
    lines.append("")
    lines.append("The Distributions Armada Tech 2025 ledger has a separate 'Gross vs Net' table inside each month's sheet that tracks **cumulative settlement balances** — i.e., it carries forward unpaid balances from prior months. That's why for example Raj's 'Gross' column in the Nov sheet shows $632.74 (vs his $287.26 monthly slice): it's settling outstanding amounts. Inside that calculation are also Coinbase/Wire/Crypto fees (~2.8%) and per-person 'Expense' allocations (Nairne Expense $1,848.20 in Nov, Alec Expense $1,828.31, etc.) which are nested INSIDE the Net amounts shown in this reconciliation.")
    lines.append("")
    lines.append("In short: don't try to back into a 'Gross' from this workbook — the formula amount per person ≈ what they were paid (the Net column), and the cost/fee layers are already absorbed.")
    lines.append("")
    lines.append("## Open Items Before Sending to Accountant")
    lines.append("")
    lines.append("1. **LLC operating agreement / capital accounts** — confirm with accountant the formal ownership % and net-income allocation for K-1s. The 60/0.5 derived from the Distributions ledger is operational, not legal.")
    lines.append("2. **Phil's last name + SSN/address** — for his $946.97 1099.")
    lines.append("3. **Operating expense reclassification** — review the GP Expenses tab. Likely needs reclass:")
    lines.append("   - **506c SPV Loans** ($4,275 Aug + $25,000 Oct = $29,275): these may be loans/capital, not expenses.")
    lines.append("   - **Insurance $18k Dec**: confirm whether annual or already pro-rated.")
    lines.append("   - **TPA fees** (Dec $600 + $4,500 = $5,100; Nov $600): confirm GP-paid vs fund-paid (fund books already have $600/mo admin).")
    lines.append("   - **PVD, Badtwin, Ad Spend, Website**: vendor names + 1099 obligations need confirmation.")
    lines.append("4. **Per-recipient SSN/address** — required for all 1099 forms (placeholder cells in workbook).")
    lines.append(f"5. **Cash-vs-accrual basis** — Total cash paid (${sum(actual_year_totals.values()):,.2f}) vs Total owed per TPA (${total_perf_fees:,.2f}) = ${total_perf_fees - sum(actual_year_totals.values()):,.2f} difference. The delta includes ~$6,999 of August TruQuant-tagged amounts that are excluded per the 'TQ is not a GP expense' policy. Confirm GP's tax basis (cash vs accrual) with accountant.")
    lines.append("")
    if unmapped:
        lines.append("## Unmapped Investors (excluded from 1099 attribution)")
        lines.append("")
        lines.append("| Period | TPA ID | Name | Perf Fee |")
        lines.append("|---|---|---|---:|")
        for u in unmapped:
            lines.append(f"| {u['period']} | {u['tpa_id']} | {u['name']} | ${u['perf_fee']:,.2f} |")
        lines.append("")
    lines.append("## Methodology")
    lines.append("")
    lines.append("- **Source of GP gross income:** TPA Reporting Package's `Performance Fees Crystallized` (per investor, in the Investor Capital Summary). Per the 2026-04-27 decision, TPA — not the internal Monthly Return — is authoritative for GP/consultant compensation.")
    lines.append("- **Investor → consultant mapping:** the IDS sheet from `BEST ONE of December 2025 Monthly Return.xlsx` (most complete 2025 mapping), augmented with the standard `CONSULTANT_OVERRIDES` from `tools/build_consultant_splits.py`.")
    lines.append("- **GP economic split (2025):** **Nairne 60%** (= Fund Mgmt 59.5% + direct 0.5%) / **Raj 0.5%** / Consultant 39% / **Phil 0.5%**. Per Nairne 2026-04-30: Fund Mgmt is Nairne's K-1 income, not a separate entity 1099. Nairne and Raj are K-1 partners; Phil and the consultants are 1099 contractors. Phil held the 0.5% slice all of 2025; Alec replaced him in April 2026.")
    lines.append("- **TruQuant:** From September 2025 onwards, TruQuant takes 18% of true gross **upstream** of Armada Prime LLP and it never enters the GP's books. **In August 2025, however**, TruQuant was paid INSIDE the GP entity as 'Trader & Developer' (13.5%) and 'Spydr' (a small slice of the consultant pool) — these Aug payments DO appear in the 1099 list.")
    lines.append("- **Tax classification (per Nairne 2026-04-30):** Raj and Nairne are LLC members → K-1. Everyone else who received GP-pool payouts is a 1099 contractor.")
    lines.append("")
    lines.append("## Reproduce")
    lines.append("")
    lines.append("```")
    lines.append("python tools/build_2025_year_end.py")
    lines.append("```")
    lines.append("")
    lines.append("Outputs `2025-armada-prime-tech-1099-k1.xlsx` and this file.")

    OUT_MD.write_text("\n".join(lines))
    print(f"Wrote {OUT_MD}")


if __name__ == "__main__":
    agg = aggregate()
    build_workbook(agg)
    build_markdown(agg)
    print("Done.")
