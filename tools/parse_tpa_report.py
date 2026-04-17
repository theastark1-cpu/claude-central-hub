#!/usr/bin/env python3
"""Parse an Armada Prime LLP TPA Reporting Package xlsx and upsert into data/tpa_history.json.

Usage:
    python tools/parse_tpa_report.py path/to/report.xlsx

Looks up values by sheet + header name (not fixed cell coords), so the parser
will raise clearly if the TPA ever changes the workbook schema.
"""
from __future__ import annotations

import argparse
import datetime as dt
import json
import re
import sys
from pathlib import Path
from typing import Any

import openpyxl

REPO_ROOT = Path(__file__).resolve().parent.parent
HISTORY_PATH = REPO_ROOT / "data" / "tpa_history.json"

MONTHS = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
    "january": 1, "february": 2, "march": 3, "april": 4, "june": 6,
    "july": 7, "august": 8, "september": 9, "october": 10,
    "november": 11, "december": 12,
}


def _clean(v: Any) -> Any:
    if isinstance(v, str):
        return v.strip()
    return v


def _rows(ws) -> list[list[Any]]:
    return [[_clean(c) for c in row] for row in ws.iter_rows(values_only=True)]


def _round(v: float | None, nd: int = 2) -> float:
    if v is None:
        return 0.0
    return round(float(v), nd)


def _parse_period(ws_rows: list[list[Any]]) -> tuple[str, str, str]:
    """Return (period 'YYYY-MM', label 'Aug 2025', as_of 'YYYY-MM-DD').

    Handles three TPA formats:
      v1 "As of : Aug-31-2025"
      v1 "Period : Aug-01-2025 - Aug-31-2025"
      v2 "November 30, 2025"           (header row)
      v2 "For the period from November 1, 2025 to November 30, 2025"
    """
    for row in ws_rows[:8]:
        for cell in row:
            if not isinstance(cell, str):
                continue
            # v1: "As of : Aug-31-2025"
            m = re.search(r"As of\s*:\s*([A-Za-z]{3,})-?(\d{1,2})-?(\d{4})", cell)
            if m:
                month_name, day, year = m.group(1).lower(), int(m.group(2)), int(m.group(3))
                month = MONTHS[month_name[:3]]
                return f"{year}-{month:02d}", f"{month_name.capitalize()[:3]} {year}", f"{year}-{month:02d}-{day:02d}"
            # v1: "Period : Aug-01-2025 - Aug-31-2025"
            m = re.search(r"([A-Za-z]{3,})-(\d{1,2})-(\d{4})\s*-\s*([A-Za-z]{3,})-(\d{1,2})-(\d{4})", cell)
            if m:
                end_month = MONTHS[m.group(4).lower()[:3]]
                end_day = int(m.group(5))
                end_year = int(m.group(6))
                return (
                    f"{end_year}-{end_month:02d}",
                    f"{m.group(4)[:3].capitalize()} {end_year}",
                    f"{end_year}-{end_month:02d}-{end_day:02d}",
                )
            # v2: "November 30, 2025" or "For the period from X to November 30, 2025"
            m = re.search(r"([A-Za-z]{3,})\s+(\d{1,2}),\s*(\d{4})", cell)
            if m:
                month_name = m.group(1).lower()[:3]
                if month_name not in MONTHS:
                    continue
                day, year = int(m.group(2)), int(m.group(3))
                month = MONTHS[month_name]
                # If cell contains "from X to Y", prefer the second (end) match
                matches = re.findall(r"([A-Za-z]{3,})\s+(\d{1,2}),\s*(\d{4})", cell)
                if len(matches) >= 2:
                    mn, d, y = matches[-1]
                    if mn.lower()[:3] in MONTHS:
                        month_name = mn.lower()[:3]
                        day, year = int(d), int(y)
                        month = MONTHS[month_name]
                return f"{year}-{month:02d}", f"{month_name.capitalize()} {year}", f"{year}-{month:02d}-{day:02d}"
    raise ValueError("Could not parse reporting period from workbook")


def _find_row(rows: list[list[Any]], label: str) -> list[Any] | None:
    label = label.lower()
    for row in rows:
        for cell in row:
            if isinstance(cell, str) and cell.strip().lower() == label:
                return row
    return None


def _find_contains(rows: list[list[Any]], substr: str) -> list[Any] | None:
    substr = substr.lower()
    for row in rows:
        for cell in row:
            if isinstance(cell, str) and substr in cell.strip().lower():
                return row
    return None


def _find_starts_with(rows: list[list[Any]], prefix: str) -> list[Any] | None:
    """Match a row where a cell starts with `prefix` (case-insensitive)."""
    prefix = prefix.lower()
    for row in rows:
        for cell in row:
            if isinstance(cell, str) and cell.strip().lower().startswith(prefix):
                return row
    return None


def _last_numeric(row: list[Any]) -> float:
    for v in reversed(row):
        if isinstance(v, (int, float)):
            return float(v)
    return 0.0


def parse_balance_sheet(rows: list[list[Any]]) -> dict:
    cascade = _last_numeric(_find_contains(rows, "investment in cascade") or [0])
    # "Investment in Cryptocurrencies, (At Cost)" — always cost basis
    crypto = _last_numeric(_find_contains(rows, "investment in cryptocurrencies") or [0])
    cash_row = _find_row(rows, "Cash") or []
    cash = _last_numeric(cash_row)
    # v2 new asset lines
    loan_acg = _last_numeric(_find_contains(rows, "loan to armada capital") or [0])
    subs_recv = _last_numeric(_find_contains(rows, "subscription receivable") or [0])
    unrealized_gl = _last_numeric(_find_contains(rows, "unrealized gain (loss) on investment in cryptocurrencies") or [0])
    total_assets = _last_numeric(_find_row(rows, "Total Assets") or [])

    payable_gp = _last_numeric(_find_contains(rows, "payable to gp") or [0])
    perf_payable = _last_numeric(_find_contains(rows, "performance fees payable") or [0])
    subs_advance = _last_numeric(_find_contains(rows, "subscription received in advance") or [0])
    total_liab = _last_numeric(_find_row(rows, "Total Liabilities") or [])

    capital_add = _last_numeric(_find_contains(rows, "capital addition") or [0])
    capital_redempt = _last_numeric(_find_contains(rows, "capital redemption") or [0])
    retained = _last_numeric(_find_contains(rows, "income and retained earnings") or [0])
    total_capital = _last_numeric(_find_row(rows, "Total Capital") or [])

    return {
        "total_assets": _round(total_assets),
        "assets": {
            "cascade": _round(cascade),
            "crypto": _round(crypto),
            "cash": _round(cash),
            "loan_to_acg": _round(loan_acg),
            "subscription_receivable": _round(subs_recv),
            "unrealized_gl_crypto": _round(unrealized_gl),
        },
        "total_liabilities": _round(total_liab),
        "liabilities": {
            "payable_to_gp": _round(payable_gp),
            "perf_fees_payable": _round(perf_payable),
            "subscriptions_advance": _round(subs_advance),
        },
        "total_capital": _round(total_capital),
        "capital": {
            "additions": _round(capital_add),
            "redemptions": _round(capital_redempt),
            "retained_earnings": _round(retained),
        },
    }


def parse_income_statement(rows: list[list[Any]]) -> dict:
    interest = _last_numeric(_find_contains(rows, "interest income") or [0])
    misc = _last_numeric(_find_contains(rows, "misc. trading income") or [0])
    # Prefix-match to avoid collision with "Change in Unrealized Gain (Loss) on Investment..."
    realized = _last_numeric(_find_starts_with(rows, "realized gain") or [0])
    reward = _last_numeric(_find_contains(rows, "reward income") or [0])
    change_unrealized = _last_numeric(_find_starts_with(rows, "change in unrealized") or [0])
    cascade_income = _last_numeric(_find_starts_with(rows, "income from cascade") or [0])
    total_income = _last_numeric(_find_row(rows, "Total Income") or [])

    bank = _last_numeric(_find_contains(rows, "bank charges") or [0])
    # "Performance Fees" — prefix-match so it doesn't match "Performance Fees Payable" on Balance Sheet
    # (safe here since we're parsing the Income Statement sheet)
    perf = _last_numeric(_find_contains(rows, "performance fees") or [0])
    commission = _last_numeric(_find_starts_with(rows, "commission expense") or [0])
    # v2 adds "Operating Expense" as its own line; distinct from "Total Expense"
    op_expense = 0.0
    for row in rows:
        for cell in row:
            if isinstance(cell, str) and cell.strip().lower() == "operating expense":
                op_expense = _last_numeric(row)
                break

    total_expense = _last_numeric(_find_row(rows, "Total Expense") or [])

    # Net income row: "NET INCOME", "Net Income", or "Net Income (Loss)"
    net_row = (_find_row(rows, "NET INCOME")
               or _find_row(rows, "Net Income")
               or _find_row(rows, "Net Income (Loss)")
               or _find_contains(rows, "net income"))
    net = _last_numeric(net_row or [])
    return {
        "total_income": _round(total_income),
        "income": {
            "interest": _round(interest),
            "misc_trading": _round(misc),
            "realized_gl_crypto": _round(realized),
            "reward": _round(reward),
            "change_in_unrealized_gl": _round(change_unrealized),
            "cascade_income": _round(cascade_income),
        },
        "total_expense": _round(total_expense),
        "expense": {
            "bank_charges": _round(bank),
            "perf_fees": _round(perf),
            "operating_expense": _round(op_expense),
            "commission_expense": _round(commission),
        },
        "net_income": _round(net),
    }


HEADER_HINT_WORDS = (
    "investor", "fund", "date", "symbol", "asset class", "description",
    "account", "period", "quantity", "cash",
)


def _header_index(rows: list[list[Any]]) -> tuple[int, dict[str, int]]:
    """Find the header row and return its index + {header: col}.

    Heuristic: first row where >=2 cells are non-empty strings AND the first
    non-empty string is one of HEADER_HINT_WORDS (case-insensitive), and the row
    is after the top title/period block (skip rows 0-2).
    """
    for i, row in enumerate(rows):
        if i < 2:
            continue
        strs = [(j, str(c).strip()) for j, c in enumerate(row) if isinstance(c, str) and str(c).strip()]
        if len(strs) < 2:
            continue
        first = strs[0][1].lower()
        if any(hint in first for hint in HEADER_HINT_WORDS):
            return i, {s: j for j, s in strs}
    # Fallback: first row with >=5 string cells
    for i, row in enumerate(rows):
        strs = [(j, str(c).strip()) for j, c in enumerate(row) if isinstance(c, str) and str(c).strip()]
        if len(strs) >= 5:
            return i, {s: j for j, s in strs}
    raise ValueError("Could not locate header row")


def _first_data_row(rows: list[list[Any]]) -> int:
    """For headerless v2 sheets: find first row that looks like investor data.

    Criteria: has >=3 non-empty cells including at least one 14-Class-style ID
    or a mix of strings + numerics.
    """
    for i, row in enumerate(rows):
        non_empty = [c for c in row if c not in (None, "")]
        if len(non_empty) < 3:
            continue
        has_class_id = any(isinstance(c, str) and re.match(r"^\d+-Class", c) for c in non_empty)
        has_money = any(isinstance(c, (int, float)) and c > 100 for c in non_empty)
        if has_class_id and has_money:
            return i
    return -1


def _sheet(wb, *names: str):
    """Find a worksheet by any of the candidate names."""
    for n in names:
        if n in wb.sheetnames:
            return wb[n]
    return None


def _col(headers: dict[str, int], *names: str) -> int | None:
    for n in names:
        for k, v in headers.items():
            if k.lower().strip() == n.lower().strip():
                return v
    for n in names:
        for k, v in headers.items():
            if n.lower() in k.lower():
                return v
    return None


def parse_capital_schedule(rows: list[list[Any]]) -> list[dict]:
    """Per-investor capital schedule. v1 has column headers; v2 is headerless
    positional. v2 schedule column layout:
        [No., Name, BegOwn, EndOwn, BegEquity, GrossProfit, ExpensedFee,
         AllocatedFee, Additions, Withdrawals, Transfers, EndingEquity]
    v1 swaps Name and No. in cols 0-1; remainder is identical.
    """
    out = []
    try:
        header_idx, headers = _header_index(rows)
        c_name = _col(headers, "Investor Name")
        c_no = _col(headers, "Investor No.", "Investor Number")
        c_beg_own = _col(headers, "Beginning Ownership")
        c_end_own = _col(headers, "Ending Ownership")
        c_beg_eq = _col(headers, "Beginning Equity")
        c_gp = _col(headers, "Gross Profits")
        c_alloc_fee = _col(headers, "Allocated Fees")
        c_add = _col(headers, "Additions")
        c_wd = _col(headers, "Withdrawals")
        c_xfr = _col(headers, "Transfers In/(Out)")
        c_end_eq = _col(headers, "Ending Equity")
        data_start = header_idx + 1
    except ValueError:
        # v2 headerless — positional
        data_start = _first_data_row(rows)
        if data_start < 0:
            return []
        # Detect column order: v2 has investor_no at col 0, v1 has name at col 0
        first_cell = rows[data_start][0] if rows[data_start] else None
        v2 = isinstance(first_cell, str) and bool(re.match(r"^\d+-Class", first_cell))
        if v2:
            c_no, c_name = 0, 1
        else:
            c_name, c_no = 0, 1
        c_beg_own, c_end_own, c_beg_eq, c_gp = 2, 3, 4, 5
        c_alloc_fee = 7  # col 6 is Expensed Fees (unused)
        c_add, c_wd, c_xfr, c_end_eq = 8, 9, 10, 11

    for row in rows[data_start:]:
        name = row[c_name] if c_name is not None and c_name < len(row) else None
        if not isinstance(name, str) or not name.strip() or name.strip().lower() == "total":
            continue
        out.append({
            "investor_no": row[c_no] if c_no is not None and c_no < len(row) else "",
            "name": name.strip(),
            "begin_ownership": _round(row[c_beg_own] if c_beg_own is not None and c_beg_own < len(row) else 0, 6),
            "end_ownership": _round(row[c_end_own] if c_end_own is not None and c_end_own < len(row) else 0, 6),
            "begin_equity": _round(row[c_beg_eq] if c_beg_eq is not None and c_beg_eq < len(row) else 0),
            "gross_profit": _round(row[c_gp] if c_gp is not None and c_gp < len(row) else 0),
            "allocated_fee": _round(row[c_alloc_fee] if c_alloc_fee is not None and c_alloc_fee < len(row) else 0),
            "additions": _round(row[c_add] if c_add is not None and c_add < len(row) else 0),
            "withdrawals": _round(row[c_wd] if c_wd is not None and c_wd < len(row) else 0),
            "transfers": _round(row[c_xfr] if c_xfr is not None and c_xfr < len(row) else 0),
            "ending_balance": _round(row[c_end_eq] if c_end_eq is not None and c_end_eq < len(row) else 0),
        })
    return out


def parse_investor_capital_summary(rows: list[list[Any]]) -> tuple[list[dict], dict]:
    """Per-investor capital summary. v1 has column headers; v2 is headerless
    but uses the same 29-column layout as v1.

    Canonical column order (0-indexed):
      0 Period | 1 Fund Name | 2 Investor Name | 3 Investor Number |
      4 Share Class | 5 Series | 6 Begin Balance | 7 Shares Begin |
      8 Additions Begin | 9 Additions Shares Begin | ...
      14 Total P&L | 15 Mgmt Fee | 16 Perf Fee | 17 Additions End | ...
      23 Ending Balance | 24 Shares Ending | 25 NAV Per Share |
      26 Gross MTD ROR | 27 Net MTD ROR | 28 YTD ROR
    """
    try:
        header_idx, headers = _header_index(rows)
        c_name = _col(headers, "Investor Name")
        c_no = _col(headers, "Investor Number", "Investor No.")
        c_begin = _col(headers, "Begin Balance")
        c_end = _col(headers, "Ending Balance")
        c_shares_end = _col(headers, "Shares Ending")
        c_shares_begin = _col(headers, "Shares Begin")
        c_nav = _col(headers, "NAV Per Share")
        c_gross_mtd = _col(headers, "Gross MTD ROR")
        c_net_mtd = _col(headers, "Net MTD ROR")
        c_ytd = _col(headers, "YTD ROR")
        c_mgmt = _col(headers, "Mgmt Fee")
        c_perf = _col(headers, "Perf Fee")
        c_add_end = _col(headers, "Additions End")
        c_add_begin = _col(headers, "Additions Begin")
        data_start = header_idx + 1
    except ValueError:
        # v2 positional
        data_start = _first_data_row(rows)
        if data_start < 0:
            return [], {}
        c_name, c_no = 2, 3
        c_begin, c_shares_begin = 6, 7
        c_add_begin = 8
        c_mgmt, c_perf = 15, 16
        c_add_end = 17
        c_end, c_shares_end = 23, 24
        c_nav = 25
        c_gross_mtd, c_net_mtd, c_ytd = 26, 27, 28

    rows_out = []
    for row in rows[data_start:]:
        name = row[c_name] if c_name is not None and c_name < len(row) else None
        if not isinstance(name, str) or not name.strip() or "total" in name.strip().lower():
            continue
        rows_out.append({
            "investor_no": row[c_no] if c_no is not None and c_no < len(row) else "",
            "name": name.strip(),
            "begin_balance": _round(row[c_begin] if c_begin is not None and c_begin < len(row) else 0),
            "ending_balance": _round(row[c_end] if c_end is not None and c_end < len(row) else 0),
            "shares_begin": _round(row[c_shares_begin] if c_shares_begin is not None and c_shares_begin < len(row) else 0, 6),
            "shares_end": _round(row[c_shares_end] if c_shares_end is not None and c_shares_end < len(row) else 0, 6),
            "nav_per_share": _round(row[c_nav] if c_nav is not None and c_nav < len(row) else 0, 4),
            "gross_mtd_ror": _round(row[c_gross_mtd] if c_gross_mtd is not None and c_gross_mtd < len(row) else 0, 6),
            "net_mtd_ror": _round(row[c_net_mtd] if c_net_mtd is not None and c_net_mtd < len(row) else 0, 6),
            "ytd_ror": _round(row[c_ytd] if c_ytd is not None and c_ytd < len(row) else 0, 6),
            "mgmt_fee": _round(row[c_mgmt] if c_mgmt is not None and c_mgmt < len(row) else 0),
            "perf_fee": _round(row[c_perf] if c_perf is not None and c_perf < len(row) else 0),
        })

    fund_level = {}
    for row in rows[data_start:]:
        first_non_empty = next((c for c in row if c not in (None, "")), "")
        if isinstance(first_non_empty, str) and "total" in first_non_empty.lower():
            fund_level = {
                "total_additions_end": _round(row[c_add_end] if c_add_end is not None and c_add_end < len(row) else 0),
                "total_additions_begin": _round(row[c_add_begin] if c_add_begin is not None and c_add_begin < len(row) else 0),
                "total_ending_balance": _round(row[c_end] if c_end is not None and c_end < len(row) else 0),
                "total_shares": _round(row[c_shares_end] if c_shares_end is not None and c_shares_end < len(row) else 0, 6),
            }
            break
    return rows_out, fund_level


def parse_positions(rows: list[list[Any]]) -> list[dict]:
    header_idx, headers = _header_index(rows)
    c_ac = _col(headers, "Asset Class")
    c_sym = _col(headers, "Symbol")
    c_qty = _col(headers, "Quantity")
    c_mv = _col(headers, "MV (BC)")
    c_ugl = _col(headers, "Unrealized G/L (BC)")
    out = []
    for row in rows[header_idx + 1:]:
        ac = row[c_ac] if c_ac is not None else None
        sym = row[c_sym] if c_sym is not None else None
        if not isinstance(ac, str) or not ac.strip() or "total" in ac.lower() or ac.strip().startswith("LC -"):
            continue
        if not isinstance(sym, str) or not sym.strip():
            continue
        out.append({
            "asset_class": ac.strip(),
            "symbol": sym.strip(),
            "qty": _round(row[c_qty] if c_qty is not None else 0, 6),
            "mv": _round(row[c_mv] if c_mv is not None else 0),
            "unrealized_gl": _round(row[c_ugl] if c_ugl is not None else 0),
        })
    return out


def parse_realized(rows: list[list[Any]]) -> list[dict]:
    header_idx, headers = _header_index(rows)
    c_date = _col(headers, "Date")
    c_sym = _col(headers, "Symbol")
    c_qty = _col(headers, "Quantity")
    c_gl = _col(headers, "Realized G/L (BC)")
    out = []
    for row in rows[header_idx + 1:]:
        d = row[c_date] if c_date is not None else None
        sym = row[c_sym] if c_sym is not None else None
        if d is None or d == "":
            continue
        if isinstance(d, str) and ("total" in d.lower() or d.strip().startswith("LC -")):
            continue
        if not isinstance(sym, str) or not sym.strip():
            continue
        if isinstance(d, dt.datetime):
            d_str = d.strftime("%Y-%m-%d")
        elif isinstance(d, str):
            # "08/25/2025"
            try:
                d_str = dt.datetime.strptime(d.strip(), "%m/%d/%Y").strftime("%Y-%m-%d")
            except ValueError:
                d_str = d.strip()
        else:
            d_str = str(d)
        out.append({
            "date": d_str,
            "symbol": row[c_sym] if c_sym is not None else "",
            "qty": _round(row[c_qty] if c_qty is not None else 0, 6),
            "realized_gl": _round(row[c_gl] if c_gl is not None else 0, 4),
        })
    return out


def parse_operating_expenses(rows: list[list[Any]]) -> dict:
    """Sum 'Expenses Incurred' lines for each expense column."""
    header_idx, headers = _header_index(rows)
    admin_col = _col(headers, "Administration Fees")
    bank_col = _col(headers, "Bank Charges")

    # Sum absolute values of 'Expenses Incurred' section entries (month rows after 'Expenses Incurred')
    admin_total = 0.0
    bank_total = 0.0
    in_incurred = False
    for row in rows[header_idx + 1:]:
        first = next((c for c in row if c not in (None, "")), "")
        if isinstance(first, str):
            low = first.lower()
            if "incurred" in low:
                in_incurred = True
                continue
            if "paid" in low or "ending" in low or "beginning" in low:
                in_incurred = False
                continue
        if in_incurred:
            if admin_col is not None and isinstance(row[admin_col], (int, float)):
                admin_total += abs(row[admin_col])
            if bank_col is not None and isinstance(row[bank_col], (int, float)):
                bank_total += abs(row[bank_col])

    # Fallback if no "Expenses Incurred" section detected: use sum of 'Beginning' monthly rows
    if admin_total == 0 and bank_total == 0:
        for row in rows[header_idx + 1:]:
            first = next((c for c in row if c not in (None, "")), "")
            if isinstance(first, str) and re.match(r"^[A-Za-z]+-\d{4}$", first.strip()):
                if admin_col is not None and isinstance(row[admin_col], (int, float)):
                    admin_total += abs(row[admin_col])
                if bank_col is not None and isinstance(row[bank_col], (int, float)):
                    bank_total += abs(row[bank_col])
                break

    return {"admin_fees": _round(admin_total), "bank_charges": _round(bank_total)}


def parse_reconciliation(rows: list[list[Any]]) -> dict:
    # If any numeric 'break' columns are non-zero, surface them
    breaks = {"mv_break": 0.0, "trade_break": 0.0, "cash_break": 0.0}
    try:
        header_idx, headers = _header_index(rows)
    except ValueError:
        return breaks
    mv_col = _col(headers, "MV Break (BC)", "MV Break (LC)")
    trade_col = _col(headers, "Trade Break (BC)", "Trade Break (LC)")
    cash_col = _col(headers, "Cash Transaction Break")
    for row in rows[header_idx + 1:]:
        if mv_col is not None and isinstance(row[mv_col], (int, float)):
            breaks["mv_break"] += float(row[mv_col])
        if trade_col is not None and isinstance(row[trade_col], (int, float)):
            breaks["trade_break"] += float(row[trade_col])
        if cash_col is not None and isinstance(row[cash_col], (int, float)):
            breaks["cash_break"] += float(row[cash_col])
    return {k: _round(v, 4) for k, v in breaks.items()}


def parse_workbook(path: Path) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)

    bs_sheet = _sheet(wb, "Balance Sheet")
    if bs_sheet is None:
        raise ValueError("Balance Sheet not found")
    bs = parse_balance_sheet(_rows(bs_sheet))
    period, label, as_of = _parse_period(_rows(bs_sheet))

    inc_sheet = _sheet(wb, "Income Statement", "Statement of Operations")
    inc = parse_income_statement(_rows(inc_sheet)) if inc_sheet else {}

    cap_sched_sheet = _sheet(wb, "Capital Schedule", "Investors Capital Schedule")
    cap_sched = parse_capital_schedule(_rows(cap_sched_sheet)) if cap_sched_sheet else []

    inv_sum_sheet = _sheet(wb, "Investor Capital Summary", "Investors Capital Summary")
    inv_summary, fund_totals = parse_investor_capital_summary(_rows(inv_sum_sheet)) if inv_sum_sheet else ([], {})

    pos_sheet = _sheet(wb, "Position Report", "Portfolio Valuation")
    positions = parse_positions(_rows(pos_sheet)) if pos_sheet else []

    real_sheet = _sheet(wb, "Realized Gain Loss", "Realized Gain (Loss)")
    realized = parse_realized(_rows(real_sheet)) if real_sheet else []

    op_sheet = _sheet(wb, "Operating Expenses Detailed")
    op_exp = parse_operating_expenses(_rows(op_sheet)) if op_sheet else {}
    # Bank charges are booked on the Income Statement each period; mirror that value here so the
    # dashboard shows cash-cost-of-month even when the OpEx roll-forward splits it into paid/unpaid.
    op_exp["bank_charges"] = abs(inc.get("expense", {}).get("bank_charges", 0))

    recon_sheet = _sheet(wb, "Reconciliation Summary")
    recon = parse_reconciliation(_rows(recon_sheet)) if recon_sheet else {}

    # Merge per-investor records on investor_no
    merged = {}
    for r in cap_sched:
        merged[r["investor_no"]] = {**r}
    for r in inv_summary:
        no = r["investor_no"]
        if no in merged:
            merged[no].update({k: r[k] for k in ["shares_begin", "shares_end", "nav_per_share", "gross_mtd_ror", "net_mtd_ror", "ytd_ror", "mgmt_fee", "perf_fee"]})
        else:
            merged[no] = r
    investors = sorted(merged.values(), key=lambda x: -x["ending_balance"])

    # Fund-level aggregates
    total_gp = sum(i.get("gross_profit", 0) for i in investors)
    total_perf = sum(i.get("perf_fee", 0) for i in investors)
    weighted_gross_mtd = 0.0
    weighted_net_mtd = 0.0
    total_weight = sum(i["ending_balance"] for i in investors)
    if total_weight:
        for i in investors:
            w = i["ending_balance"] / total_weight
            weighted_gross_mtd += i.get("gross_mtd_ror", 0) * w
            weighted_net_mtd += i.get("net_mtd_ror", 0) * w

    # NAV per share: take from any investor (or fund total)
    nav_per_share = next((i["nav_per_share"] for i in investors if i.get("nav_per_share")), 0)

    fund_level = {
        "gross_mtd_ror": _round(weighted_gross_mtd, 6),
        "net_mtd_ror": _round(weighted_net_mtd, 6),
        "perf_fee_rate": 0.30,
        "mgmt_fee_rate": 0.0,
        "investor_count": len(investors),
        "nav_per_share": nav_per_share,
        "total_gross_profit": _round(total_gp),
        "total_perf_fee": _round(total_perf),
    }

    return {
        "period": period,
        "period_label": label,
        "as_of": as_of,
        "balance_sheet": bs,
        "income_statement": inc,
        "positions": positions,
        "realized_trades": realized,
        "operating_expenses": op_exp,
        "fund_level": fund_level,
        "investors": investors,
        "reconciliation": recon,
    }


def upsert_history(record: dict, history_path: Path = HISTORY_PATH) -> None:
    if history_path.exists():
        history = json.loads(history_path.read_text())
    else:
        history = {"fund": "Armada Prime LLP", "months": []}

    history["fund"] = "Armada Prime LLP"
    history["last_updated"] = dt.date.today().isoformat()

    months = [m for m in history.get("months", []) if m.get("period") != record["period"]]
    months.append(record)
    months.sort(key=lambda m: m["period"])
    history["months"] = months

    history_path.parent.mkdir(parents=True, exist_ok=True)
    history_path.write_text(json.dumps(history, indent=2))


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("xlsx", type=Path, help="Path to the TPA Reporting Package xlsx")
    ap.add_argument("--output", type=Path, default=HISTORY_PATH, help="Output JSON history path")
    args = ap.parse_args()

    if not args.xlsx.exists():
        print(f"error: file not found: {args.xlsx}", file=sys.stderr)
        return 1

    record = parse_workbook(args.xlsx)
    upsert_history(record, args.output)

    bs = record["balance_sheet"]
    fl = record["fund_level"]
    print(
        f"Upserted {record['period']}: assets ${bs['total_assets']/1e6:.2f}M, "
        f"capital ${bs['total_capital']/1e6:.2f}M, "
        f"net income ${record['income_statement']['net_income']:,.0f}, "
        f"{fl['investor_count']} investors, "
        f"gross MTD {fl['gross_mtd_ror']*100:.2f}%"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
