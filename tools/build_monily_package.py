#!/usr/bin/env python3
"""Build the consolidated Monily Partnership Tax Organizer document for
Armada Prime Tech LLC, tax year 2025.

Produces ONE combined xlsx with all financial schedules as tabs:
  /Users/nairne/claude-central-hub/monily-package/Monily_Tax_Package_2025.xlsx
    Tabs:
      1. Cover & Summary
      2. P&L Statement
      3. Balance Sheet
      4. General Ledger
      5. Asset Schedule
      6. K-1 Partners
      7. 1099 Recipients
      8. Tax Organizer Answers

Plus the narrative cover memo files:
  05_Cover_Memo_for_Monily.md
  05_Cover_Memo_for_Monily.docx (built separately by md_to_docx.py)

Per Nairne 2026-05-05: Phil is a K-1 partner (was previously thought to
be a 1099 contractor — corrected here).

Usage:
    python tools/build_monily_package.py
"""
from __future__ import annotations

import sys
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
from build_2025_year_end import (
    ACTUAL_PAID,
    GP_OP_EXPENSES,
    K1_RECIPIENTS,
    NAIRNE_ALIASES,
    PERIOD_LABELS,
)

REPO_ROOT = Path(__file__).resolve().parent.parent
OUT_DIR = REPO_ROOT / "monily-package"
OUT_DIR.mkdir(exist_ok=True)
OUT_XLSX = OUT_DIR / "Monily_Tax_Package_2025.xlsx"
OUT_MD = OUT_DIR / "05_Cover_Memo_for_Monily.md"

# Styling
HDR_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
SECTION_FONT = Font(bold=True, color="FFFFFF", size=12)
SUBTOTAL_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
WARN_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
K1_FILL = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
THIN = Side(border_style="thin", color="999999")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MONEY = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
DATE_FMT = "yyyy-mm-dd"

ENTITY_NAME = "Armada Prime Tech LLC"
TAX_YEAR = "2025"
PERIOD_DESC = "Aug 1, 2025 – Dec 31, 2025"


# ---------------------------------------------------------------------------
# Year-totals computation (centralized so all tabs use the same numbers)
# ---------------------------------------------------------------------------

def compute_year_totals():
    """Per Nairne 2026-05-07:
       - 59.5% Fund Mgmt slice split 50/50 between Nairne (K-1) and AJ (1099)
       - AJ stays a 1099 contractor; her cash already reflects both her 39%-pool
         consultant share AND her 50%×Fund Mgmt share
       - Nairne K-1 economic interest: 50%×59.5% + 0.5% direct = 30.25%
       - Raj/Phil K-1 economic interest: 0.5% each
       - Total partner interest: 31.25%
       - Tax-Reclass only (GAAP column removed)
    """
    year_totals = {}
    for period, recipients in ACTUAL_PAID.items():
        for k, v in recipients.items():
            year_totals[k] = year_totals.get(k, 0) + v

    revenue = sum([15355.51, 12474.75, 51906.99, 57074.09, 16211.69])
    nairne_total = sum(year_totals.get(a, 0) for a in NAIRNE_ALIASES)
    raj_total = year_totals.get("Raj", 0)
    phil_total = year_totals.get("Phil", 0)
    contractor_total = sum(v for k, v in year_totals.items() if k not in K1_RECIPIENTS)

    op_expenses = aggregate_op_expenses()
    op_total_gross = sum(op_expenses.values())
    spv_reclass = 4275  # Aug only — moves to Balance Sheet
    insurance_prorate = 18000 - 7500
    op_total = op_total_gross - spv_reclass - insurance_prorate  # tax-reclass amount

    net_income = revenue - contractor_total - op_total

    # NEW K-1 ownership %: Nairne 30.25/31.25, Raj 0.5/31.25, Phil 0.5/31.25
    nairne_pct = 30.25 / 31.25  # 96.80%
    raj_pct = 0.5 / 31.25       # 1.60%
    phil_pct = 0.5 / 31.25      # 1.60%

    return {
        "year_totals": year_totals,
        "revenue": revenue,
        "nairne_cash": nairne_total,
        "raj_cash": raj_total,
        "phil_cash": phil_total,
        "contractor_total": contractor_total,
        "op_expenses": op_expenses,
        "op_total_gross": op_total_gross,
        "op_total": op_total,
        "spv_reclass": spv_reclass,
        "insurance_prorate": insurance_prorate,
        "net_income": net_income,
        "nairne_pct": nairne_pct,
        "raj_pct": raj_pct,
        "phil_pct": phil_pct,
    }


def aggregate_op_expenses() -> dict:
    out = {}
    mapping = {
        "506c SPV Loan": "506c SPV Loan",
        "PVD": "PVD",
        "Website": "Website",
        "Ad Spend": "Ad Spend",
        "Chris": "Chris",
        "Alpha Verification": "Alpha Verification",
        "Formidium (TPA)": "TPA",
        "TPA (Formidium)": "TPA",
        "TPA (second line)": "TPA",
        "Insurance": "Insurance",
    }
    for period, items in GP_OP_EXPENSES.items():
        for vendor, amount in items:
            key = mapping.get(vendor, vendor)
            out[key] = out.get(key, 0) + amount
    return out


def categorize_account(vendor: str) -> str:
    v = vendor.lower()
    if "insurance" in v: return "Insurance Expense"
    if "spv" in v or "loan" in v: return "506c SPV Loan (RECLASS to Asset)"
    if "tpa" in v or "formidium" in v: return "Professional Fees — TPA Admin"
    if "chris" in v: return "Contractor Labor — Chris"
    if "pvd" in v: return "Professional Fees — PVD"
    if "website" in v: return "Marketing — Website"
    if "ad spend" in v or "ads" in v or "badtwin" in v: return "Marketing — Advertising"
    if "alpha" in v: return "Compliance — Verification"
    return "Other Operating Expense"


# ---------------------------------------------------------------------------
# Tab helpers
# ---------------------------------------------------------------------------

def title_block(ws, doc_title: str, ncols: int = 5) -> int:
    ws.cell(row=1, column=1, value=ENTITY_NAME).font = Font(bold=True, size=16)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.cell(row=2, column=1, value=doc_title).font = Font(bold=True, size=13, color="1F4E79")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    ws.cell(row=3, column=1, value=f"Tax Year {TAX_YEAR}  |  Period: {PERIOD_DESC}  |  Generated {date.today().isoformat()}").font = Font(italic=True, color="666666")
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=ncols)
    return 5


def set_header_row(ws, row: int, ncols: int):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BOX


def write_section(ws, row: int, title: str, ncols: int) -> int:
    ws.cell(row=row, column=1, value=title)
    for c in range(1, ncols + 1):
        ws.cell(row=row, column=c).fill = SECTION_FILL
        ws.cell(row=row, column=c).font = SECTION_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    return row + 1


# ---------------------------------------------------------------------------
# Tab 1: Cover & Summary
# ---------------------------------------------------------------------------

def build_cover_tab(wb, T):
    ws = wb.create_sheet("1. Cover & Summary")

    r = title_block(ws, "Tax Package — Cover & Summary", ncols=4)

    # Filing summary
    info_rows = [
        ("Entity Name", ENTITY_NAME),
        ("Filing for Tax Year", TAX_YEAR),
        ("Period of Operations", PERIOD_DESC),
        ("Filing Status", "First year of filing (entity formed at Armada Prime relaunch)"),
        ("Tax Classification", "Multi-member LLC, taxed as partnership"),
        ("Number of Partners", "3 (Nairne, Raj Duggal, Phil)"),
        ("Calendar Year Filer", "Yes"),
    ]
    for label, value in info_rows:
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws.cell(row=r, column=2, value=value)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        r += 1
    r += 1

    r = write_section(ws, r, "HEADLINE NUMBERS (Tax-Reclass Basis)", ncols=4)
    headers = ["Line Item", "Amount ($)", "Notes", ""]
    for i, h in enumerate(headers, 1):
        ws.cell(row=r, column=i, value=h)
    set_header_row(ws, r, 4)
    r += 1
    rows = [
        ("Revenue (Performance Fees from Armada Prime LLP)", T["revenue"], "TPA-authoritative; Aug-Dec 2025"),
        ("Less: 1099 Contractor Expenses", -T["contractor_total"], "Alec, Jake, AJ, Issac (Luke + Nikki <$600 — no 1099 required, but still expense)"),
        ("Less: Operating Expenses (after reclass)", -T["op_total"], "Strips $4,275 SPV loan to balance sheet; pro-rates $18K Insurance to $7,500"),
        ("PARTNERSHIP NET INCOME", T["net_income"], "Allocated to K-1 partners (Nairne, Raj, Phil)"),
    ]
    for label, amount, note in rows:
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=amount).number_format = MONEY
        ws.cell(row=r, column=3, value=note).alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        if "NET INCOME" in label:
            for c in range(1, 5):
                ws.cell(row=r, column=c).fill = TOTAL_FILL
                ws.cell(row=r, column=c).font = Font(bold=True)
        r += 1
    r += 1

    r = write_section(ws, r, "K-1 PARTNERS (3 partners) — Per Nairne 2026-05-07: Fund Mgmt 59.5% split 50/50 between Nairne (K-1) and AJ (1099)", ncols=4)
    headers = ["Partner", "Ownership %", "Cash Distributions ($)", "Allocated K-1 Income ($)"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=r, column=i, value=h)
    set_header_row(ws, r, 4)
    r += 1
    k1_rows = [
        (f"Nairne (50%×Fund Mgmt 59.5% + direct 0.5% = 30.25% of perf fees)", T["nairne_pct"], T["nairne_cash"], T["net_income"] * T["nairne_pct"]),
        ("Raj Duggal (direct 0.5%)", T["raj_pct"], T["raj_cash"], T["net_income"] * T["raj_pct"]),
        ("Phil (direct 0.5%)", T["phil_pct"], T["phil_cash"], T["net_income"] * T["phil_pct"]),
    ]
    for partner, pct, cash, allocated in k1_rows:
        ws.cell(row=r, column=1, value=partner)
        ws.cell(row=r, column=2, value=f"{pct*100:.2f}%")
        ws.cell(row=r, column=3, value=cash).number_format = MONEY
        ws.cell(row=r, column=4, value=allocated).number_format = MONEY
        for c in range(1, 5):
            ws.cell(row=r, column=c).fill = K1_FILL
        r += 1
    r += 1

    r = write_section(ws, r, "TAB GUIDE", ncols=4)
    tab_guide = [
        ("1. Cover & Summary", "This tab — entity info, headline numbers, K-1 partner allocations, file overview"),
        ("2. P&L Statement", "Profit & Loss (Tax-Reclass basis), line-by-line detail"),
        ("3. Balance Sheet", "As of Dec 31, 2025. Best-effort with placeholders flagged for accountant"),
        ("4. General Ledger", "Transaction-level ledger of all cash movements (revenue, distributions, expenses)"),
        ("5. Asset Schedule", "Fixed asset / depreciation schedule. No fixed assets recorded."),
        ("6. K-1 Partners", "Detailed partner schedule for K-1 prep (Nairne, Raj, Phil)"),
        ("7. 1099 Recipients", "Detailed contractor schedule for 1099-NEC prep"),
        ("8. Tax Organizer Answers", "Pre-filled answers for the Monily Partnership Tax Organizer form"),
    ]
    for tab_name, desc in tab_guide:
        ws.cell(row=r, column=1, value=tab_name).font = Font(bold=True)
        ws.cell(row=r, column=2, value=desc)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True)
        r += 1

    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 50


# ---------------------------------------------------------------------------
# Tab 2: P&L Statement
# ---------------------------------------------------------------------------

def build_pnl_tab(wb, T):
    ws = wb.create_sheet("2. P&L Statement")
    r = title_block(ws, "Profit & Loss Statement (Tax-Reclass Basis)", ncols=4)
    headers = ["Line Item", "Amount ($)", "Notes", ""]
    for i, h in enumerate(headers, 1):
        ws.cell(row=r, column=i, value=h)
    set_header_row(ws, r, 4)
    r += 1

    def line(label, amount=None, note="", bold=False, fill=None):
        nonlocal r
        ws.cell(row=r, column=1, value=label)
        if amount is not None:
            ws.cell(row=r, column=2, value=amount).number_format = MONEY
        ws.cell(row=r, column=3, value=note).alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        if bold:
            for c in range(1, 5):
                ws.cell(row=r, column=c).font = Font(bold=True)
        if fill:
            for c in range(1, 5):
                ws.cell(row=r, column=c).fill = fill
        r += 1

    r = write_section(ws, r, "REVENUE", 4)
    line("Performance Fees Income (from Armada Prime LLP)", T["revenue"],
         "Per TPA Reporting Packages, Performance Fees Crystallized line, Aug-Dec 2025.")
    line("Total Revenue", T["revenue"], bold=True, fill=SUBTOTAL_FILL)
    r += 1

    r = write_section(ws, r, "DIRECT COSTS — Capital Raiser Commissions (1099 contractors)", 4)
    yt = T["year_totals"]
    contractors = [
        ("Alec Atkinson", yt.get("Alec Atkinson", 0), "1099-NEC required (>$600)"),
        ("AJ Affleck", yt.get("AJ Affleck", 0), "1099-NEC required (>$600). Includes 50%×Fund Mgmt 59.5% + her 39%-pool consultant share per Nairne 2026-05-07."),
        ("Jake Gordon", yt.get("Jake Gordon", 0), "1099-NEC required (>$600)"),
        ("Issac Morris", yt.get("Issac", 0), "1099-NEC required (>$600)"),
        ("Luke Affleck", yt.get("Luke", 0), "Less than $600 — NO 1099 required (still expense to GP)"),
        ("Nikki", yt.get("Nikki", 0), "Less than $600 — NO 1099 required (still expense to GP)"),
    ]
    for name, amt, note in contractors:
        line(f"  {name}", amt, note)
    line("Total Direct Costs", T["contractor_total"], bold=True, fill=SUBTOTAL_FILL)
    r += 1

    r = write_section(ws, r, "OPERATING EXPENSES (Tax-Reclass)", 4)
    op = T["op_expenses"]
    op_lines = [
        ("Chris (Contractor labor — 1099)", op.get("Chris", 0), "Issue 1099-NEC; verify SSN/address"),
        ("Consulting (Nov)", op.get("Consulting", 0), "Operating consulting; verify recipient for 1099"),
        ("Insurance (D&O)", 7500, "Pro-rated: $18k Dec is annual D&O — only 5/12 ($7,500) hits 2025; remaining $10,500 → Prepaid Asset on Balance Sheet"),
        ("PVD (Tech/Ads)", op.get("PVD", 0), "Vendor identification needed for 1099 obligation"),
        ("Website", op.get("Website", 0), "Marketing/web build"),
        ("Alpha Verification", op.get("Alpha Verification", 0), "Compliance/verification service"),
        ("TPA Admin Fees (Formidium)", op.get("TPA", 0), "Verify GP-paid vs fund-paid (fund books also have $600/mo)"),
    ]
    for label, amt, note in op_lines:
        line(f"  {label}", amt, note)
    line("Total Operating Expenses", T["op_total"], bold=True, fill=SUBTOTAL_FILL)
    r += 1

    r = write_section(ws, r, "RECLASSIFIED OFF P&L (To Balance Sheet)", 4)
    line("  506c SPV Loan disbursement (Aug)", T["spv_reclass"],
         "Loan receivable / equity investment — appears on Balance Sheet (Tab 3), NOT on P&L")
    line("  Insurance prepaid portion (covers 2026)", T["insurance_prorate"],
         "$10,500 of Dec $18K Insurance covers 2026 — booked as Prepaid Asset on Balance Sheet")
    r += 1

    r = write_section(ws, r, "NET INCOME (Partnership)", 4)
    line("Total Revenue", T["revenue"])
    line("Less: Direct Costs (1099 contractors)", -T["contractor_total"])
    line("Less: Operating Expenses (after reclass)", -T["op_total"])
    line("= NET INCOME", T["net_income"], bold=True, fill=TOTAL_FILL,
         note="This is what flows to Schedule K Line 1 → allocated to K-1 partners.")
    r += 1

    r = write_section(ws, r, "K-1 PARTNER ALLOCATION (Per Nairne 2026-05-07: Fund Mgmt 59.5% split 50/50 between Nairne K-1 and AJ 1099)", 4)
    line(f"Nairne — Cash distributions received", T["nairne_cash"],
         "Includes 50%×Fund Mgmt 59.5% + direct 0.5%. K-1 capital account.")
    line(f"Nairne — Allocated share of Net Income ({T['nairne_pct']*100:.2f}%)",
         T["net_income"] * T["nairne_pct"],
         "K-1 Box 1 (Ordinary Income). Ownership = 30.25/31.25 of partner share.")
    line("Raj Duggal — Cash distributions received", T["raj_cash"], "K-1 capital account.")
    line(f"Raj Duggal — Allocated share of Net Income ({T['raj_pct']*100:.2f}%)",
         T["net_income"] * T["raj_pct"], "K-1 Box 1. Ownership = 0.5/31.25 of partner share.")
    line("Phil — Cash distributions received", T["phil_cash"], "K-1 capital account.")
    line(f"Phil — Allocated share of Net Income ({T['phil_pct']*100:.2f}%)",
         T["net_income"] * T["phil_pct"], "K-1 Box 1. Ownership = 0.5/31.25 of partner share.")
    r += 1

    r = write_section(ws, r, "PREPARER NOTES", 4)
    notes = [
        "1. Entity formed at the Armada Prime relaunch in August 2025; first year of operations.",
        "2. Accounting method: ACCRUAL basis (recommended for tax minimization).",
        "3. Revenue source: TPA (Formidium) Performance Fees Crystallized line, Aug-Dec 2025.",
        "4. Member/economic structure (per Nairne 2026-05-07): Fund Mgmt 59.5% slice is split 50/50 between AJ Affleck (1099 contractor) and Nairne (K-1 partner). Plus 0.5% direct slices to each of Nairne, Raj, Phil. All three of Nairne/Raj/Phil are K-1 partners.",
        "5. AJ Affleck's 1099 amount thus includes BOTH her 39%-pool consultant share AND her 50%×Fund Mgmt 29.75% share.",
        "6. K-1 ownership %: Nairne 30.25/31.25 = 96.80%; Raj 0.5/31.25 = 1.60%; Phil 0.5/31.25 = 1.60%. Total partner interest = 31.25% of pre-distribution gross.",
        "7. TruQuant payments are NOT included. August 'Trader & Developer' $6,909.93 + 'Spydr' $88.78 excluded per 2026-04-30 policy (TQ is upstream of GP entity).",
        "8. RECOMMEND: $4,275 of August '506c SPV Loan' is a loan/capital item — reclassified to Balance Sheet.",
        "9. RECOMMEND: $18,000 Dec Insurance line is annual D&O — $7,500 pro-rated to 2025; remaining $10,500 booked as Prepaid Asset on Balance Sheet.",
        "10. 1099 threshold: contractors receiving <$600/year (Luke $164.90, Nikki $223.00) do NOT require 1099-NEC issuance, but their amounts ARE still expense to GP entity.",
        "11. K-1: NO threshold — partners receive K-1 every year regardless of dollar amount. Raj and Phil get K-1s even at $125 allocations.",
    ]
    for note in notes:
        line(note)
        ws.row_dimensions[r-1].height = 30

    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 30


# ---------------------------------------------------------------------------
# Tab 3: Balance Sheet
# ---------------------------------------------------------------------------

def build_balance_sheet_tab(wb, T):
    ws = wb.create_sheet("3. Balance Sheet")
    r = title_block(ws, "Balance Sheet (as of Dec 31, 2025)", ncols=4)
    headers = ["Account", "Amount ($)", "Source / Method", "Notes"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=r, column=i, value=h)
    set_header_row(ws, r, 4)
    r += 1

    def line(label, amount=None, source="", note="", bold=False, fill=None, warn=False):
        nonlocal r
        ws.cell(row=r, column=1, value=label)
        if amount is not None:
            ws.cell(row=r, column=2, value=amount).number_format = MONEY
        ws.cell(row=r, column=3, value=source)
        ws.cell(row=r, column=4, value=note).alignment = Alignment(wrap_text=True)
        if bold:
            for c in range(1, 5):
                ws.cell(row=r, column=c).font = Font(bold=True)
        if warn:
            for c in range(1, 5):
                ws.cell(row=r, column=c).fill = WARN_FILL
        elif fill:
            for c in range(1, 5):
                ws.cell(row=r, column=c).fill = fill
        r += 1

    spv_loans = T["spv_reclass"]

    r = write_section(ws, r, "ASSETS", 4)
    line("Cash & Cash Equivalents (Bank + Crypto Wallets)", None,
         "PLACEHOLDER", "⚠️ Need bank statements / wallet balances as of 12/31/2025", warn=True)
    line("506c SPV Loan Receivable / SPV Investment", spv_loans,
         "Reclassified from P&L", "Aug $4,275 + Oct $25,000")
    line("Prepaid Insurance", 10500,
         "Reclassified from P&L", "Dec $18k less $7,500 pro-rated to 2025 = $10,500 prepaid for 2026")
    line("Other Receivables / Accruals", None,
         "PLACEHOLDER", "⚠️ Any TPA accruals?", warn=True)
    line("Total Assets (excl. placeholders)", spv_loans + 10500, bold=True, fill=SUBTOTAL_FILL)
    r += 1

    r = write_section(ws, r, "LIABILITIES", 4)
    line("Accounts Payable", None, "PLACEHOLDER", "⚠️ Any unpaid contractor amounts? Distributions Ledger shows ~$11,588 cumulative cash-vs-accrual delta", warn=True)
    line("Accrued Expenses", None, "PLACEHOLDER", "⚠️ Any Q4 op expenses incurred but not paid?", warn=True)
    line("Total Liabilities", 0, bold=True, fill=SUBTOTAL_FILL)
    r += 1

    r = write_section(ws, r, "MEMBERS' EQUITY", 4)
    line("Member Capital — Nairne (60% ownership)", None,
         "PLACEHOLDER", "⚠️ Initial capital contributions in 2025?", warn=True)
    line("Member Capital — Raj Duggal (0.5% ownership)", None,
         "PLACEHOLDER", "⚠️ Initial capital contributions in 2025?", warn=True)
    line("Member Capital — Phil (0.5% ownership)", None,
         "PLACEHOLDER", "⚠️ Initial capital contributions in 2025?", warn=True)
    line("Cumulative Distributions — Nairne", -T["nairne_cash"],
         "Cash distributions made in 2025", f"Fund Mgmt ${T['nairne_cash'] - 946.97:,.2f} + direct ${946.97:,.2f}")
    line("Cumulative Distributions — Raj Duggal", -T["raj_cash"], "Cash distributions made in 2025")
    line("Cumulative Distributions — Phil", -T["phil_cash"], "Cash distributions made in 2025")
    line("Retained Earnings (Net Income for the year)", T["net_income"],
         "From P&L Statement (Tax-Reclass basis)", f"${T['net_income']:,.2f}")
    equity_total = T["net_income"] - T["nairne_cash"] - T["raj_cash"] - T["phil_cash"]
    line("Total Members' Equity (excl. placeholders)", equity_total, bold=True, fill=SUBTOTAL_FILL)
    r += 1

    r = write_section(ws, r, "BALANCING NOTE", 4)
    line("Total Assets", spv_loans + 10500)
    line("Total Liabilities + Equity", equity_total,
         note="When member capital + cash + payables are filled in, this should balance.")
    line("Imbalance (placeholder gap)", (spv_loans + 10500) - equity_total,
         note="Reflects the missing cash + member capital data.", warn=True)
    r += 1

    r = write_section(ws, r, "METHODOLOGY", 4)
    method_notes = [
        "1. Best-effort balance sheet built from transactional data. PLACEHOLDER lines need population by accountant.",
        "2. 506c SPV Loans ($4,275) reclassified from P&L to assets.",
        "3. $18,000 Dec Insurance split: $7,500 to 2025 P&L, $10,500 booked as Prepaid Asset.",
        "4. Cumulative Distributions shown as negative equity. Member capital contributions need to be added.",
        "5. The $11,588 cash-vs-accrual delta on Distributions ledger may flow through Accounts Payable. Confirm tax basis with accountant.",
    ]
    for note in method_notes:
        line(note)
        ws.row_dimensions[r-1].height = 30

    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 60


# ---------------------------------------------------------------------------
# Tab 4: General Ledger
# ---------------------------------------------------------------------------

def build_gl_tab(wb, T):
    ws = wb.create_sheet("4. General Ledger")
    r = title_block(ws, "General Ledger — Cash Transactions", ncols=7)
    headers = ["Date", "Type", "Account", "Counterparty", "Description", "Debit ($)", "Credit ($)"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=r, column=i, value=h)
    set_header_row(ws, r, 7)
    r += 1

    perf_fees = {
        "2025-08": 15355.51, "2025-09": 12474.75, "2025-10": 51906.99,
        "2025-11": 57074.09, "2025-12": 16211.69,
    }
    eom = {
        "2025-08": "2025-08-31", "2025-09": "2025-09-30", "2025-10": "2025-10-31",
        "2025-11": "2025-11-30", "2025-12": "2025-12-31",
    }

    txn_count = 0
    total_debit = 0.0
    total_credit = 0.0

    for period in ["2025-08", "2025-09", "2025-10", "2025-11", "2025-12"]:
        d = eom[period]
        amt = perf_fees[period]
        write_txn(ws, r, d, "Revenue", "Performance Fees Income", "Armada Prime LLP (TPA)",
                  f"{PERIOD_LABELS[period]} GP cut received from fund", debit=amt, credit=None)
        r += 1
        txn_count += 1
        total_debit += amt

        for recipient, payout in ACTUAL_PAID.get(period, {}).items():
            if payout == 0:
                continue
            if recipient == "Fund Mgmt":
                acct = "K-1 Distribution to Member"
                desc = "Distribution to Nairne (Fund Mgmt 59.5% slice)"
                cp = "Nairne"
            elif recipient == "Nairne":
                acct = "K-1 Distribution to Member"
                desc = "Distribution to Nairne (direct 0.5%)"
                cp = "Nairne"
            elif recipient == "Raj":
                acct = "K-1 Distribution to Member"
                desc = "Distribution to Raj Duggal (direct 0.5%)"
                cp = "Raj Duggal"
            elif recipient == "Phil":
                acct = "K-1 Distribution to Member"
                desc = "Distribution to Phil (direct 0.5%, K-1 partner per Nairne 2026-05-05)"
                cp = "Phil"
            else:
                acct = "Capital Raiser Commission Expense (1099)"
                desc = f"1099-NEC payment to {recipient}"
                cp = recipient

            if payout >= 0:
                debit_val = None
                credit_val = payout
                total_credit += payout
            else:
                debit_val = -payout
                credit_val = None
                total_debit += -payout

            ttype = "Distribution" if recipient in K1_RECIPIENTS else "Expense"
            write_txn(ws, r, d, ttype, acct, cp, desc, debit=debit_val, credit=credit_val)
            r += 1
            txn_count += 1

        for vendor, amount in GP_OP_EXPENSES.get(period, []):
            acct = categorize_account(vendor)
            write_txn(ws, r, d, "Expense", acct, vendor,
                      f"GP-paid expense: {vendor}", debit=None, credit=amount)
            r += 1
            txn_count += 1
            total_credit += amount

    r += 1
    ws.cell(row=r, column=1, value=f"TOTAL TRANSACTIONS: {txn_count}").font = Font(bold=True)
    r += 1
    ws.cell(row=r, column=1, value="TOTAL DEBITS").font = Font(bold=True)
    ws.cell(row=r, column=6, value=total_debit).number_format = MONEY
    ws.cell(row=r, column=6).font = Font(bold=True)
    r += 1
    ws.cell(row=r, column=1, value="TOTAL CREDITS").font = Font(bold=True)
    ws.cell(row=r, column=7, value=total_credit).number_format = MONEY
    ws.cell(row=r, column=7).font = Font(bold=True)
    r += 1
    ws.cell(row=r, column=1, value=f"NET (Debit − Credit) = Net Income proxy: ${total_debit - total_credit:,.2f}").font = Font(italic=True, bold=True)

    r += 2
    ws.cell(row=r, column=1, value="NOTES").font = Font(bold=True, color="C00000")
    r += 1
    notes = [
        "Each row = one cash transaction. Revenue receipts are debits to Cash; distributions and expense payments are credits to Cash.",
        "K-1 Distribution to Member: Nairne, Raj, Phil (all K-1 partners).",
        "Capital Raiser Commission Expense: Alec, Jake, AJ, Issac, Luke (1099 contractors).",
        "Dates are end-of-month. Actual settlement may have been days/weeks later.",
        "Source: Distributions Armada Tech 2025 ledger + BEST ONE Dec 2025 Costs.",
        "TruQuant payments excluded entirely (per Nairne 2026-04-30).",
        "Issac's December −$278.48 entry is a clawback against prior month overpayment.",
    ]
    for note in notes:
        ws.cell(row=r, column=1, value=note).alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        ws.row_dimensions[r].height = 30
        r += 1

    for col, w in zip("ABCDEFG", [12, 13, 35, 22, 50, 14, 14]):
        ws.column_dimensions[col].width = w


def write_txn(ws, row, dt, ttype, account, counterparty, desc, debit=None, credit=None):
    ws.cell(row=row, column=1, value=dt).number_format = DATE_FMT
    ws.cell(row=row, column=2, value=ttype)
    ws.cell(row=row, column=3, value=account)
    ws.cell(row=row, column=4, value=counterparty)
    ws.cell(row=row, column=5, value=desc)
    if debit is not None:
        ws.cell(row=row, column=6, value=debit).number_format = MONEY
    if credit is not None:
        ws.cell(row=row, column=7, value=credit).number_format = MONEY


# ---------------------------------------------------------------------------
# Tab 5: Asset Schedule
# ---------------------------------------------------------------------------

def build_asset_tab(wb, T):
    ws = wb.create_sheet("5. Asset Schedule")
    r = title_block(ws, "Asset Schedule (Fixed Assets / Depreciation)", ncols=8)
    headers = ["Asset Description", "Acquisition Date", "Cost", "Useful Life", "Depreciation Method", "Prior Accum Depr", "2025 Depr Expense", "Net Book Value"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=r, column=i, value=h)
    set_header_row(ws, r, 8)
    r += 1

    ws.cell(row=r, column=1, value="(No fixed assets recorded for 2025)").font = Font(italic=True, color="999999")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    r += 2

    ws.cell(row=r, column=1, value="OTHER ITEMS THAT MAY BELONG ON SCHEDULE").font = Font(bold=True, color="C00000")
    r += 1
    other_items = [
        ("506c SPV Loan disbursement (Aug)", "2025-08-31", 4275.00, "—", "Loan/Investment (NOT depreciable)", 0, 0, 4275.00),
        ("506c SPV Loan disbursement (Oct)", "2025-10-31", 25000.00, "—", "Loan/Investment (NOT depreciable)", 0, 0, 25000.00),
    ]
    for desc, dt, cost, life, method, prior_dep, this_dep, nbv in other_items:
        ws.cell(row=r, column=1, value=desc)
        ws.cell(row=r, column=2, value=dt)
        ws.cell(row=r, column=3, value=cost).number_format = MONEY
        ws.cell(row=r, column=4, value=life)
        ws.cell(row=r, column=5, value=method)
        ws.cell(row=r, column=6, value=prior_dep).number_format = MONEY
        ws.cell(row=r, column=7, value=this_dep).number_format = MONEY
        ws.cell(row=r, column=8, value=nbv).number_format = MONEY
        r += 1

    r += 2
    ws.cell(row=r, column=1, value="NOTES").font = Font(bold=True)
    r += 1
    notes = [
        "Armada Prime Tech LLC has no fixed assets / depreciable property recorded for 2025. Services-only LLC.",
        "506c SPV Loan disbursements ($4,275 total) are NOT depreciable. They are loan receivables / equity investments — Balance Sheet items.",
        "If equipment / software / capital assets were purchased in 2025, please add them.",
    ]
    for note in notes:
        ws.cell(row=r, column=1, value=note).alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        ws.row_dimensions[r].height = 30
        r += 1

    for i, w in enumerate([35, 14, 14, 14, 28, 16, 16, 16], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---------------------------------------------------------------------------
# Tab 6: K-1 Partners
# ---------------------------------------------------------------------------

def build_k1_tab(wb, T):
    ws = wb.create_sheet("6. K-1 Partners")
    r = title_block(ws, "K-1 Partner Schedule", ncols=6)

    ws.cell(row=r, column=1, value="Per Nairne 2026-05-07: 3 K-1 partners. Fund Mgmt 59.5% slice is split 50/50 — half to AJ Affleck (1099 contractor) and half to Nairne (K-1). Plus direct 0.5% slices to Nairne, Raj, and Phil. Total partner economic interest = 31.25% of pre-distribution gross.").font = Font(italic=True, color="C00000")
    ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws.row_dimensions[r].height = 45
    r += 2

    headers = ["Partner Name", "Economic Interest", "Ownership % (Normalized)", "Cash Distributions ($)", "K-1 Allocated Net Income ($)", "SSN/Address (fill in)"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=r, column=i, value=h)
    set_header_row(ws, r, 6)
    r += 1

    partners = [
        ("Nairne", "30.25% (= 50%×Fund Mgmt 59.5% + direct 0.5%)", T["nairne_pct"], T["nairne_cash"]),
        ("Raj Duggal", "0.50% (direct)", T["raj_pct"], T["raj_cash"]),
        ("Phil", "0.50% (direct)", T["phil_pct"], T["phil_cash"]),
    ]
    total_cash = 0
    total_alloc = 0
    for name, interest, pct, cash in partners:
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=interest)
        ws.cell(row=r, column=3, value=f"{pct*100:.2f}%")
        ws.cell(row=r, column=4, value=cash).number_format = MONEY
        alloc = T["net_income"] * pct
        ws.cell(row=r, column=5, value=alloc).number_format = MONEY
        ws.cell(row=r, column=6, value="")
        for c in range(1, 7):
            ws.cell(row=r, column=c).fill = K1_FILL
        total_cash += cash
        total_alloc += alloc
        r += 1

    ws.cell(row=r, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=r, column=2, value="31.25% partner / 68.75% non-partner")
    ws.cell(row=r, column=3, value="100.00%")
    ws.cell(row=r, column=4, value=total_cash).number_format = MONEY
    ws.cell(row=r, column=5, value=total_alloc).number_format = MONEY
    for c in range(1, 7):
        ws.cell(row=r, column=c).fill = TOTAL_FILL
        ws.cell(row=r, column=c).font = Font(bold=True)
    r += 2

    ws.cell(row=r, column=1, value="NOTES").font = Font(bold=True)
    r += 1
    notes = [
        "Per Nairne 2026-05-07: the 59.5% Fund Management slice is split 50/50 between AJ Affleck (1099 contractor) and Nairne (K-1 partner). AJ's $37,139.89 in 2025 1099 income includes BOTH her 39%-pool consultant share AND her 50%×Fund Mgmt share.",
        "Nairne's K-1 economic interest is 30.25% of pre-distribution gross (= 50%×59.5% + 0.5% direct).",
        "Total K-1 partner interest = 31.25% (Nairne 30.25 + Raj 0.5 + Phil 0.5). Normalized partner ownership: Nairne 96.80%, Raj 1.60%, Phil 1.60%.",
        "Cash Distributions and Allocated K-1 Income are different concepts in partnership tax. K-1 Box 1 reports allocated income (taxable to partner); Box L tracks capital account changes from cash distributions (NOT additionally taxed).",
        "K-1 issuance: ALL partners receive a K-1 every year regardless of dollar amount — even tiny allocations like Raj's $125 require a K-1. There is no $600 threshold for K-1s (unlike 1099s).",
        "The actual LLC operating agreement governs the legal allocation. Confirm with accountant before filing.",
    ]
    for note in notes:
        ws.cell(row=r, column=1, value=note).alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        ws.row_dimensions[r].height = 30
        r += 1

    for i, w in enumerate([20, 14, 22, 28, 28, 35], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---------------------------------------------------------------------------
# Tab 7: 1099 Recipients
# ---------------------------------------------------------------------------

def build_1099_tab(wb, T):
    ws = wb.create_sheet("7. 1099 Recipients")
    r = title_block(ws, "1099-NEC Recipients Schedule", ncols=5)

    headers = ["Recipient", "2025 Total ($)", "1099-NEC Required?", "Source of Payment", "EIN/SSN + Address (fill in)"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=r, column=i, value=h)
    set_header_row(ws, r, 5)
    r += 1

    yt = T["year_totals"]
    contractors = [
        ("Alec Atkinson", yt.get("Alec Atkinson", 0), True, "Capital raiser commission (39% of his investors' perf fees)"),
        ("AJ Affleck", yt.get("AJ Affleck", 0), True, "Capital raiser commission + 50%×Fund Mgmt 59.5% (per Nairne 2026-05-07)"),
        ("Jake Gordon", yt.get("Jake Gordon", 0), True, "Capital raiser commission"),
        ("Chris (last name TBD)", 11500.00, True, "Operating contractor labor (Nov $7,500 + Dec $4,000)"),
        ("Issac Morris", yt.get("Issac", 0), True, "Capital raiser commission"),
        ("Nikki (last name TBD)", yt.get("Nikki", 0), False, "Capital raiser commission (Nov 2025 only) — under $600"),
        ("Luke Affleck", yt.get("Luke", 0), False, "Capital raiser commission (Dec 2025 only) — under $600"),
    ]
    total_required = 0
    total_all = 0
    for name, amt, required, src in contractors:
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=amt).number_format = MONEY
        ws.cell(row=r, column=3, value="YES" if required else "NO (<$600 threshold)")
        ws.cell(row=r, column=4, value=src).alignment = Alignment(wrap_text=True)
        ws.cell(row=r, column=5, value="")
        if not required:
            for c in range(1, 6):
                ws.cell(row=r, column=c).fill = WARN_FILL
        total_all += amt
        if required:
            total_required += amt
        r += 1

    ws.cell(row=r, column=1, value="TOTAL — All Contractor Payments (P&L expense)").font = Font(bold=True)
    ws.cell(row=r, column=2, value=total_all).number_format = MONEY
    for c in range(1, 6):
        ws.cell(row=r, column=c).fill = SUBTOTAL_FILL
        ws.cell(row=r, column=c).font = Font(bold=True)
    r += 1
    ws.cell(row=r, column=1, value="TOTAL — 1099-NEC Forms Required to Issue").font = Font(bold=True)
    ws.cell(row=r, column=2, value=total_required).number_format = MONEY
    ws.cell(row=r, column=3, value="(only contractors >$600)")
    for c in range(1, 6):
        ws.cell(row=r, column=c).fill = TOTAL_FILL
        ws.cell(row=r, column=c).font = Font(bold=True)
    r += 2

    ws.cell(row=r, column=1, value="VENDORS THAT MAY ALSO REQUIRE 1099").font = Font(bold=True, color="C00000")
    r += 1
    vendors_check = [
        ("PVD (Tech/Ads)", 6000.00, "If individual/sole prop and >$600 → 1099-NEC. Verify entity type."),
        ("Website builder", 7500.00, "If individual contractor. Verify."),
        ("Alpha Verification", 2250.00, "Likely corporate; unlikely 1099 needed. Verify."),
        ("Insurance carrier", 18000.00, "Corporate — no 1099."),
        ("Formidium (TPA)", 5700.00, "Corporate — no 1099."),
        ("Consulting (Nov)", 7500.00, "Verify recipient identity for 1099 requirement."),
    ]
    for name, amt, note in vendors_check:
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=amt).number_format = MONEY
        ws.cell(row=r, column=4, value=note).alignment = Alignment(wrap_text=True)
        r += 1
    r += 1

    notes = [
        "1099-NEC threshold: $600/year per recipient. Below that, NO 1099 required (but still expense to GP entity).",
        "Luke ($164.90) and Nikki ($223.00) are under $600 — no 1099-NEC issuance required for them.",
        "Phil is NOT on this 1099 list — Phil is a K-1 partner (per Nairne 2026-05-05). See K-1 Partners tab.",
        "AJ Affleck IS a 1099 contractor (NOT a K-1 partner). Her $37,139.89 includes BOTH her 39%-pool consultant share AND her 50%×Fund Mgmt 59.5% share (per Nairne 2026-05-07).",
        "All 1099-NECs must be furnished to recipients by January 31 of the following year.",
    ]
    for note in notes:
        ws.cell(row=r, column=1, value=note).alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        ws.row_dimensions[r].height = 30
        r += 1

    for i, w in enumerate([22, 16, 24, 50, 30], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---------------------------------------------------------------------------
# Tab 8: Tax Organizer Answers
# ---------------------------------------------------------------------------

def build_organizer_tab(wb, T):
    ws = wb.create_sheet("8. Tax Organizer Answers")
    r = title_block(ws, "Monily Partnership Tax Organizer — Pre-Filled Answers", ncols=3)

    sections = [
        ("CORPORATION GENERAL INFORMATION", [
            ("Legal name of business", "Armada Prime Tech LLC"),
            ("Filing for the year", "2025"),
            ("EIN", "⚠️ TO PROVIDE — from IRS letter"),
            ("Phone Number", "⚠️ TO PROVIDE"),
            ("Email", "⚠️ TO PROVIDE"),
            ("Corporation address / State / City / Zip / Country", "⚠️ TO PROVIDE"),
            ("Above address is new", "No"),
            ("Is it first year of Filing", "Yes"),
            ("Partnership state residence", "⚠️ TO PROVIDE (state of formation)"),
        ]),
        ("DOCUMENTS (Yes/No/N/A)", [
            ("EIN letter", "Yes (need to upload)"),
            ("Letter of Incorporation", "Yes (need to upload)"),
            ("Profit & Loss Statement", "Yes — see Tab 2"),
            ("Balance Sheet", "Partial — see Tab 3 (placeholders flagged)"),
            ("General Ledgers", "Yes — see Tab 4"),
            ("Asset Schedule Template", "Yes (no fixed assets) — see Tab 5"),
            ("Payroll Report and Filings", "N/A (no W-2 employees)"),
            ("Last Filed Tax Year", "No (first year)"),
            ("Sales Tax Filings", "N/A"),
            ("Estimated State Tax Payments", "No (unless made — confirm)"),
            ("First time filing with Monily", "Yes"),
        ]),
        ("PARTNERS INFORMATION", [
            ("Number of Share Holders (Partners)", "3 (Nairne, Raj Duggal, Phil)"),
            ("Change of business name during year", "No"),
            ("Calendar year filer", "Yes"),
            ("Foreign account interest/signature authority", "⚠️ Verify — depends on crypto exchange custody"),
            ("Any shareholder a disregarded entity / trust / S-corp", "No (members are individuals — confirm)"),
            ("Owns 20%+ of foreign/domestic corp", "No"),
            ("Outstanding restricted stock", "No"),
            ("Outstanding stock options/warrants", "No"),
            ("Distribution of property or transfer of shareholder interest", "No (Phil → Alec was 2026, not 2025)"),
            ("Accessibility expenses", "No"),
            ("FICA on tips above min wage", "No (no W-2 employees)"),
            ("Own residential rental buildings (low-income housing)", "No"),
            ("R&D expenditures during year", "No"),
        ]),
        ("SIGNATURE", [
            ("Taxpayer Sign", "Nairne (Managing Member)"),
            ("Taxpayer Title", "Managing Member"),
            ("Date", date.today().isoformat()),
        ]),
    ]

    for section_title, fields in sections:
        r = write_section(ws, r, section_title, 3)
        for label, value in fields:
            ws.cell(row=r, column=1, value=label).font = Font(bold=True)
            ws.cell(row=r, column=2, value=value)
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
            ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True)
            if "⚠️" in str(value):
                for c in range(1, 4):
                    ws.cell(row=r, column=c).fill = WARN_FILL
            r += 1
        r += 1

    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 35


# ---------------------------------------------------------------------------
# Cover memo (markdown)
# ---------------------------------------------------------------------------

def build_cover_memo(T):
    md = f"""# Monily Tax Organizer — Cover Memo
## Armada Prime Tech LLC — Tax Year 2025

**Prepared:** {date.today().isoformat()}
**Period of operations:** Aug 1, 2025 – Dec 31, 2025 (entity formed at Armada Prime LLP relaunch)
**Filing status:** First year of filing
**Tax classification:** Multi-member LLC, partnership for tax purposes

---

## What's in this package

| File | Purpose |
|---|---|
| `Monily_Tax_Package_2025.xlsx` | **Single combined workbook** with 8 tabs (Cover & Summary, P&L, Balance Sheet, GL, Asset Schedule, K-1 Partners, 1099 Recipients, Tax Organizer Answers) |
| `05_Cover_Memo_for_Monily.md` / `.docx` | This narrative cover memo |
| `06_Bonus_Detailed_Workbook.xlsx` | Standalone 15-tab reconciliation backup |

---

## Headline Numbers (2025) — Tax-Reclass Basis

| Line | Amount |
|---|---:|
| **Revenue** (Performance Fees from Armada Prime LLP) | ${T['revenue']:,.2f} |
| Less: Direct Costs (1099 contractor commissions) | (${T['contractor_total']:,.2f}) |
| Less: Operating Expenses (after reclass) | (${T['op_total']:,.2f}) |
| **PARTNERSHIP NET INCOME** | **${T['net_income']:,.2f}** |

Tax-reclass adjustments: $4,275 SPV Loan moved to Balance Sheet; $10,500 Insurance pro-rated to Prepaid Asset (only $7,500 of the $18K Dec premium hits 2025 P&L for the 5-month operating period).

---

## Member Structure & K-1 Allocation (3 Partners) — Per Nairne 2026-05-07

The 59.5% Fund Mgmt slice is split 50/50 between **AJ Affleck (1099 contractor)** and **Nairne (K-1 partner)**.

| Member | Economic Interest | Ownership % | Cash Distributions | K-1 Allocated Net Income |
|---|---|---:|---:|---:|
| **Nairne** | 30.25% (50%×Fund Mgmt + 0.5% direct) | {T['nairne_pct']*100:.2f}% | ${T['nairne_cash']:,.2f} | ${T['net_income'] * T['nairne_pct']:,.2f} |
| **Raj Duggal** | 0.50% (direct) | {T['raj_pct']*100:.2f}% | ${T['raj_cash']:,.2f} | ${T['net_income'] * T['raj_pct']:,.2f} |
| **Phil** | 0.50% (direct) | {T['phil_pct']*100:.2f}% | ${T['phil_cash']:,.2f} | ${T['net_income'] * T['phil_pct']:,.2f} |

Total partner economic interest = 31.25% of pre-distribution gross. Normalized partner ownership: Nairne 96.80%, Raj 1.60%, Phil 1.60%.

---

## 1099-NEC Recipients (Aug–Dec 2025 totals)

| Recipient | 2025 Total | 1099 Required? |
|---|---:|---|
| Alec Atkinson | ${T['year_totals'].get('Alec Atkinson', 0):,.2f} | YES (>$600) |
| **AJ Affleck** | **${T['year_totals'].get('AJ Affleck', 0):,.2f}** | **YES — includes 50%×Fund Mgmt** |
| Jake Gordon | ${T['year_totals'].get('Jake Gordon', 0):,.2f} | YES (>$600) |
| Chris (operating) | $11,500.00 | YES (>$600) |
| Issac Morris | ${T['year_totals'].get('Issac', 0):,.2f} | YES (>$600) |
| Nikki | ${T['year_totals'].get('Nikki', 0):,.2f} | NO — under $600 |
| Luke Affleck | ${T['year_totals'].get('Luke', 0):,.2f} | NO — under $600 |
| **Total Contractor Expense** | **${T['contractor_total'] + 11500:,.2f}** | |

**Phil is NOT on the 1099 list** — Phil is a K-1 partner. See K-1 Partners tab.

**Threshold rules:**
- 1099-NEC: required when payment ≥$600/year per recipient. Luke + Nikki are below — no 1099 issuance, but expenses still flow through P&L.
- K-1: NO threshold. All partners (Nairne, Raj, Phil) receive K-1s every year regardless of dollar amount.

---

## Operating Expenses Summary (Tax-Reclass)

| Category | Amount | Notes |
|---|---:|---|
| Insurance (D&O, pro-rated 5/12) | $7,500.00 | Remaining $10,500 → Prepaid Asset on Balance Sheet |
| Chris (contractor labor) | $4,000.00 | Dec only; 1099-NEC required |
| Consulting | $7,500.00 | Nov only |
| PVD (Tech/Ads) | $6,000.00 | Aug only; verify 1099 obligation |
| Website | $7,500.00 | Oct only |
| Alpha Verification | $2,250.00 | Nov only |
| TPA Admin Fees (Formidium) | $5,700.00 | $600 Nov + $5,100 Dec |
| **Total Operating Expenses** | **${T['op_total']:,.2f}** | |
| (Reclassified to Balance Sheet) | $4,275.00 | 506c SPV Loan → loan receivable |

---

## Methodology

### Revenue Recognition
- Source: TPA (Formidium) Reporting Packages, "Performance Fees Crystallized" line, monthly Aug–Dec 2025
- TruQuant's 18% upstream cut and the August "Trader & Developer" / "Spydr" amounts are EXCLUDED — they belong upstream of this entity per a 2026-04-30 policy decision

### Distribution Tracking
- Source: Internal "Distributions Armada Tech 2025 (INTERNAL ONLY)" ledger
- All payments are CASH BASIS — what was actually disbursed each month
- Per-recipient amounts are NET (already after weighted costs / Coinbase fees)

### Member Structure (Per Nairne 2026-05-07)
- **Nairne**: K-1 partner. Receives 50%×Fund Mgmt 59.5% + 0.5% direct = **30.25% economic interest**. Normalized ownership: 96.80%.
- **AJ Affleck**: 1099 CONTRACTOR (NOT a partner). Receives 50%×Fund Mgmt 59.5% + her 39%-pool share. Total $37,139.89 in 2025.
- **Raj Duggal**: K-1 partner. 0.5% direct. Normalized ownership: 1.60%.
- **Phil**: K-1 partner. 0.5% direct. Normalized ownership: 1.60%. (Phil held the slice all of 2025; Alec replaced him in April 2026.)
- Total K-1 partner economic interest: 31.25%

### Accounting Method
- **ACCRUAL basis** elected for tax minimization (recognize Aug-Dec deductions in 2025 even if some cash settled in early 2026).

---

## Open Items the Accountant Will Need

1. **EIN** — from IRS letter (CP-575 / 147C)
2. **Formation documents** — Articles of Organization for the LLC
3. **Registered address + state of formation**
4. **Business phone + email**
5. **Member SSNs + addresses** for Nairne, Raj, AND Phil (3 K-1s)
6. **1099 recipient SSN/EIN + addresses** for Alec, Jake, AJ, Issac, Luke, Chris, plus PVD/Ad Spend/Website vendors crossing $600
7. **Bank/wallet statements** as of 12/31/2025 to populate Balance Sheet Cash line
8. **Initial member capital contributions** for all 3 members at formation
9. **Confirm cash basis vs accrual basis** for tax reporting
10. **Foreign financial account question** — verify crypto wallets/exchanges
11. **506c SPV structure** — confirm whether the $4,275 is loan or equity investment
12. **Insurance pro-ration** — confirm $18K Dec is annual D&O

---

## Reproduce This Package

```bash
python tools/build_monily_package.py
python tools/md_to_docx.py     # to regenerate the .docx
```
"""
    OUT_MD.write_text(md)
    print(f"Wrote {OUT_MD}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    T = compute_year_totals()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    build_cover_tab(wb, T)
    build_pnl_tab(wb, T)
    build_balance_sheet_tab(wb, T)
    build_gl_tab(wb, T)
    build_asset_tab(wb, T)
    build_k1_tab(wb, T)
    build_1099_tab(wb, T)
    build_organizer_tab(wb, T)

    wb.save(OUT_XLSX)
    print(f"Wrote {OUT_XLSX}")

    build_cover_memo(T)
    print(f"\nDone. Package files in {OUT_DIR}/")
