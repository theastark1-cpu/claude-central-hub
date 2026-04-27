#!/usr/bin/env python3
"""Build consultant-splits dynamic Excel + dashboard JSON from a TPA Reporting Package.

Reads:
  1. The TPA reporting package xlsx (per-investor performance fees = the GP pool).
  2. The internal Monthly Return xlsx (IDS sheet → investor → consultant mapping).

Writes:
  1. <output_xlsx>: a self-contained workbook with live formulas (named cells for split %,
     xlookups for consultant attribution, sumifs for consultant aggregation).
  2. <repo>/data/consultant_splits.json: month-keyed snapshot for the dashboard.

Reconciliation insight (March 2026):
  - TPA "Gross Income" $404,850.79  = the Fund's 82% cut (post-TruQuant).
  - TPA "Performance Fees Crystallized" $121,404.24 = GP Cut (30% of Fund cut).
  - TPA "Net to Investors" $283,276.56 = Investor Cut (70% of Fund cut).
  TruQuant's 18% is taken upstream and never enters Armada Prime's books.

Each TPA investor's perf_fee column is therefore the per-investor GP pool. We attribute
that pool to the consultant who raised the investor (per IDS), then split it:
  Fund Mgmt:  59.5%
  Consultant: 39.0%
  Raj:         0.5%
  Nairne:      0.5%
  Phil:        0.5%

Usage:
  python tools/build_consultant_splits.py <tpa_xlsx> [--internal <internal_xlsx>]
                                                     [--output-xlsx <path>]
                                                     [--label "Mar 2026"]
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

sys.path.insert(0, str(Path(__file__).resolve().parent))
from parse_tpa_report import parse_workbook  # noqa: E402

REPO_ROOT = Path(__file__).resolve().parent.parent
JSON_PATH = REPO_ROOT / "data" / "consultant_splits.json"
DEFAULT_INTERNAL = Path("/Users/nairne/Downloads/March 2026 Monthly Return (1).xlsx")
DEFAULT_TPA = Path(
    "/Users/nairne/Downloads/Reporting Package - March 2026_v2_final/"
    "Armada_Prime_LLP_Reporting Package_2026-03-31_1777286642911.xlsx"
)
DEFAULT_OUTPUT_XLSX = REPO_ROOT / "Armada_Consultant_Splits.xlsx"

SPLIT_PCTS = {
    "fund_mgmt": 0.595,
    "consultant": 0.39,
    "raj": 0.005,
    "nairne": 0.005,
    "alec": 0.005,  # Alec is the new GP (was Phil before 2026-04-27)
}

# Confirmed by user 2026-04-27. Augments / overrides the internal IDS sheet.
# Fund Hub SPV = sub-LP container with mixed AJ/Alec investors; will be split out
# in a separate buildout. For now its 39% portion is held in "Fund Hub SPV
# (Pending Split)" so it stays visible.
CONSULTANT_OVERRIDES = {
    "14-Class B-1027-1": "Fund Hub SPV (Pending Split)",  # Fund Hub Investments LLC
    "14-Class B-1076-1": "Alec Atkinson",                  # Philippe Henriques
    "14-Class B-1073-1": "Alec Atkinson",                  # PGJCHoldings LLC
    "14-Class B-1072-1": "Alec Atkinson",                  # Mashirito LLC
    "14-Class B-1068-1": "Alec Atkinson",                  # Weston Shea Christensen
}

# Default Payor Group per cost line item, used when the internal Costs sheet
# leaves column D empty. Confirmed by user 2026-04-27 from their existing
# internal cost sheet. Edit when adding new line item types.
#   "TQ/Armada"     → split among everyone receiving income (incl. TruQuant)
#                     in proportion to each party's income.
#   "Armada"        → split among Armada-side income recipients only
#                     (excludes TruQuant). Pro-rata by income.
#   "Fund Management" → 100% paid by Fund Mgmt. No consultant or fixed-GP burden.
PAYOR_GROUP_DEFAULTS = {
    "Chris": "TQ/Armada",
    "TPA": "Armada",
    "Chris (Hotel+Ticket)": "TQ/Armada",
    "Charalece": "Armada",
    "Alec (Ticket)": "Fund Management",
    "Alec (Hotel)": "Fund Management",
    "Taxes": "Armada",
    "Legal": "Armada",
}

VALID_PAYOR_GROUPS = ("TQ/Armada", "Armada", "Fund Management")

# Seed distributions for the Distributions tab. Each entry tracks one cash
# outflow from the GP pool. "Type" is either "Expense" (cash to a vendor —
# doesn't count toward any party's payout) or "Payout" (cash to a specific
# party — counts toward their settlement).
# Edit the Distributions sheet directly for new transfers; the Per-Party
# Settlement section uses SUMIFS so Outstanding updates automatically.
DISTRIBUTIONS_DEFAULT = [
    {
        "date": None,
        "type": "Expense",
        "party": "Expenses",
        "description": "Royal KKS Collective + Atkinson Group + Chris March (per app screenshot)",
        "amount": 56026.00,
        "method": "Wire",
        "status": "Completed",
        "notes": "Includes Feb/March reimbursements — Nairne, Chris hotel/ticket, Charalece, Alec NYC, Taxes, Chris",
    },
    {
        "date": None,
        "type": "Payout",
        "party": "Fund Mgmt",
        "description": "Management cut (partial)",
        "amount": 30604.06,
        "method": "Wire",
        "status": "Completed",
        "notes": "",
    },
    {
        "date": None,
        "type": "Payout",
        "party": "AJ Affleck",
        "description": "Consultant cut (partial)",
        "amount": 937.76,
        "method": "Wire",
        "status": "Completed",
        "notes": "",
    },
]


# ---------------------------------------------------------------------------
# IDS loader
# ---------------------------------------------------------------------------

def load_ids(internal_xlsx: Path) -> dict[str, dict]:
    """Return {tpa_id: {name, consultant, position_id}} from the internal IDS sheet,
    with overrides applied. Skips rows missing TPA ID or consultant.
    """
    wb = openpyxl.load_workbook(internal_xlsx, data_only=True)
    ws = wb["IDS"]
    out: dict[str, dict] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        name, tpa_id, pos_id, consultant = row[0], row[1], row[2], row[3]
        if not tpa_id:
            continue
        out[str(tpa_id).strip()] = {
            "name": name.strip() if isinstance(name, str) else name,
            "consultant": consultant.strip() if isinstance(consultant, str) else consultant,
            "position_id": pos_id,
        }
    for tpa_id, cons in CONSULTANT_OVERRIDES.items():
        if tpa_id in out:
            out[tpa_id]["consultant"] = cons
        else:
            out[tpa_id] = {"name": None, "consultant": cons, "position_id": None}
    return out


# ---------------------------------------------------------------------------
# Costs loader
# ---------------------------------------------------------------------------

def load_costs(internal_xlsx: Path) -> list[dict]:
    """Read the Costs sheet from the internal Monthly Return file.

    Returns a list of {name, amount, payor_group} dicts (excludes the TOTAL row).
    Payor Group reads column D if populated; otherwise falls back to
    PAYOR_GROUP_DEFAULTS keyed by name. Raises if a name has no default.
    """
    wb = openpyxl.load_workbook(internal_xlsx, data_only=True)
    ws = wb["Costs"]
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        name = row[0]
        amount = row[1] if len(row) > 1 else None
        payor_group = row[3] if len(row) > 3 else None
        if not isinstance(name, str):
            continue
        name = name.strip()
        if not name or name.lower() == "total":
            break
        if not isinstance(amount, (int, float)):
            continue
        if isinstance(payor_group, str):
            payor_group = payor_group.strip()
        if not payor_group:
            payor_group = PAYOR_GROUP_DEFAULTS.get(name)
        if payor_group not in VALID_PAYOR_GROUPS:
            raise ValueError(
                f"Cost {name!r}: payor_group {payor_group!r} not in {VALID_PAYOR_GROUPS}. "
                "Either fill column D in the internal Costs sheet or add a default to "
                "PAYOR_GROUP_DEFAULTS in tools/build_consultant_splits.py"
            )
        out.append({"name": name, "amount": float(amount), "payor_group": payor_group})
    return out


def compute_tq_income(tpa_record: dict) -> float:
    """TruQuant takes 18% of TRUE gross, upstream of Armada Prime's books.

    The TPA's 'Total Income' is Armada's 82% cut. Inverse-grossed:
        TQ_income = total_income * 18 / 82
    For March 2026: $404,850.79 * 18/82 = $88,869.69
    """
    fund_total_income = tpa_record.get("income_statement", {}).get("total_income", 0) or 0
    return fund_total_income * 0.18 / 0.82


def build_income_map(records: list[dict], tpa_record: dict) -> dict[str, float]:
    """Each party's income for the period. Used as weighting denominator
    when allocating costs.
    """
    consultant_totals: dict[str, float] = {}
    for r in records:
        consultant_totals.setdefault(r["consultant"], 0.0)
        consultant_totals[r["consultant"]] += r["perf_fee"] * SPLIT_PCTS["consultant"]
    total_perf = sum(r["perf_fee"] for r in records)
    income = dict(consultant_totals)
    income["Fund Mgmt"] = total_perf * SPLIT_PCTS["fund_mgmt"]
    income["Raj (GP fixed 0.5%)"] = total_perf * SPLIT_PCTS["raj"]
    income["Nairne (GP fixed 0.5%)"] = total_perf * SPLIT_PCTS["nairne"]
    income["Alec (GP fixed 0.5%)"] = total_perf * SPLIT_PCTS["alec"]
    income["TruQuant"] = compute_tq_income(tpa_record)
    return income


def allocate_cost(cost: dict, income: dict[str, float]) -> dict[str, float]:
    """Returns {party: amount} that this single cost is split into."""
    pg = cost["payor_group"]
    amount = cost["amount"]
    if pg == "Fund Management":
        return {"Fund Mgmt": amount}
    if pg == "Armada":
        # Exclude TruQuant; split among Armada participants by income share
        armada = {p: i for p, i in income.items() if p != "TruQuant" and i > 0}
        total = sum(armada.values())
        if total == 0:
            return {}
        return {p: amount * i / total for p, i in armada.items()}
    if pg == "TQ/Armada":
        nz = {p: i for p, i in income.items() if i > 0}
        total = sum(nz.values())
        if total == 0:
            return {}
        return {p: amount * i / total for p, i in nz.items()}
    raise ValueError(f"Unknown payor_group: {pg!r}")


# ---------------------------------------------------------------------------
# Reconciliation
# ---------------------------------------------------------------------------

def reconcile(tpa_investors: list[dict], ids_map: dict[str, dict]) -> list[dict]:
    """Return per-investor rows with consultant attribution and reconciliation flag."""
    out = []
    for inv in tpa_investors:
        tpa_id = str(inv["investor_no"]).strip()
        ids_row = ids_map.get(tpa_id, {})
        consultant = ids_row.get("consultant") or "Unmapped"
        out.append({
            "tpa_id": tpa_id,
            "name": inv["name"],
            "consultant": consultant,
            "position_id": ids_row.get("position_id"),
            "begin_balance": inv.get("begin_balance", 0) or 0,
            "ending_balance": inv.get("ending_balance", 0) or 0,
            "gross_profit": inv.get("gross_profit", 0) or 0,
            "perf_fee": inv.get("perf_fee", 0) or 0,
            "additions": inv.get("additions", 0) or 0,
            "withdrawals": inv.get("withdrawals", 0) or 0,
        })
    return out


# ---------------------------------------------------------------------------
# Excel builder (formulas, not values)
# ---------------------------------------------------------------------------

ACCENT = "1f6feb"
HEADER_FILL = PatternFill("solid", fgColor="1a2236")
HEADER_FONT = Font(bold=True, color="e2e8f0", size=10)
TITLE_FONT = Font(bold=True, color="ffffff", size=12)
INPUT_FILL = PatternFill("solid", fgColor="d9ead3")
TOTAL_FILL = PatternFill("solid", fgColor="334155")
OK_FILL = PatternFill("solid", fgColor="14532d")
WARN_FILL = PatternFill("solid", fgColor="7f1d1d")
BORDER = Border(*(Side(style="thin", color="2a3550"),) * 4)


def _style_header(cell):
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.border = BORDER
    cell.alignment = Alignment(horizontal="left", vertical="center")


def _autosize(ws, min_w: int = 10, max_w: int = 48):
    for col_cells in ws.columns:
        col = col_cells[0].column_letter
        widest = max((len(str(c.value)) if c.value is not None else 0 for c in col_cells), default=0)
        ws.column_dimensions[col].width = max(min_w, min(max_w, widest + 2))


def build_workbook(records: list[dict], period_label: str, output_path: Path,
                   costs: list[dict], tpa_record: dict) -> dict:
    """Write the dynamic Excel. Returns summary dict."""
    income_map = build_income_map(records, tpa_record)
    # cost_alloc[i] = {party: $} for cost i (same order as `costs`)
    cost_alloc = [allocate_cost(c, income_map) for c in costs]
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # --------- 1. Inputs ---------
    ws_in = wb.create_sheet("Inputs")
    ws_in["A1"] = "Armada Prime LLP — Consultant Splits"
    ws_in["A1"].font = TITLE_FONT
    ws_in.merge_cells("A1:F1")
    ws_in["A2"] = f"Period: {period_label}"
    ws_in["A2"].font = Font(italic=True, color="94a3b8")

    # Named cells for split percentages
    ws_in["A4"] = "GP Pool Split %"
    ws_in["A4"].font = Font(bold=True, color="e2e8f0")
    pct_rows = [
        ("Fund Mgmt", "GP_FundMgmt_Pct", SPLIT_PCTS["fund_mgmt"]),
        ("Consultant", "GP_Consultant_Pct", SPLIT_PCTS["consultant"]),
        ("Raj", "GP_Raj_Pct", SPLIT_PCTS["raj"]),
        ("Nairne", "GP_Nairne_Pct", SPLIT_PCTS["nairne"]),
        ("Alec (GP)", "GP_Alec_Pct", SPLIT_PCTS["alec"]),
    ]
    for i, (label, name, val) in enumerate(pct_rows, start=5):
        ws_in.cell(row=i, column=1, value=label)
        c = ws_in.cell(row=i, column=2, value=val)
        c.fill = INPUT_FILL
        c.number_format = "0.00%"
        wb.defined_names[name] = DefinedName(name=name, attr_text=f"Inputs!$B${i}")

    ws_in["A11"] = "Total split (sanity)"
    ws_in["B11"] = "=SUM(B5:B9)"
    ws_in["B11"].number_format = "0.00%"
    ws_in["B11"].font = Font(bold=True, color="10b981")

    ws_in["A13"] = "Expense Pool"
    ws_in["A13"].font = Font(bold=True, color="e2e8f0")
    ws_in["A14"] = "Operating expenses (live-linked to Costs sheet)"
    ws_in["B14"] = "=Costs_Total"  # named range defined when Costs sheet is built
    ws_in["B14"].number_format = '"$"#,##0.00'
    ws_in["B14"].font = Font(bold=True, color="10b981")
    wb.defined_names["Expense_Pool"] = DefinedName(name="Expense_Pool", attr_text="Inputs!$B$14")

    # TruQuant income — 18% of TRUE gross. The TPA shows Armada's 82% cut,
    # so TQ_Income = TPA_total_income * 18/82. Hardcoded value (computed in
    # Python from the TPA package); used by the Consultant Summary's TQ row.
    ws_in["A16"] = "TruQuant Income (upstream)"
    ws_in["A16"].font = Font(bold=True, color="e2e8f0")
    ws_in["A17"] = "TQ takes 18% before Armada's books"
    tq_income = compute_tq_income(tpa_record)
    ws_in["B17"] = round(tq_income, 2)
    ws_in["B17"].number_format = '"$"#,##0.00'
    ws_in["B17"].fill = INPUT_FILL
    wb.defined_names["TQ_Income"] = DefinedName(name="TQ_Income", attr_text="Inputs!$B$17")

    ws_in["A19"] = "Notes"
    ws_in["A19"].font = Font(bold=True, color="e2e8f0")
    ws_in["A20"] = (
        "TPA per-investor performance-fee crystallized = each investor's GP pool. "
        "Consultant attribution flows from IDS → Per-Investor Allocation → Consultant Summary. "
        "Cost line items live on the Costs sheet — edit there and amounts re-flow to Consultant Summary."
    )
    ws_in["A20"].alignment = Alignment(wrap_text=True, vertical="top")
    ws_in.merge_cells("A20:F22")

    ws_in.column_dimensions["A"].width = 38
    ws_in.column_dimensions["B"].width = 18

    # --------- 2. IDS Mapping ---------
    ws_ids = wb.create_sheet("IDS Mapping")
    ids_headers = ["TPA ID", "Investor Name", "Position ID", "Consultant", "Source"]
    for i, h in enumerate(ids_headers, start=1):
        c = ws_ids.cell(row=1, column=i, value=h)
        _style_header(c)

    seen_ids: set[str] = set()
    for r in records:
        seen_ids.add(r["tpa_id"])
    # All IDS rows (mapped + unmapped + overrides) — so we can xlookup safely
    for i, rec in enumerate(records, start=2):
        ws_ids.cell(row=i, column=1, value=rec["tpa_id"])
        ws_ids.cell(row=i, column=2, value=rec["name"])
        ws_ids.cell(row=i, column=3, value=rec["position_id"])
        ws_ids.cell(row=i, column=4, value=rec["consultant"])
        src = "TPA + IDS"
        if rec["tpa_id"] in CONSULTANT_OVERRIDES:
            src = "Override (confirmed 2026-04-27)"
        if rec["consultant"] == "Unmapped":
            src = "UNMAPPED"
        ws_ids.cell(row=i, column=5, value=src)

    last_id_row = len(records) + 1
    wb.defined_names["IDS_TPA"] = DefinedName(
        name="IDS_TPA",
        attr_text=f"'IDS Mapping'!$A$2:$A${last_id_row}",
    )
    wb.defined_names["IDS_Consultant"] = DefinedName(
        name="IDS_Consultant",
        attr_text=f"'IDS Mapping'!$D$2:$D${last_id_row}",
    )
    _autosize(ws_ids)

    # --------- 3. Per-Investor Allocation ---------
    ws_p = wb.create_sheet("Per-Investor Allocation")
    p_headers = [
        "TPA ID", "Investor", "Consultant",
        "Begin Balance", "Ending Balance", "Gross P&L", "Perf Fee (GP Pool)",
        "Fund Mgmt", "Consultant Cut", "Raj", "Nairne", "Alec (GP)",
        "Sum Check",
    ]
    for i, h in enumerate(p_headers, start=1):
        c = ws_p.cell(row=1, column=i, value=h)
        _style_header(c)

    for i, rec in enumerate(records, start=2):
        ws_p.cell(row=i, column=1, value=rec["tpa_id"])
        ws_p.cell(row=i, column=2, value=rec["name"])
        # Plain string value (not XLOOKUP) so SUMIFs match reliably across all
        # Excel versions. The IDS Mapping sheet is the authoritative source —
        # rerun the build script after editing it. To override a single row,
        # edit this cell directly.
        ws_p.cell(row=i, column=3, value=rec["consultant"])
        ws_p.cell(row=i, column=4, value=rec["begin_balance"]).number_format = '"$"#,##0.00'
        ws_p.cell(row=i, column=5, value=rec["ending_balance"]).number_format = '"$"#,##0.00'
        ws_p.cell(row=i, column=6, value=rec["gross_profit"]).number_format = '"$"#,##0.00'
        c_pf = ws_p.cell(row=i, column=7, value=rec["perf_fee"])
        c_pf.number_format = '"$"#,##0.00'
        c_pf.fill = INPUT_FILL
        # Split formulas reference the named cells
        ws_p.cell(row=i, column=8, value=f"=G{i}*GP_FundMgmt_Pct").number_format = '"$"#,##0.00'
        ws_p.cell(row=i, column=9, value=f"=G{i}*GP_Consultant_Pct").number_format = '"$"#,##0.00'
        ws_p.cell(row=i, column=10, value=f"=G{i}*GP_Raj_Pct").number_format = '"$"#,##0.00'
        ws_p.cell(row=i, column=11, value=f"=G{i}*GP_Nairne_Pct").number_format = '"$"#,##0.00'
        ws_p.cell(row=i, column=12, value=f"=G{i}*GP_Alec_Pct").number_format = '"$"#,##0.00'
        ws_p.cell(row=i, column=13,
                  value=f"=ROUND(SUM(H{i}:L{i})-G{i}, 2)").number_format = '"$"#,##0.00'

    last_inv_row = len(records) + 1
    total_row = last_inv_row + 1
    ws_p.cell(row=total_row, column=2, value="TOTAL").font = Font(bold=True)
    for col in range(4, 13):
        cl = get_column_letter(col)
        cell = ws_p.cell(row=total_row, column=col,
                         value=f"=SUM({cl}2:{cl}{last_inv_row})")
        cell.number_format = '"$"#,##0.00'
        cell.font = Font(bold=True)
        cell.fill = TOTAL_FILL
    ws_p.cell(row=total_row, column=13,
              value=f"=ROUND(SUM(H{total_row}:L{total_row})-G{total_row}, 2)").number_format = '"$"#,##0.00'

    wb.defined_names["PI_Consultant"] = DefinedName(
        name="PI_Consultant",
        attr_text=f"'Per-Investor Allocation'!$C$2:$C${last_inv_row}",
    )
    wb.defined_names["PI_PerfFee"] = DefinedName(
        name="PI_PerfFee",
        attr_text=f"'Per-Investor Allocation'!$G$2:$G${last_inv_row}",
    )
    wb.defined_names["PI_ConsultantCut"] = DefinedName(
        name="PI_ConsultantCut",
        attr_text=f"'Per-Investor Allocation'!$I$2:$I${last_inv_row}",
    )
    wb.defined_names["PI_FundMgmt"] = DefinedName(
        name="PI_FundMgmt",
        attr_text=f"'Per-Investor Allocation'!$H$2:$H${last_inv_row}",
    )
    wb.defined_names["PI_Capital"] = DefinedName(
        name="PI_Capital",
        attr_text=f"'Per-Investor Allocation'!$E$2:$E${last_inv_row}",
    )
    _autosize(ws_p)

    # --------- 4. Consultant Summary ---------
    # Two sections:
    #   1. CONSULTANTS — only people who raised capital. Their numbers come from
    #      SUMIF on Per-Investor!C (which is now plain text, so SUMIFs work in
    #      every Excel version).
    #   2. GP POOL — OTHER RECIPIENTS — Fund Mgmt + Raj/Nairne/Alec(GP) fixed.
    #      These rows do NOT have investors or capital raised — Fund Mgmt's
    #      59.5% is GP-pool-wide, not tied to specific investors.
    consultant_names = sorted({r["consultant"] for r in records if r["consultant"] != "Unmapped"})
    if any(r["consultant"] == "Unmapped" for r in records):
        consultant_names.append("Unmapped")

    ws_c = wb.create_sheet("Consultant Summary")
    c_headers = [
        "Consultant", "# Investors", "Capital Raised", "Gross GP",
        "% of GP Pool", "Weighted Expense", "Net Profit",
    ]
    for i, h in enumerate(c_headers, start=1):
        c = ws_c.cell(row=1, column=i, value=h)
        _style_header(c)

    # Section 1: consultants
    for i, name in enumerate(consultant_names, start=2):
        ws_c.cell(row=i, column=1, value=name)
        ws_c.cell(row=i, column=2,
                  value=f'=COUNTIF(PI_Consultant, A{i})')
        ws_c.cell(row=i, column=3,
                  value=f'=SUMIF(PI_Consultant, A{i}, PI_Capital)').number_format = '"$"#,##0.00'
        ws_c.cell(row=i, column=4,
                  value=f'=SUMIF(PI_Consultant, A{i}, PI_ConsultantCut)').number_format = '"$"#,##0.00'

    last_consultant_row = len(consultant_names) + 1
    consultant_subtotal_row = last_consultant_row + 1

    # Subtotal row for consultants
    ws_c.cell(row=consultant_subtotal_row, column=1, value="CONSULTANTS SUBTOTAL").font = Font(bold=True, color="a78bfa")
    for col_letter in ("B", "C", "D"):
        cell = ws_c.cell(row=consultant_subtotal_row, column={'B':2,'C':3,'D':4}[col_letter])
        cell.value = f"=SUM({col_letter}2:{col_letter}{last_consultant_row})"
        cell.font = Font(bold=True)
        cell.fill = TOTAL_FILL
        if col_letter != "B":
            cell.number_format = '"$"#,##0.00'

    # Spacer row, then Section 2
    section_label_row = consultant_subtotal_row + 2
    ws_c.cell(row=section_label_row, column=1,
              value="GP POOL — OTHER RECIPIENTS").font = Font(bold=True, color="93c5fd")

    other_start = section_label_row + 1

    # Fund Mgmt — 59.5% of pool, no investor attribution
    fm_row = other_start
    ws_c.cell(row=fm_row, column=1, value="Fund Mgmt").font = Font(italic=True, color="93c5fd")
    ws_c.cell(row=fm_row, column=2, value="—")
    ws_c.cell(row=fm_row, column=3, value="—")
    ws_c.cell(row=fm_row, column=4, value=f'=SUM(PI_PerfFee)*GP_FundMgmt_Pct').number_format = '"$"#,##0.00'

    # Fixed partners (Raj / Nairne / Alec GP)
    fixed_specs = [
        ("Raj (GP fixed 0.5%)", "GP_Raj_Pct"),
        ("Nairne (GP fixed 0.5%)", "GP_Nairne_Pct"),
        ("Alec (GP fixed 0.5%)", "GP_Alec_Pct"),
    ]
    for j, (label, pct_name) in enumerate(fixed_specs, start=1):
        r = fm_row + j
        ws_c.cell(row=r, column=1, value=label).font = Font(italic=True, color="d8b4fe")
        ws_c.cell(row=r, column=2, value="—")
        ws_c.cell(row=r, column=3, value="—")
        ws_c.cell(row=r, column=4,
                  value=f'=SUM(PI_PerfFee)*{pct_name}').number_format = '"$"#,##0.00'

    last_other_row = fm_row + len(fixed_specs)

    # GP Pool grand total — at the very bottom, equals the perf fee pool
    grand_total_row = last_other_row + 1
    ws_c.cell(row=grand_total_row, column=1, value="GP POOL TOTAL").font = Font(bold=True)
    ws_c.cell(row=grand_total_row, column=4,
              value=f"=SUM(D2:D{last_consultant_row})+SUM(D{fm_row}:D{last_other_row})").number_format = '"$"#,##0.00'

    pool_cell = f"$D${grand_total_row}"

    # Weighted Expense column → INDEX/MATCH against the per-party totals row
    # in the Costs sheet matrix (named ranges Costs_PartyTotals + Costs_PartyHeaders
    # are defined when the Costs sheet is built later).
    fillable_rows = list(range(2, last_consultant_row + 1)) + list(range(fm_row, last_other_row + 1))
    for r in fillable_rows:
        ws_c.cell(row=r, column=5, value=f"=IFERROR(D{r}/{pool_cell},0)").number_format = "0.00%"
        ws_c.cell(row=r, column=6,
                  value=f'=IFERROR(INDEX(Costs_PartyTotals, MATCH(A{r}, Costs_PartyHeaders, 0)), 0)').number_format = '"$"#,##0.00'
        ws_c.cell(row=r, column=7, value=f"=D{r}-F{r}").number_format = '"$"#,##0.00'

    # Subtotal row for consultants
    ws_c.cell(row=consultant_subtotal_row, column=5,
              value=f"=IFERROR(D{consultant_subtotal_row}/{pool_cell},0)").number_format = "0.00%"
    ws_c.cell(row=consultant_subtotal_row, column=5).font = Font(bold=True)
    ws_c.cell(row=consultant_subtotal_row, column=5).fill = TOTAL_FILL
    ws_c.cell(row=consultant_subtotal_row, column=6,
              value=f"=SUM(F2:F{last_consultant_row})").number_format = '"$"#,##0.00'
    ws_c.cell(row=consultant_subtotal_row, column=6).font = Font(bold=True)
    ws_c.cell(row=consultant_subtotal_row, column=6).fill = TOTAL_FILL
    ws_c.cell(row=consultant_subtotal_row, column=7,
              value=f"=SUM(G2:G{last_consultant_row})").number_format = '"$"#,##0.00'
    ws_c.cell(row=consultant_subtotal_row, column=7).font = Font(bold=True)
    ws_c.cell(row=consultant_subtotal_row, column=7).fill = TOTAL_FILL

    # GP Pool total — Armada-side only ($121,404.22). Excludes TruQuant.
    ws_c.cell(row=grand_total_row, column=5, value=f"=D{grand_total_row}/{pool_cell}").number_format = "0.00%"
    # The GP pool's share of total costs = sum of every Armada party's allocation in Costs
    ws_c.cell(row=grand_total_row, column=6,
              value=f"=SUM(F2:F{last_consultant_row})+SUM(F{fm_row}:F{last_other_row})").number_format = '"$"#,##0.00'
    ws_c.cell(row=grand_total_row, column=7, value=f"=D{grand_total_row}-F{grand_total_row}").number_format = '"$"#,##0.00'
    for col in range(1, 8):
        ws_c.cell(row=grand_total_row, column=col).font = Font(bold=True)
        ws_c.cell(row=grand_total_row, column=col).fill = TOTAL_FILL

    # Section 3: External — TruQuant (18% upstream). Not in the GP pool, but
    # TQ pays its share of any TQ/Armada-classified costs.
    tq_section_row = grand_total_row + 2
    ws_c.cell(row=tq_section_row, column=1,
              value="EXTERNAL — TruQuant (18% upstream of Armada's books)").font = Font(bold=True, color="fbbf24")
    tq_row = tq_section_row + 1
    ws_c.cell(row=tq_row, column=1, value="TruQuant").font = Font(italic=True, color="fbbf24")
    ws_c.cell(row=tq_row, column=2, value="—")
    ws_c.cell(row=tq_row, column=3, value="—")
    # TruQuant income = (TPA total income / 0.82) × 0.18, computed from a
    # named cell defined in Inputs (TQ_Income).
    ws_c.cell(row=tq_row, column=4, value=f"=TQ_Income").number_format = '"$"#,##0.00'
    ws_c.cell(row=tq_row, column=5, value="—")  # TQ isn't part of the GP pool
    ws_c.cell(row=tq_row, column=6,
              value=f'=IFERROR(INDEX(Costs_PartyTotals, MATCH(A{tq_row}, Costs_PartyHeaders, 0)), 0)').number_format = '"$"#,##0.00'
    ws_c.cell(row=tq_row, column=7, value=f"=D{tq_row}-F{tq_row}").number_format = '"$"#,##0.00'

    # Section 4: Grand total (incl. TQ)
    grand_total_with_tq = tq_row + 2
    ws_c.cell(row=grand_total_with_tq, column=1, value="GRAND TOTAL (Armada + TQ)").font = Font(bold=True)
    ws_c.cell(row=grand_total_with_tq, column=4,
              value=f"=D{grand_total_row}+D{tq_row}").number_format = '"$"#,##0.00'
    ws_c.cell(row=grand_total_with_tq, column=6,
              value=f"=F{grand_total_row}+F{tq_row}").number_format = '"$"#,##0.00'
    ws_c.cell(row=grand_total_with_tq, column=7,
              value=f"=D{grand_total_with_tq}-F{grand_total_with_tq}").number_format = '"$"#,##0.00'
    for col in range(1, 8):
        ws_c.cell(row=grand_total_with_tq, column=col).font = Font(bold=True)
        ws_c.cell(row=grand_total_with_tq, column=col).fill = TOTAL_FILL

    last_summary_row = grand_total_with_tq

    _autosize(ws_c)

    # --------- 5. Costs ---------
    # Matrix layout: rows = cost line items, columns = parties.
    # Each cell shows the dollar amount that party pays for that cost,
    # determined by Payor Group:
    #   Fund Management → 100% to Fund Mgmt
    #   Armada          → split among Armada parties (excl. TruQuant) by income
    #   TQ/Armada       → split among everyone (incl. TruQuant) by income
    # The TOTAL row is each party's full expense burden, which Consultant
    # Summary picks up via INDEX/MATCH on the Costs_PartyTotals named range.
    party_order = consultant_names + ["Fund Mgmt", "TruQuant"] + [s[0] for s in fixed_specs]

    ws_costs = wb.create_sheet("Costs")
    ws_costs["A1"] = "Costs — Allocation by Payor Group"
    ws_costs["A1"].font = TITLE_FONT
    ws_costs.merge_cells(f"A1:{get_column_letter(4 + len(party_order))}1")
    ws_costs["A2"] = f"Period: {period_label}"
    ws_costs["A2"].font = Font(italic=True, color="94a3b8")
    ws_costs["A3"] = ("Payor Groups: 'TQ/Armada' = split among everyone (incl. TruQuant) by income share. "
                      "'Armada' = split among Armada parties only (excludes TruQuant) by income. "
                      "'Fund Management' = 100% Fund Mgmt.")
    ws_costs["A3"].font = Font(italic=True, color="94a3b8", size=9)
    ws_costs["A3"].alignment = Alignment(wrap_text=True)
    ws_costs.merge_cells(f"A3:{get_column_letter(4 + len(party_order))}3")

    cost_header_row = 5
    fixed_headers = ["Expense Item", "Total Cost", "Status", "Payor Group"]
    for i, h in enumerate(fixed_headers, start=1):
        c = ws_costs.cell(row=cost_header_row, column=i, value=h)
        _style_header(c)
    for j, party in enumerate(party_order, start=5):
        c = ws_costs.cell(row=cost_header_row, column=j, value=party)
        _style_header(c)

    cost_data_start = cost_header_row + 1
    for i, (item, alloc) in enumerate(zip(costs, cost_alloc), start=cost_data_start):
        ws_costs.cell(row=i, column=1, value=item["name"])
        c = ws_costs.cell(row=i, column=2, value=item["amount"])
        c.number_format = '"$"#,##0.00'
        c.fill = INPUT_FILL
        ws_costs.cell(row=i, column=3, value=item.get("status") or "")
        c = ws_costs.cell(row=i, column=4, value=item["payor_group"])
        c.font = Font(italic=True, color="a78bfa")
        for j, party in enumerate(party_order, start=5):
            amt = alloc.get(party, 0)
            c = ws_costs.cell(row=i, column=j, value=round(amt, 2) if amt else 0)
            c.number_format = '"$"#,##0.00'

    last_cost_row = cost_data_start + len(costs) - 1
    total_row_costs = last_cost_row + 1
    ws_costs.cell(row=total_row_costs, column=1, value="TOTAL").font = Font(bold=True)
    # Total Cost column = sum of all costs (input column)
    c = ws_costs.cell(row=total_row_costs, column=2,
                      value=f"=SUM(B{cost_data_start}:B{last_cost_row})")
    c.number_format = '"$"#,##0.00'
    c.font = Font(bold=True)
    c.fill = TOTAL_FILL
    # Each party column → sum of their allocations across all cost rows
    for j in range(5, 5 + len(party_order)):
        col = get_column_letter(j)
        c = ws_costs.cell(row=total_row_costs, column=j,
                          value=f"=SUM({col}{cost_data_start}:{col}{last_cost_row})")
        c.number_format = '"$"#,##0.00'
        c.font = Font(bold=True)
        c.fill = TOTAL_FILL

    # Named ranges that Consultant Summary uses to look up each party's burden
    first_party_col = get_column_letter(5)
    last_party_col = get_column_letter(4 + len(party_order))
    wb.defined_names["Costs_PartyHeaders"] = DefinedName(
        name="Costs_PartyHeaders",
        attr_text=f"Costs!${first_party_col}${cost_header_row}:${last_party_col}${cost_header_row}",
    )
    wb.defined_names["Costs_PartyTotals"] = DefinedName(
        name="Costs_PartyTotals",
        attr_text=f"Costs!${first_party_col}${total_row_costs}:${last_party_col}${total_row_costs}",
    )
    wb.defined_names["Costs_Total"] = DefinedName(
        name="Costs_Total",
        attr_text=f"Costs!$B${total_row_costs}",
    )

    # Income reference table — shows what we used as the weighting denominator
    # for cost allocation. Editable but: changing these here does NOT propagate
    # because cost allocations were pre-computed in Python. Re-run the script
    # after editing.
    ref_start = total_row_costs + 3
    ws_costs.cell(row=ref_start, column=1,
                  value="Income Reference (used for cost weighting)").font = Font(bold=True, color="e2e8f0")
    for i, h in enumerate(["Party", "Income", "% TQ+Armada", "% Armada-only"], start=1):
        c = ws_costs.cell(row=ref_start + 1, column=i, value=h)
        _style_header(c)
    total_with_tq = sum(income_map.values())
    total_armada = sum(v for k, v in income_map.items() if k != "TruQuant")
    for i, party in enumerate(party_order, start=ref_start + 2):
        inc = income_map.get(party, 0)
        ws_costs.cell(row=i, column=1, value=party)
        c = ws_costs.cell(row=i, column=2, value=round(inc, 2))
        c.number_format = '"$"#,##0.00'
        c = ws_costs.cell(row=i, column=3, value=inc / total_with_tq if total_with_tq else 0)
        c.number_format = "0.00%"
        if party == "TruQuant":
            ws_costs.cell(row=i, column=4, value="—").alignment = Alignment(horizontal="right")
        else:
            c = ws_costs.cell(row=i, column=4, value=inc / total_armada if total_armada else 0)
            c.number_format = "0.00%"

    ws_costs.column_dimensions["A"].width = 30
    for col in range(2, 5 + len(party_order)):
        ws_costs.column_dimensions[get_column_letter(col)].width = 14
    ws_costs.row_dimensions[3].height = 28

    # --------- 6. Distributions ---------
    # Tracks every cash outflow from the GP pool. Three sections:
    #   1. Pool status — total pool, cash distributed, remaining
    #   2. Distributions log — each transfer (Expense or Payout) with details
    #   3. Per-Party Settlement — per party: expected net, paid, outstanding
    ws_dist = wb.create_sheet("Distributions")
    ws_dist["A1"] = "GP Pool Distribution Tracker"
    ws_dist["A1"].font = TITLE_FONT
    ws_dist.merge_cells("A1:I1")
    ws_dist["A2"] = f"Period: {period_label}"
    ws_dist["A2"].font = Font(italic=True, color="94a3b8")

    # --- Pool Status ---
    ws_dist["A4"] = "Pool Status"
    ws_dist["A4"].font = Font(bold=True, color="e2e8f0")
    pool_status_rows = [
        ("Total GP Pool", f"=SUM(PI_PerfFee)"),
        ("Cash Distributed (this period)", None),  # placeholder, fill later
        ("Remaining in Pool", None),
    ]
    for i, (label, formula) in enumerate(pool_status_rows, start=5):
        ws_dist.cell(row=i, column=1, value=label).font = Font(bold=True)
        if formula is not None:
            c = ws_dist.cell(row=i, column=2, value=formula)
            c.number_format = '"$"#,##0.00'

    # --- Distributions Log ---
    log_header_row = 9
    ws_dist.cell(row=log_header_row - 1, column=1,
                 value="Distributions Log").font = Font(bold=True, color="e2e8f0")
    log_headers = ["#", "Date", "Type", "Party / Payee", "Description", "Amount", "Method", "Status", "Notes"]
    for i, h in enumerate(log_headers, start=1):
        c = ws_dist.cell(row=log_header_row, column=i, value=h)
        _style_header(c)

    log_data_start = log_header_row + 1
    for i, dist in enumerate(DISTRIBUTIONS_DEFAULT, start=log_data_start):
        ws_dist.cell(row=i, column=1, value=i - log_data_start + 1)
        ws_dist.cell(row=i, column=2, value=dist["date"] or "").alignment = Alignment(horizontal="center")
        c = ws_dist.cell(row=i, column=3, value=dist["type"])
        c.fill = PatternFill("solid", fgColor=("fef3c7" if dist["type"] == "Expense" else "dbeafe"))
        c.font = Font(bold=True, color=("92400e" if dist["type"] == "Expense" else "1e3a8a"))
        c.alignment = Alignment(horizontal="center")
        ws_dist.cell(row=i, column=4, value=dist["party"])
        ws_dist.cell(row=i, column=5, value=dist["description"])
        c = ws_dist.cell(row=i, column=6, value=dist["amount"])
        c.number_format = '"$"#,##0.00'
        c.fill = INPUT_FILL
        ws_dist.cell(row=i, column=7, value=dist["method"])
        ws_dist.cell(row=i, column=8, value=dist["status"])
        ws_dist.cell(row=i, column=9, value=dist["notes"])

    # Reserve extra empty rows so the user can add more distributions in-line
    extra_rows = 10
    last_log_row = log_data_start + len(DISTRIBUTIONS_DEFAULT) + extra_rows - 1
    log_total_row = last_log_row + 1
    ws_dist.cell(row=log_total_row, column=4, value="TOTAL").font = Font(bold=True)
    c = ws_dist.cell(row=log_total_row, column=6,
                     value=f"=SUM(F{log_data_start}:F{last_log_row})")
    c.number_format = '"$"#,##0.00'
    c.font = Font(bold=True)
    c.fill = TOTAL_FILL

    # Now wire the Pool Status's "Cash Distributed" + "Remaining"
    ws_dist["B6"] = f"=F{log_total_row}"
    ws_dist["B6"].number_format = '"$"#,##0.00'
    ws_dist["B7"] = "=B5-B6"
    ws_dist["B7"].number_format = '"$"#,##0.00'
    ws_dist["B7"].font = Font(bold=True, color="10b981")

    # Named ranges for Per-Party Settlement to reference
    wb.defined_names["Dist_Type"] = DefinedName(
        name="Dist_Type",
        attr_text=f"Distributions!$C${log_data_start}:$C${last_log_row}",
    )
    wb.defined_names["Dist_Party"] = DefinedName(
        name="Dist_Party",
        attr_text=f"Distributions!$D${log_data_start}:$D${last_log_row}",
    )
    wb.defined_names["Dist_Amount"] = DefinedName(
        name="Dist_Amount",
        attr_text=f"Distributions!$F${log_data_start}:$F${last_log_row}",
    )

    # --- Per-Party Settlement ---
    settlement_label_row = log_total_row + 2
    ws_dist.cell(row=settlement_label_row, column=1,
                 value="Per-Party Settlement").font = Font(bold=True, color="e2e8f0")
    settlement_header_row = settlement_label_row + 1
    settlement_headers = [
        "Party", "Expected Net", "Cash Paid Out", "Outstanding",
        "Suggested Recipient / Wire Info", "Status",
    ]
    for i, h in enumerate(settlement_headers, start=1):
        c = ws_dist.cell(row=settlement_header_row, column=i, value=h)
        _style_header(c)

    # Each Armada party: pull Expected Net from Consultant Summary's column G,
    # sum payouts to that party from the distributions log, outstanding = expected - paid.
    armada_parties = consultant_names + ["Fund Mgmt"] + [s[0] for s in fixed_specs]
    for i, party in enumerate(armada_parties, start=settlement_header_row + 1):
        ws_dist.cell(row=i, column=1, value=party)
        ws_dist.cell(row=i, column=2,
                     value=f"=IFERROR(VLOOKUP(A{i}, 'Consultant Summary'!$A$2:$G${last_summary_row}, 7, FALSE), 0)").number_format = '"$"#,##0.00'
        ws_dist.cell(row=i, column=3,
                     value=f'=SUMIFS(Dist_Amount, Dist_Party, A{i}, Dist_Type, "Payout")').number_format = '"$"#,##0.00'
        ws_dist.cell(row=i, column=4, value=f"=B{i}-C{i}").number_format = '"$"#,##0.00'
        ws_dist.cell(row=i, column=6,
                     value=f'=IF(ROUND(D{i},2)=0,"Settled",IF(ROUND(D{i},2)<0,"Overpaid","Outstanding"))')
        c = ws_dist.cell(row=i, column=5, value="")
        c.fill = INPUT_FILL

    settle_total_row = settlement_header_row + 1 + len(armada_parties)
    ws_dist.cell(row=settle_total_row, column=1, value="TOTAL (Armada GP Pool)").font = Font(bold=True)
    for col_letter in ("B", "C", "D"):
        col_idx = {"B": 2, "C": 3, "D": 4}[col_letter]
        c = ws_dist.cell(row=settle_total_row, column=col_idx,
                         value=f'=SUM({col_letter}{settlement_header_row + 1}:{col_letter}{settle_total_row - 1})')
        c.number_format = '"$"#,##0.00'
        c.font = Font(bold=True)
        c.fill = TOTAL_FILL

    # TruQuant — shown below the total since it's external (paid upstream)
    tq_settle_row = settle_total_row + 2
    ws_dist.cell(row=tq_settle_row - 1, column=1,
                 value="External (does not draw from this pool)").font = Font(italic=True, color="94a3b8")
    ws_dist.cell(row=tq_settle_row, column=1, value="TruQuant").font = Font(italic=True, color="fbbf24")
    ws_dist.cell(row=tq_settle_row, column=2,
                 value=f"=TQ_Income-IFERROR(INDEX(Costs_PartyTotals, MATCH(A{tq_settle_row}, Costs_PartyHeaders, 0)),0)").number_format = '"$"#,##0.00'
    ws_dist.cell(row=tq_settle_row, column=3, value="—").alignment = Alignment(horizontal="right")
    ws_dist.cell(row=tq_settle_row, column=4, value="—").alignment = Alignment(horizontal="right")
    ws_dist.cell(row=tq_settle_row, column=5,
                 value="External — TruQuant's 18% comes upstream of this pool. Their share of TQ/Armada costs is also paid externally.")
    ws_dist.cell(row=tq_settle_row, column=6, value="External")

    # Note row
    note_row = tq_settle_row + 2
    ws_dist.cell(row=note_row, column=1,
                 value=("Note: Expenses (vendor cash) reduce the pool but don't count toward any party's "
                        "Cash Paid Out — only Payout-type distributions do. Cash Paid Out aggregates from "
                        "the Distributions Log via SUMIFS on Type='Payout'. To record a new transfer, add "
                        "a row in the Distributions Log above and the Settlement table updates automatically.")
                 ).font = Font(italic=True, color="94a3b8")
    ws_dist.merge_cells(f"A{note_row}:I{note_row}")
    ws_dist.cell(row=note_row, column=1).alignment = Alignment(wrap_text=True)

    # Column widths
    ws_dist.column_dimensions["A"].width = 32
    ws_dist.column_dimensions["B"].width = 18
    ws_dist.column_dimensions["C"].width = 14
    ws_dist.column_dimensions["D"].width = 18
    ws_dist.column_dimensions["E"].width = 50
    ws_dist.column_dimensions["F"].width = 18
    ws_dist.column_dimensions["G"].width = 12
    ws_dist.column_dimensions["H"].width = 12
    ws_dist.column_dimensions["I"].width = 38

    # --------- 7. Reconciliation ---------
    ws_r = wb.create_sheet("Reconciliation")
    ws_r["A1"] = "Reconciliation Checks"
    ws_r["A1"].font = TITLE_FONT
    ws_r.merge_cells("A1:D1")

    checks = [
        ("Sum of per-investor Perf Fee (GP pool from TPA)",
         f"=SUM(PI_PerfFee)",
         "TPA crystallized total — should match the package's reported number"),
        ("Sum of all per-investor split columns (H–L)",
         f"='Per-Investor Allocation'!H{total_row}+'Per-Investor Allocation'!I{total_row}+'Per-Investor Allocation'!J{total_row}+'Per-Investor Allocation'!K{total_row}+'Per-Investor Allocation'!L{total_row}",
         "Should equal sum of perf fees (no leakage)"),
        ("Difference (H+I+J+K+L − Perf Fee total)",
         f"=ROUND(B4-B3,2)",
         "Should be 0.00"),
        ("# investors with consultant = Unmapped",
         f'=COUNTIF(PI_Consultant, "Unmapped")',
         "Should be 0; if not, fix the IDS Mapping sheet"),
        ("Split % sanity (must equal 100%)",
         f"=Inputs!$B$11",
         "Sum of all five split percentages"),
    ]
    ws_r.cell(row=2, column=1, value="Check").font = Font(bold=True)
    ws_r.cell(row=2, column=2, value="Result").font = Font(bold=True)
    ws_r.cell(row=2, column=3, value="Notes").font = Font(bold=True)
    for i, (label, formula, note) in enumerate(checks, start=3):
        ws_r.cell(row=i, column=1, value=label)
        c = ws_r.cell(row=i, column=2, value=formula)
        if "$" in note or "TPA" in label or "Sum" in label:
            c.number_format = '"$"#,##0.00'
        ws_r.cell(row=i, column=3, value=note)

    ws_r.column_dimensions["A"].width = 50
    ws_r.column_dimensions["B"].width = 18
    ws_r.column_dimensions["C"].width = 60

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    return {
        "investor_count": len(records),
        "unmapped": [r for r in records if r["consultant"] == "Unmapped"],
    }


# ---------------------------------------------------------------------------
# JSON builder (snapshot for the dashboard)
# ---------------------------------------------------------------------------

def build_json_snapshot(
    records: list[dict],
    tpa_record: dict,
    period_label: str,
    costs: list[dict],
) -> dict:
    """Compute static snapshot used by the dashboard."""
    total_perf = sum(r["perf_fee"] for r in records)
    total_costs = sum(c["amount"] for c in costs)
    income_map = build_income_map(records, tpa_record)
    cost_alloc = [allocate_cost(c, income_map) for c in costs]
    tq_income = compute_tq_income(tpa_record)

    # Per-consultant aggregation (computed values, not formulas)
    by_cons: dict[str, dict] = {}
    for r in records:
        bucket = by_cons.setdefault(r["consultant"], {
            "consultant": r["consultant"],
            "investor_count": 0,
            "capital_raised": 0.0,
            "gp_earned_consultant_cut": 0.0,
            "investors": [],
        })
        bucket["investor_count"] += 1
        bucket["capital_raised"] += r["ending_balance"]
        bucket["gp_earned_consultant_cut"] += r["perf_fee"] * SPLIT_PCTS["consultant"]
        bucket["investors"].append({
            "tpa_id": r["tpa_id"],
            "name": r["name"],
            "ending_balance": round(r["ending_balance"], 2),
            "gross_profit": round(r["gross_profit"], 2),
            "perf_fee": round(r["perf_fee"], 2),
            "consultant_cut": round(r["perf_fee"] * SPLIT_PCTS["consultant"], 2),
        })

    consultants = []
    for k, v in sorted(by_cons.items(), key=lambda kv: -kv[1]["gp_earned_consultant_cut"]):
        consultants.append({
            **v,
            "capital_raised": round(v["capital_raised"], 2),
            "gp_earned_consultant_cut": round(v["gp_earned_consultant_cut"], 2),
            "pct_of_gp_pool": round(v["gp_earned_consultant_cut"] / total_perf, 6) if total_perf else 0,
        })

    fund_mgmt_total = total_perf * SPLIT_PCTS["fund_mgmt"]
    fixed = {
        "raj": total_perf * SPLIT_PCTS["raj"],
        "nairne": total_perf * SPLIT_PCTS["nairne"],
        "alec_gp": total_perf * SPLIT_PCTS["alec"],
    }

    # Per-party total expense burden (sum of their share across every cost line item)
    party_burden: dict[str, float] = {}
    for alloc in cost_alloc:
        for party, amt in alloc.items():
            party_burden[party] = party_burden.get(party, 0.0) + amt

    # Cost allocation matrix for the dashboard (rows = costs, cols = parties)
    cost_matrix_parties = sorted(income_map.keys(), key=lambda p: -income_map[p])
    cost_matrix = []
    for cost, alloc in zip(costs, cost_alloc):
        cost_matrix.append({
            "name": cost["name"],
            "amount": round(cost["amount"], 2),
            "payor_group": cost["payor_group"],
            "by_party": {p: round(alloc.get(p, 0.0), 2) for p in cost_matrix_parties},
        })

    income_breakdown = [
        {
            "party": p,
            "income": round(i, 2),
            "pct_tq_armada": round(i / sum(income_map.values()), 6) if sum(income_map.values()) else 0,
            "pct_armada": round(i / sum(v for k, v in income_map.items() if k != "TruQuant"), 6)
                          if p != "TruQuant" else None,
        }
        for p, i in sorted(income_map.items(), key=lambda kv: -kv[1])
    ]

    bs = tpa_record.get("balance_sheet", {})
    inc = tpa_record.get("income_statement", {})
    fl = tpa_record.get("fund_level", {})

    return {
        "period": tpa_record["period"],
        "period_label": period_label,
        "as_of": tpa_record["as_of"],
        "split_pcts": SPLIT_PCTS,
        "fund_totals": {
            "tpa_gross_income": round(inc.get("total_income", 0), 2),
            "tpa_net_income": round(inc.get("net_income", 0), 2),
            "tpa_perf_fees_crystallized": round(total_perf, 2),
            "fund_mgmt_pool": round(fund_mgmt_total, 2),
            "consultants_pool": round(total_perf * SPLIT_PCTS["consultant"], 2),
            "raj_pool": round(fixed["raj"], 2),
            "nairne_pool": round(fixed["nairne"], 2),
            "alec_gp_pool": round(fixed["alec_gp"], 2),
            "investor_count": len(records),
            "unmapped_count": sum(1 for r in records if r["consultant"] == "Unmapped"),
            "ending_aum": round(bs.get("total_capital", 0), 2),
            "gross_mtd_ror": fl.get("gross_mtd_ror", 0),
            "net_mtd_ror": fl.get("net_mtd_ror", 0),
            "total_costs": round(total_costs, 2),
            "tq_income": round(tq_income, 2),
        },
        "consultants": [
            {
                **c,
                "weighted_expense": round(party_burden.get(c["consultant"], 0), 2),
                "net_profit": round(c["gp_earned_consultant_cut"] - party_burden.get(c["consultant"], 0), 2),
            }
            for c in consultants
        ],
        "investors": [
            {
                "tpa_id": r["tpa_id"],
                "name": r["name"],
                "consultant": r["consultant"],
                "ending_balance": round(r["ending_balance"], 2),
                "gross_profit": round(r["gross_profit"], 2),
                "perf_fee": round(r["perf_fee"], 2),
                "fund_mgmt_cut": round(r["perf_fee"] * SPLIT_PCTS["fund_mgmt"], 2),
                "consultant_cut": round(r["perf_fee"] * SPLIT_PCTS["consultant"], 2),
                "raj_cut": round(r["perf_fee"] * SPLIT_PCTS["raj"], 2),
                "nairne_cut": round(r["perf_fee"] * SPLIT_PCTS["nairne"], 2),
                "alec_gp_cut": round(r["perf_fee"] * SPLIT_PCTS["alec"], 2),
            }
            for r in sorted(records, key=lambda x: -x["perf_fee"])
        ],
        "costs": {
            "total": round(total_costs, 2),
            "matrix_parties": cost_matrix_parties,
            "line_items": cost_matrix,
            "party_burden": [
                {"party": p, "amount": round(party_burden.get(p, 0), 2)}
                for p in sorted(party_burden, key=lambda x: -party_burden[x])
            ],
            "income_breakdown": income_breakdown,
        },
    }


def upsert_json(snapshot: dict, json_path: Path = JSON_PATH) -> None:
    if json_path.exists():
        history = json.loads(json_path.read_text())
    else:
        history = {"fund": "Armada Prime LLP", "months": []}
    history["fund"] = "Armada Prime LLP"
    months = [m for m in history.get("months", []) if m.get("period") != snapshot["period"]]
    months.append(snapshot)
    months.sort(key=lambda m: m["period"])
    history["months"] = months
    json_path.parent.mkdir(parents=True, exist_ok=True)
    json_path.write_text(json.dumps(history, indent=2))


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("tpa_xlsx", type=Path, nargs="?", default=DEFAULT_TPA,
                    help="Path to TPA Reporting Package xlsx")
    ap.add_argument("--internal", type=Path, default=DEFAULT_INTERNAL,
                    help="Path to internal Monthly Return xlsx (for IDS sheet)")
    ap.add_argument("--output-xlsx", type=Path, default=DEFAULT_OUTPUT_XLSX,
                    help="Where to write the dynamic Excel")
    ap.add_argument("--label", type=str, default=None, help="Override period_label")
    args = ap.parse_args()

    if not args.tpa_xlsx.exists():
        print(f"error: TPA file not found: {args.tpa_xlsx}", file=sys.stderr)
        return 1
    if not args.internal.exists():
        print(f"error: internal file not found: {args.internal}", file=sys.stderr)
        return 1

    print(f"Loading TPA package:  {args.tpa_xlsx.name}")
    tpa = parse_workbook(args.tpa_xlsx)
    print(f"  period={tpa['period']}  investors={len(tpa['investors'])}  perf_fees=${sum(i.get('perf_fee',0) for i in tpa['investors']):,.2f}")

    print(f"Loading internal IDS: {args.internal.name}")
    ids_map = load_ids(args.internal)
    print(f"  IDS rows: {len(ids_map)}  (incl. {len(CONSULTANT_OVERRIDES)} confirmed overrides)")

    records = reconcile(tpa["investors"], ids_map)
    unmapped = [r for r in records if r["consultant"] == "Unmapped"]
    print(f"Reconciled: {len(records)} investors, {len(unmapped)} unmapped")
    for u in unmapped:
        print(f"  UNMAPPED: {u['tpa_id']:22}  {u['name']}")

    costs = load_costs(args.internal)
    print(f"Costs: {len(costs)} line items, total ${sum(c['amount'] for c in costs):,.2f}")

    label = args.label or tpa["period_label"]
    summary = build_workbook(records, label, args.output_xlsx, costs, tpa)
    print(f"Wrote Excel: {args.output_xlsx}")

    snapshot = build_json_snapshot(records, tpa, label, costs)
    upsert_json(snapshot)
    print(f"Wrote JSON: {JSON_PATH}")

    print(f"\nGP pool ${snapshot['fund_totals']['tpa_perf_fees_crystallized']:,.2f}  →  "
          f"Fund Mgmt ${snapshot['fund_totals']['fund_mgmt_pool']:,.2f}  +  "
          f"Consultants ${snapshot['fund_totals']['consultants_pool']:,.2f}  +  "
          f"Raj/Nairne/Alec(GP) ${snapshot['fund_totals']['raj_pool']*3:,.2f}")
    print(f"Total costs: ${snapshot['fund_totals']['total_costs']:,.2f} (allocated pro-rata by GP %)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
