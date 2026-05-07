#!/usr/bin/env python3
"""Build a Form 1065 (Accrual Basis) workbook for Armada Prime Tech LLC, tax year 2025.

Tax-optimized strategy:
  - Accounting method: ACCRUAL (recognize Dec 2025 deductions in 2025, not 2026)
  - Insurance $18K kept fully deductible in 2025 (12-month prepaid rule)
  - 506c SPV Loans $29,275 reclassified to Balance Sheet (Asset, not P&L)
  - Phil treated as K-1 partner (not 1099)

Output:
  /Users/nairne/claude-central-hub/monily-package/Form_1065_Accrual_2025.xlsx

Tabs (mirror Form 1065 structure):
  1. Cover & Filing Info
  2. Form 1065 Page 1 — Income & Deductions
  3. Schedule K — Partners' Distributive Share
  4. Schedule K-1 — Nairne
  5. Schedule K-1 — Raj Duggal
  6. Schedule K-1 — Phil
  7. Schedule L — Balance Sheet (per books)
  8. Schedule M-1 — Income Reconciliation
  9. Schedule M-2 — Partners' Capital Analysis
  10. Supporting — 1099 Contractor Detail
  11. Supporting — Op Expenses Detail

Usage:
    python tools/build_form_1065_accrual.py
"""
from __future__ import annotations

import sys
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
from build_2025_year_end import (
    ACTUAL_PAID,
    GP_OP_EXPENSES,
    K1_RECIPIENTS,
    NAIRNE_ALIASES,
    PERIOD_LABELS,
)

REPO_ROOT = Path(__file__).resolve().parent.parent
OUT_DIR = REPO_ROOT / "monily-package"
OUT_XLSX = OUT_DIR / "Form_1065_Accrual_2025.xlsx"

# Styling
HDR_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
SECTION_FONT = Font(bold=True, color="FFFFFF", size=12)
SUBTOTAL_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
WARN_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
K1_FILL = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
THIN = Side(border_style="thin", color="999999")
MEDIUM = Side(border_style="medium", color="333333")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MONEY = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
INT_MONEY = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'

ENTITY_NAME = "Armada Prime Tech LLC"
TAX_YEAR = "2025"
PERIOD_DESC = "Aug 1, 2025 – Dec 31, 2025"


# ---------------------------------------------------------------------------
# Year-totals (accrual basis, tax-optimized)
# ---------------------------------------------------------------------------

def compute_totals():
    yt = {}
    for period, recipients in ACTUAL_PAID.items():
        for k, v in recipients.items():
            yt[k] = yt.get(k, 0) + v

    # Revenue: TPA Performance Fees Crystallized (accrual = recognized when earned)
    revenue = sum([15355.51, 12474.75, 51906.99, 57074.09, 16211.69])

    # K-1 partners
    nairne_cash = sum(yt.get(a, 0) for a in NAIRNE_ALIASES)
    raj_cash = yt.get("Raj", 0)
    phil_cash = yt.get("Phil", 0)
    contractor_total = sum(v for k, v in yt.items() if k not in K1_RECIPIENTS)

    # Operating expenses (accrual + tax-optimized reclass)
    # Keep full $18K insurance (12-month prepaid rule)
    # Move $29,275 SPV Loans to Balance Sheet (NOT a P&L expense)
    op_raw = {}
    mapping = {
        "506c SPV Loan": "506c SPV Loan",
        "PVD": "PVD",
        "Website": "Website",
        "Ad Spend": "Ad Spend",
        "Chris": "Chris",
        "Alpha Verification": "Alpha Verification",
        "Formidium (TPA)": "TPA",
        "TPA (Formidium)": "TPA",
        "TPA (second line)": "TPA",
        "Insurance": "Insurance",
    }
    for period, items in GP_OP_EXPENSES.items():
        for vendor, amount in items:
            key = mapping.get(vendor, vendor)
            op_raw[key] = op_raw.get(key, 0) + amount

    spv_amount = op_raw.pop("506c SPV Loan", 0)  # remove from P&L
    op_total = sum(op_raw.values())

    # Form 1065 line classifications:
    # Line 9: Salaries & wages = $0 (no W-2 employees)
    # Line 10: Guaranteed payments to partners = $0 (no GP structure for now)
    # Line 13: Rent = $0
    # Line 14: Taxes & licenses = $0 (state LLC fee tracked separately if known)
    # Line 19: Employee benefit programs = $0
    # Line 20: Other deductions (statement attached) — most ops fall here
    # Form 1065 also has: cost of goods sold (Line 2) for product businesses — N/A here

    # Map our categories to Form 1065 lines
    line_breakdown = {
        "L9_salaries": 0,  # No W-2
        "L10_guaranteed_payments": 0,  # None elected
        "L13_rent": 0,
        "L14_taxes_licenses": 0,
        "L15_interest": 0,
        "L16_depreciation": 0,
        "L17_depletion": 0,
        "L18_retirement": 0,
        "L19_benefits": 0,
        "L20_other": op_total + contractor_total,  # Bulk of deductions go here
        # "Other deductions" sub-categories (from statement):
        "subL20_capital_raiser_commissions": contractor_total,
        "subL20_contractor_labor_chris": op_raw.get("Chris", 0),
        "subL20_insurance_DnO": op_raw.get("Insurance", 0),
        "subL20_marketing_website": op_raw.get("Website", 0),
        "subL20_marketing_advertising": op_raw.get("Ad Spend", 0),
        "subL20_compliance_verification": op_raw.get("Alpha Verification", 0),
        "subL20_admin_TPA_fees": op_raw.get("TPA", 0),
        "subL20_professional_fees_PVD": op_raw.get("PVD", 0),
    }

    total_deductions = (
        line_breakdown["L9_salaries"]
        + line_breakdown["L10_guaranteed_payments"]
        + line_breakdown["L13_rent"]
        + line_breakdown["L14_taxes_licenses"]
        + line_breakdown["L15_interest"]
        + line_breakdown["L16_depreciation"]
        + line_breakdown["L17_depletion"]
        + line_breakdown["L18_retirement"]
        + line_breakdown["L19_benefits"]
        + line_breakdown["L20_other"]
    )
    ordinary_business_income = revenue - total_deductions

    # K-1 ownership %
    nairne_pct = 60.0 / 61.0
    raj_pct = 0.5 / 61.0
    phil_pct = 0.5 / 61.0

    return {
        "yt": yt,
        "revenue": revenue,
        "contractor_total": contractor_total,
        "op_raw": op_raw,
        "op_total": op_total,
        "spv_amount": spv_amount,
        "line_breakdown": line_breakdown,
        "total_deductions": total_deductions,
        "ordinary_business_income": ordinary_business_income,
        "nairne_cash": nairne_cash,
        "raj_cash": raj_cash,
        "phil_cash": phil_cash,
        "nairne_pct": nairne_pct,
        "raj_pct": raj_pct,
        "phil_pct": phil_pct,
        "nairne_k1_income": ordinary_business_income * nairne_pct,
        "raj_k1_income": ordinary_business_income * raj_pct,
        "phil_k1_income": ordinary_business_income * phil_pct,
    }


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def title(ws, doc_title: str, ncols: int = 4) -> int:
    ws.cell(row=1, column=1, value=ENTITY_NAME).font = Font(bold=True, size=16)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.cell(row=2, column=1, value=doc_title).font = Font(bold=True, size=13, color="1F4E79")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    ws.cell(row=3, column=1, value=f"Form 1065 — Tax Year {TAX_YEAR} — ACCRUAL BASIS  |  Period: {PERIOD_DESC}  |  Generated {date.today().isoformat()}").font = Font(italic=True, color="666666")
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=ncols)
    return 5


def hdr_row(ws, row: int, ncols: int):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BOX


def section(ws, row: int, title_text: str, ncols: int) -> int:
    ws.cell(row=row, column=1, value=title_text)
    for c in range(1, ncols + 1):
        ws.cell(row=row, column=c).fill = SECTION_FILL
        ws.cell(row=row, column=c).font = SECTION_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    return row + 1


# ---------------------------------------------------------------------------
# Tab 1: Cover & Filing Info
# ---------------------------------------------------------------------------

def build_cover(wb, T):
    ws = wb.create_sheet("1. Cover & Filing Info")
    r = title(ws, "Cover & Filing Info")

    info = [
        ("Entity Name", "Armada Prime Tech LLC"),
        ("Tax Year", "2025"),
        ("Tax Form", "Form 1065 (U.S. Return of Partnership Income)"),
        ("Accounting Method", "★ ACCRUAL (recommended for tax minimization)"),
        ("Period of Operations", "Aug 1, 2025 – Dec 31, 2025"),
        ("Filing Status", "First year of filing"),
        ("Tax Classification", "Multi-member LLC, taxed as partnership"),
        ("Number of Partners (K-1 recipients)", "3"),
        ("Number of 1099-NEC Contractors", "6 (Alec, Jake, AJ, Issac, Luke, Nikki) + Chris"),
        ("EIN", "⚠️ TO PROVIDE"),
        ("State of Formation", "⚠️ TO PROVIDE"),
        ("Business Address", "⚠️ TO PROVIDE"),
        ("Phone / Email", "⚠️ TO PROVIDE"),
    ]
    for label, value in info:
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws.cell(row=r, column=2, value=value)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        if "⚠️" in str(value):
            for c in range(1, 5):
                ws.cell(row=r, column=c).fill = WARN_FILL
        if "★" in str(value):
            for c in range(1, 5):
                ws.cell(row=r, column=c).fill = TOTAL_FILL
                ws.cell(row=r, column=c).font = Font(bold=True)
        r += 1
    r += 1

    r = section(ws, r, "TAX-OPTIMIZED METHOD CHOICES", 4)
    choices = [
        ("Accounting Method", "ACCRUAL", "Recognize Dec 2025 expenses in 2025 (when only 5 months of revenue). Defer income recognition compared to ad-hoc cash treatment of late-arriving wires."),
        ("Insurance Treatment", "Full $18K deducted in 2025", "12-month prepaid rule (IRS Pub 535) allows full deduction of prepaid expenses with benefit period ≤12 months from payment."),
        ("506c SPV Loans Treatment", "Reclassified to Balance Sheet (Asset)", "$29,275 in SPV Loan disbursements are loan receivables / equity investments — NOT P&L expenses."),
        ("Phil's 0.5% Slice", "K-1 Partner (NOT 1099 contractor)", "Per Nairne 2026-05-05 — Phil is a partner; his slice is a partner allocation, not a contractor expense."),
    ]
    for label, choice, rationale in choices:
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws.cell(row=r, column=2, value=choice).font = Font(bold=True, color="00703C")
        ws.cell(row=r, column=3, value=rationale).alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        ws.row_dimensions[r].height = 35
        r += 1
    r += 1

    r = section(ws, r, "HEADLINE — FORM 1065 PAGE 1 RESULT", 4)
    headline = [
        ("Total Revenue (Line 8)", T["revenue"]),
        ("Total Deductions (Line 21)", T["total_deductions"]),
        ("Ordinary Business Income (Line 22)", T["ordinary_business_income"]),
    ]
    for label, val in headline:
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws.cell(row=r, column=2, value=val).number_format = MONEY
        for c in range(1, 5):
            ws.cell(row=r, column=c).fill = TOTAL_FILL
            ws.cell(row=r, column=c).font = Font(bold=True)
        r += 1
    r += 1

    r = section(ws, r, "K-1 PARTNERS (3 partners)", 4)
    ws.cell(row=r, column=1, value="Partner").font = Font(bold=True)
    ws.cell(row=r, column=2, value="Ownership %").font = Font(bold=True)
    ws.cell(row=r, column=3, value="K-1 Box 1 Ordinary Income").font = Font(bold=True)
    ws.cell(row=r, column=4, value="Cash Distributions").font = Font(bold=True)
    hdr_row(ws, r, 4)
    r += 1
    partners = [
        ("Nairne", T["nairne_pct"], T["nairne_k1_income"], T["nairne_cash"]),
        ("Raj Duggal", T["raj_pct"], T["raj_k1_income"], T["raj_cash"]),
        ("Phil", T["phil_pct"], T["phil_k1_income"], T["phil_cash"]),
    ]
    for name, pct, k1, cash in partners:
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=f"{pct*100:.4f}%")
        ws.cell(row=r, column=3, value=k1).number_format = MONEY
        ws.cell(row=r, column=4, value=cash).number_format = MONEY
        for c in range(1, 5):
            ws.cell(row=r, column=c).fill = K1_FILL
        r += 1
    r += 1

    r = section(ws, r, "TAB GUIDE", 4)
    tabs = [
        ("1. Cover & Filing Info", "This tab — entity, method elections, headline numbers"),
        ("2. Form 1065 Page 1", "Income & Deductions (Lines 1–22)"),
        ("3. Schedule K", "Partners' Distributive Share (aggregate)"),
        ("4. K-1 Nairne", "Schedule K-1 for Nairne (60%)"),
        ("5. K-1 Raj Duggal", "Schedule K-1 for Raj Duggal (0.5%)"),
        ("6. K-1 Phil", "Schedule K-1 for Phil (0.5%)"),
        ("7. Schedule L", "Balance Sheet per Books (Beginning + End)"),
        ("8. Schedule M-1", "Reconciliation of Income"),
        ("9. Schedule M-2", "Analysis of Partners' Capital Accounts"),
        ("10. 1099 Detail", "Per-contractor breakdown for 1099-NEC issuance"),
        ("11. Op Expenses Detail", "Vendor-level expense detail"),
    ]
    for tab, desc in tabs:
        ws.cell(row=r, column=1, value=tab).font = Font(bold=True)
        ws.cell(row=r, column=2, value=desc)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        r += 1

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 30


# ---------------------------------------------------------------------------
# Tab 2: Form 1065 Page 1 — Income & Deductions
# ---------------------------------------------------------------------------

def build_page1(wb, T):
    ws = wb.create_sheet("2. Form 1065 Page 1")
    r = title(ws, "Form 1065 Page 1 — Income & Deductions (Accrual Basis)")
    headers = ["Form 1065 Line", "Description", "Amount ($)", "Notes / Source"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=r, column=i, value=h)
    hdr_row(ws, r, 4)
    r += 1

    def line(line_no, desc, amount=None, notes="", bold=False, fill=None):
        nonlocal r
        ws.cell(row=r, column=1, value=line_no)
        ws.cell(row=r, column=2, value=desc)
        if amount is not None:
            ws.cell(row=r, column=3, value=amount).number_format = MONEY
        ws.cell(row=r, column=4, value=notes).alignment = Alignment(wrap_text=True)
        if bold:
            for c in range(1, 5):
                ws.cell(row=r, column=c).font = Font(bold=True)
        if fill:
            for c in range(1, 5):
                ws.cell(row=r, column=c).fill = fill
        r += 1

    # INCOME SECTION
    r = section(ws, r, "INCOME (Lines 1–8)", 4)
    line("1a", "Gross receipts or sales", T["revenue"], "Performance Fees Crystallized from Armada Prime LLP, accrued Aug-Dec 2025 per TPA Reporting Packages")
    line("1b", "Returns and allowances", 0, "")
    line("1c", "Balance (1a − 1b)", T["revenue"], "", bold=True)
    line("2", "Cost of goods sold", 0, "N/A — services-only LLC")
    line("3", "Gross profit (1c − 2)", T["revenue"], "", bold=True)
    line("4", "Ordinary income/(loss) from other partnerships, estates, trusts", 0, "")
    line("5", "Net farm profit/(loss)", 0, "")
    line("6", "Net gain/(loss) from Form 4797 (Disposition of business property)", 0, "")
    line("7", "Other income/(loss) (statement)", 0, "")
    line("8", "TOTAL INCOME (lines 3-7)", T["revenue"], "", bold=True, fill=SUBTOTAL_FILL)
    r += 1

    # DEDUCTIONS SECTION
    r = section(ws, r, "DEDUCTIONS (Lines 9–21)", 4)
    L = T["line_breakdown"]
    line("9", "Salaries and wages (less employment credits)", L["L9_salaries"], "$0 — no W-2 employees")
    line("10", "Guaranteed payments to partners", L["L10_guaranteed_payments"], "$0 — no guaranteed payment structure (could be elected for Nairne in future)")
    line("11", "Repairs and maintenance", 0, "")
    line("12", "Bad debts", 0, "")
    line("13", "Rent", L["L13_rent"], "$0 — no leased office space")
    line("14", "Taxes and licenses", L["L14_taxes_licenses"], "$0 — state LLC fee tracked separately if applicable; ⚠️ verify")
    line("15", "Interest", L["L15_interest"], "")
    line("16a", "Depreciation (from Form 4562)", L["L16_depreciation"], "$0 — no fixed assets to depreciate")
    line("16b", "Less depreciation reported elsewhere", 0, "")
    line("16c", "Subtract line 16b from 16a", L["L16_depreciation"], "")
    line("17", "Depletion (Do not deduct oil/gas)", L["L17_depletion"], "")
    line("18", "Retirement plans, etc.", L["L18_retirement"], "$0 — Solo 401(k) recommended for 2026; nothing for 2025")
    line("19", "Employee benefit programs", L["L19_benefits"], "")
    line("20", "Other deductions (statement attached — see below)", L["L20_other"], "See Other Deductions Statement")
    line("21", "TOTAL DEDUCTIONS (lines 9-20)", T["total_deductions"], "", bold=True, fill=SUBTOTAL_FILL)
    r += 1

    # ORDINARY BUSINESS INCOME
    r = section(ws, r, "ORDINARY BUSINESS INCOME (Line 22)", 4)
    line("22", "Ordinary business income/(loss) (Line 8 − Line 21)", T["ordinary_business_income"],
         "Flows to Schedule K Line 1 → allocated to partners on K-1 Box 1", bold=True, fill=TOTAL_FILL)
    r += 1

    # OTHER DEDUCTIONS STATEMENT (Line 20 detail)
    r = section(ws, r, "OTHER DEDUCTIONS STATEMENT (supporting Line 20)", 4)
    other = [
        ("Capital raiser commissions (1099-NEC contractors)", L["subL20_capital_raiser_commissions"], "Alec, Jake, AJ, Issac, Luke, Nikki — see Tab 10"),
        ("Contractor labor — Chris (1099-NEC)", L["subL20_contractor_labor_chris"], "Operating contractor"),
        ("Insurance — D&O", L["subL20_insurance_DnO"], "Full annual policy premium; deductible under 12-month prepaid rule (Reg §1.263(a)-4(f))"),
        ("Marketing — Website", L["subL20_marketing_website"], ""),
        ("Marketing — Advertising / Ad Spend", L["subL20_marketing_advertising"], ""),
        ("Compliance — Verification (Alpha)", L["subL20_compliance_verification"], ""),
        ("Administrative — TPA fees (Formidium)", L["subL20_admin_TPA_fees"], ""),
        ("Professional fees — PVD", L["subL20_professional_fees_PVD"], "Vendor type to verify for 1099 obligation"),
    ]
    sub_total = 0
    for label, amt, notes in other:
        line("20", "  " + label, amt, notes)
        sub_total += amt
    line("20", "TOTAL OTHER DEDUCTIONS (matches Line 20 above)", sub_total, "", bold=True, fill=SUBTOTAL_FILL)

    r += 1
    r = section(ws, r, "ITEMS NOT ON P&L (TRACK SEPARATELY)", 4)
    line("—", "506c SPV Loan disbursements (RECLASSIFIED to Balance Sheet)", T["spv_amount"],
         "Aug $4,275 + Oct $25,000 = $29,275. Loan receivables / SPV equity investment; appears on Schedule L (Tab 7), not Form 1065 Page 1.", fill=WARN_FILL)

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 70


# ---------------------------------------------------------------------------
# Tab 3: Schedule K — Partners' Distributive Share
# ---------------------------------------------------------------------------

def build_schedule_k(wb, T):
    ws = wb.create_sheet("3. Schedule K")
    r = title(ws, "Schedule K — Partners' Distributive Share Items (AGGREGATE)")
    headers = ["Line", "Description", "Amount ($)", "Notes"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=r, column=i, value=h)
    hdr_row(ws, r, 4)
    r += 1

    def line(line_no, desc, amount=None, notes="", bold=False, fill=None):
        nonlocal r
        ws.cell(row=r, column=1, value=line_no)
        ws.cell(row=r, column=2, value=desc)
        if amount is not None:
            ws.cell(row=r, column=3, value=amount).number_format = MONEY
        ws.cell(row=r, column=4, value=notes).alignment = Alignment(wrap_text=True)
        if bold:
            for c in range(1, 5):
                ws.cell(row=r, column=c).font = Font(bold=True)
        if fill:
            for c in range(1, 5):
                ws.cell(row=r, column=c).fill = fill
        r += 1

    r = section(ws, r, "INCOME (LOSS)", 4)
    line("1", "Ordinary business income (loss)", T["ordinary_business_income"], "From Form 1065 Page 1, Line 22", bold=True)
    line("2", "Net rental real estate income (loss)", 0, "")
    line("3a", "Other gross rental income (loss)", 0, "")
    line("4a", "Guaranteed payments — services", 0, "$0")
    line("4b", "Guaranteed payments — capital", 0, "")
    line("5", "Interest income", 0, "")
    line("6a", "Ordinary dividends", 0, "")
    line("6b", "Qualified dividends", 0, "")
    line("7", "Royalties", 0, "")
    line("8", "Net short-term capital gain (loss)", 0, "")
    line("9a", "Net long-term capital gain (loss)", 0, "")
    line("10", "Net section 1231 gain (loss)", 0, "")
    line("11", "Other income (loss)", 0, "")
    r += 1

    r = section(ws, r, "DEDUCTIONS", 4)
    line("12", "Section 179 deduction", 0, "")
    line("13a", "Contributions", 0, "")
    line("13c(1)", "Investment interest expense", 0, "")
    line("13d", "Other deductions", 0, "")
    r += 1

    r = section(ws, r, "SELF-EMPLOYMENT", 4)
    line("14a", "Net earnings (loss) from self-employment", T["ordinary_business_income"],
         "Active managing members' shares are SE-taxable. Nairne is active; Raj/Phil status to confirm.", bold=True)
    line("14b", "Gross farming or fishing income", 0, "")
    line("14c", "Gross nonfarm income", T["revenue"], "")
    r += 1

    r = section(ws, r, "ALTERNATIVE MINIMUM TAX (AMT) ITEMS", 4)
    line("17a-17f", "AMT items", 0, "")
    r += 1

    r = section(ws, r, "OTHER INFORMATION", 4)
    line("19a", "Distributions of cash and marketable securities",
         T["nairne_cash"] + T["raj_cash"] + T["phil_cash"],
         "Total cash distributions to all partners in 2025")
    line("19b", "Distributions of other property", 0, "")
    line("20a", "Investment income", 0, "")
    line("20b", "Investment expenses", 0, "")
    r += 1

    r = section(ws, r, "ANALYSIS OF NET INCOME", 4)
    line("—", "Net income per Schedule K (Line 1)", T["ordinary_business_income"],
         "Allocated to partners per ownership: Nairne 98.36%, Raj 0.82%, Phil 0.82%",
         bold=True, fill=TOTAL_FILL)

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 60


# ---------------------------------------------------------------------------
# Tabs 4-6: Schedule K-1 per partner
# ---------------------------------------------------------------------------

def build_k1(wb, partner_name: str, ownership_pct: float, k1_income: float,
             cash_dist: float, tab_num: int, ssn_label: str = "⚠️ TO PROVIDE"):
    ws = wb.create_sheet(f"{tab_num}. K-1 {partner_name}")
    r = title(ws, f"Schedule K-1 (Form 1065) — {partner_name}")
    ws.cell(row=r-1, column=1, value=f"Partner: {partner_name}  |  Ownership: {ownership_pct*100:.4f}%").font = Font(italic=True, color="666666")
    r += 1

    # Part I — Information About the Partnership
    r = section(ws, r, "PART I — INFORMATION ABOUT THE PARTNERSHIP", 4)
    info_p1 = [
        ("A. Partnership EIN", "⚠️ TO PROVIDE"),
        ("B. Partnership name + address", "Armada Prime Tech LLC  |  ⚠️ ADDRESS TO PROVIDE"),
        ("C. IRS Center where partnership filed", "Form 1065 e-filed"),
        ("D. Check if publicly traded partnership", "No"),
    ]
    for label, value in info_p1:
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws.cell(row=r, column=2, value=value)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        if "⚠️" in str(value):
            for c in range(1, 5):
                ws.cell(row=r, column=c).fill = WARN_FILL
        r += 1
    r += 1

    # Part II — Information About the Partner
    r = section(ws, r, "PART II — INFORMATION ABOUT THE PARTNER", 4)
    info_p2 = [
        ("E. Partner SSN/EIN", ssn_label),
        ("F. Partner name + address", f"{partner_name}  |  ⚠️ ADDRESS TO PROVIDE"),
        ("G. General/limited partner", "General partner (LLC member-manager)"),
        ("H1. Domestic or foreign partner", "Domestic"),
        ("H2. Disregarded entity TIN", "N/A"),
        ("I1. Partner type", "Individual"),
        ("J. Partner's share of profit/loss/capital — beginning", "0.00% (new entity)"),
        ("J. Partner's share of profit/loss/capital — ending", f"{ownership_pct*100:.4f}%"),
        ("K. Partner's share of liabilities", "$0 (no recourse/non-recourse debt)"),
        ("L. Partner's capital account analysis", "See Schedule M-2 (Tab 9)"),
        ("M. Did partner contribute property with built-in gain/loss?", "No"),
        ("N. Partner's share of unrecognized §704(c) gain/loss", "$0"),
    ]
    for label, value in info_p2:
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws.cell(row=r, column=2, value=value)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        if "⚠️" in str(value):
            for c in range(1, 5):
                ws.cell(row=r, column=c).fill = WARN_FILL
        r += 1
    r += 1

    # Part III — Partner's Share of Current Year Income, Deductions, Credits, Other
    r = section(ws, r, "PART III — PARTNER'S SHARE OF INCOME / DEDUCTIONS / CREDITS", 4)
    headers = ["Box", "Item", "Amount ($)", "Notes"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=r, column=i, value=h)
    hdr_row(ws, r, 4)
    r += 1

    def kline(box, label, amount=None, notes="", bold=False, fill=None):
        nonlocal r
        ws.cell(row=r, column=1, value=box)
        ws.cell(row=r, column=2, value=label)
        if amount is not None:
            ws.cell(row=r, column=3, value=amount).number_format = MONEY
        ws.cell(row=r, column=4, value=notes).alignment = Alignment(wrap_text=True)
        if bold:
            for c in range(1, 5):
                ws.cell(row=r, column=c).font = Font(bold=True)
        if fill:
            for c in range(1, 5):
                ws.cell(row=r, column=c).fill = fill
        r += 1

    kline("1", "Ordinary business income (loss)", k1_income,
          f"= {ownership_pct*100:.4f}% of partnership ordinary income (Form 1065 Line 22). Reported on Schedule E Part II of partner's 1040.",
          bold=True, fill=TOTAL_FILL)
    kline("2", "Net rental real estate income (loss)", 0)
    kline("3", "Other net rental income (loss)", 0)
    kline("4a", "Guaranteed payments for services", 0)
    kline("4b", "Guaranteed payments for capital", 0)
    kline("5", "Interest income", 0)
    kline("6a", "Ordinary dividends", 0)
    kline("6b", "Qualified dividends", 0)
    kline("7", "Royalties", 0)
    kline("8", "Net short-term capital gain (loss)", 0)
    kline("9a", "Net long-term capital gain (loss)", 0)
    kline("10", "Net section 1231 gain (loss)", 0)
    kline("11", "Other income (loss)", 0)
    kline("12", "Section 179 deduction", 0)
    kline("13", "Other deductions", 0)
    kline("14", "Self-employment earnings (loss)", k1_income,
          "If partner is an active managing member, this is SE-taxable. Verify Raj/Phil's status.")
    kline("19A", "Cash distributions", cash_dist,
          "Cash actually received from the partnership in 2025. Reduces partner's outside basis / capital account.")
    kline("20Z", "QBI deduction info (Section 199A)", k1_income,
          "Qualified Business Income for §199A purposes. Capital management may be SSTB — confirm phase-out applicability.")
    r += 1

    # Summary box
    r = section(ws, r, "PARTNER'S TAX OBLIGATIONS — QUICK REFERENCE", 4)
    notes = [
        f"Box 1 amount (${k1_income:,.2f}) is ordinary income. Add to your Schedule E Part II, then to your 1040 Line 8.",
        f"Box 14 (${k1_income:,.2f}) is subject to Self-Employment Tax (Schedule SE) if you're an active managing member.",
        f"Box 19A (${cash_dist:,.2f}) cash distributions are NOT additionally taxed — they reduce your capital account / outside basis.",
        "Box 20Z provides QBI info for §199A 20% deduction (subject to income phase-out).",
        "Filing deadline: K-1 must be furnished to partner by March 15 (or extended deadline if partnership extended).",
    ]
    for note in notes:
        ws.cell(row=r, column=1, value=note).alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        ws.row_dimensions[r].height = 30
        r += 1

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 55


# ---------------------------------------------------------------------------
# Tab 7: Schedule L — Balance Sheet per Books
# ---------------------------------------------------------------------------

def build_schedule_l(wb, T):
    ws = wb.create_sheet("7. Schedule L")
    r = title(ws, "Schedule L — Balance Sheets per Books", ncols=5)
    headers = ["Line", "Asset / Liability / Equity", "Beginning of Year ($)", "End of Year ($)", "Notes"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=r, column=i, value=h)
    hdr_row(ws, r, 5)
    r += 1

    def line(line_no, label, beg=None, end=None, notes="", bold=False, fill=None, warn=False):
        nonlocal r
        ws.cell(row=r, column=1, value=line_no)
        ws.cell(row=r, column=2, value=label)
        if beg is not None:
            ws.cell(row=r, column=3, value=beg).number_format = MONEY
        if end is not None:
            ws.cell(row=r, column=4, value=end).number_format = MONEY
        ws.cell(row=r, column=5, value=notes).alignment = Alignment(wrap_text=True)
        if bold:
            for c in range(1, 6):
                ws.cell(row=r, column=c).font = Font(bold=True)
        if warn:
            for c in range(1, 6):
                ws.cell(row=r, column=c).fill = WARN_FILL
        elif fill:
            for c in range(1, 6):
                ws.cell(row=r, column=c).fill = fill
        r += 1

    # ASSETS
    r = section(ws, r, "ASSETS (Lines 1–14)", 5)
    line("1", "Cash", 0, None,
         "⚠️ TO PROVIDE — bank + crypto wallet balances as of 12/31/2025", warn=True)
    line("2a", "Trade notes and accounts receivable", 0, 0)
    line("2b", "Less: allowance for bad debts", 0, 0)
    line("3", "Inventories", 0, 0, "N/A — services only")
    line("4", "U.S. government obligations", 0, 0)
    line("5", "Tax-exempt securities", 0, 0)
    line("6", "Other current assets", 0, 10500,
         "Prepaid Insurance — $18K Dec policy less $7,500 to 2025 = $10,500 prepaid for 2026 coverage. ⚠️ Verify accountant chooses 12-month prepaid rule.")
    line("7", "Loans to partners (or persons related to partners)", 0, 0)
    line("8", "Mortgage and real estate loans", 0, 0)
    line("9", "Other investments", 0, T["spv_amount"],
         f"506c SPV Loan / Investment — ${T['spv_amount']:,.2f} (Aug $4,275 + Oct $25,000). Reclassified from P&L. ⚠️ Confirm whether loan receivable or equity investment.")
    line("10a", "Buildings and other depreciable assets", 0, 0)
    line("10b", "Less accumulated depreciation", 0, 0)
    line("11", "Depletable assets", 0, 0)
    line("11b", "Less accumulated depletion", 0, 0)
    line("12", "Land", 0, 0)
    line("13", "Intangible assets (amortizable only)", 0, 0)
    line("13b", "Less accumulated amortization", 0, 0)
    line("14", "Other assets (statement)", 0, 0)
    line("15", "TOTAL ASSETS",
         0, 10500 + T["spv_amount"],
         "Excludes cash placeholder", bold=True, fill=SUBTOTAL_FILL)
    r += 1

    # LIABILITIES + CAPITAL
    r = section(ws, r, "LIABILITIES AND CAPITAL (Lines 16–22)", 5)
    line("16", "Accounts payable", 0, 0,
         "⚠️ Verify with accountant — any unpaid contractor amounts as of 12/31?", warn=True)
    line("17", "Mortgages, notes, bonds payable in less than 1 year", 0, 0)
    line("18", "Other current liabilities (statement)", 0, 0)
    line("19a", "All nonrecourse loans", 0, 0)
    line("19b", "Loans from partners (or persons related to partners)", 0, 0)
    line("20", "Mortgages, notes, bonds payable in 1 year or more", 0, 0)
    line("21", "Other liabilities (statement)", 0, 0)
    line("—", "TOTAL LIABILITIES", 0, 0, "", bold=True, fill=SUBTOTAL_FILL)
    r += 1

    # PARTNERS' CAPITAL
    r = section(ws, r, "PARTNERS' CAPITAL ACCOUNTS", 5)
    line("22", "Partners' capital accounts", 0, T["ordinary_business_income"] - (T["nairne_cash"] + T["raj_cash"] + T["phil_cash"]),
         "= Net Income for the year − Cumulative Distributions. ⚠️ Add any member capital contributions.")
    r += 1

    line("23", "TOTAL LIABILITIES + CAPITAL",
         0,
         T["ordinary_business_income"] - (T["nairne_cash"] + T["raj_cash"] + T["phil_cash"]),
         "Should equal Total Assets when fully populated", bold=True, fill=TOTAL_FILL)

    r += 2
    r = section(ws, r, "BALANCE CHECK / NOTES", 5)
    notes = [
        "This Schedule L is best-effort built from transaction data. PLACEHOLDER cells (cash, AP, member capital contributions) need to be populated by the accountant from bank statements and capital records.",
        f"Total Assets shown (${10500 + T['spv_amount']:,.2f}) excludes cash. Once Cash + member capital contributions are filled in, the sheet should balance.",
        "First-year filing: beginning-of-year balances are $0 (entity was formed at the Armada Prime relaunch in Aug 2025).",
        "If the operating agreement specifies different capital accounts at formation (e.g., contributions made by Nairne / Raj / Phil), those go on Schedule M-2 and roll into partners' capital here.",
    ]
    for note in notes:
        ws.cell(row=r, column=1, value=note).alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        ws.row_dimensions[r].height = 35
        r += 1

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 60


# ---------------------------------------------------------------------------
# Tab 8: Schedule M-1 — Income Reconciliation
# ---------------------------------------------------------------------------

def build_schedule_m1(wb, T):
    ws = wb.create_sheet("8. Schedule M-1")
    r = title(ws, "Schedule M-1 — Reconciliation of Income (Loss) per Books With Income (Loss) per Return")
    headers = ["Line", "Description", "Amount ($)", "Notes"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=r, column=i, value=h)
    hdr_row(ws, r, 4)
    r += 1

    def line(line_no, label, amount=None, notes="", bold=False, fill=None):
        nonlocal r
        ws.cell(row=r, column=1, value=line_no)
        ws.cell(row=r, column=2, value=label)
        if amount is not None:
            ws.cell(row=r, column=3, value=amount).number_format = MONEY
        ws.cell(row=r, column=4, value=notes).alignment = Alignment(wrap_text=True)
        if bold:
            for c in range(1, 5):
                ws.cell(row=r, column=c).font = Font(bold=True)
        if fill:
            for c in range(1, 5):
                ws.cell(row=r, column=c).fill = fill
        r += 1

    line("1", "Net income (loss) per books", T["ordinary_business_income"],
         "Should equal Form 1065 Page 1 Line 22 (assuming books and tax return reconcile)")
    line("2", "Income included on Schedule K (lines 1, 2, etc.) NOT recorded on books this year", 0, "")
    line("3", "Guaranteed payments to partners (other than health insurance)", 0, "")
    line("4", "Expenses recorded on books this year not deducted on Schedule K (e.g., depreciation)", 0, "")
    line("5", "Add lines 1–4", T["ordinary_business_income"], "", bold=True, fill=SUBTOTAL_FILL)
    line("6", "Income recorded on books this year not included on Schedule K", 0, "")
    line("7", "Deductions included on Schedule K not charged against book income this year", 0, "")
    line("8", "Add lines 6 and 7", 0, "")
    line("9", "Income (loss) — Analysis of Net Income (Loss) (Sch K Line 1)", T["ordinary_business_income"],
         "= Line 5 − Line 8. Should match Schedule K Line 1.", bold=True, fill=TOTAL_FILL)

    r += 2
    r = section(ws, r, "NOTES", 4)
    notes = [
        "Schedule M-1 reconciles BOOK income (your accounting records) to TAXABLE income (Schedule K).",
        "For Armada Prime Tech LLC's first year, books and tax should reconcile cleanly — no permanent or temporary differences expected.",
        "The 506c SPV reclassification ($29,275 from P&L to balance sheet) is reflected in BOTH books and tax — no M-1 adjustment needed.",
        "Insurance ($18K full deduction in 2025) is consistent between books and tax under the 12-month prepaid rule.",
    ]
    for note in notes:
        ws.cell(row=r, column=1, value=note).alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        ws.row_dimensions[r].height = 30
        r += 1

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 55


# ---------------------------------------------------------------------------
# Tab 9: Schedule M-2 — Partners' Capital Accounts
# ---------------------------------------------------------------------------

def build_schedule_m2(wb, T):
    ws = wb.create_sheet("9. Schedule M-2")
    r = title(ws, "Schedule M-2 — Analysis of Partners' Capital Accounts", ncols=5)
    headers = ["Line", "Description", "Nairne ($)", "Raj Duggal ($)", "Phil ($)"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=r, column=i, value=h)
    hdr_row(ws, r, 5)
    r += 1

    def line(line_no, label, nairne=None, raj=None, phil=None, bold=False, fill=None, warn=False):
        nonlocal r
        ws.cell(row=r, column=1, value=line_no)
        ws.cell(row=r, column=2, value=label)
        if nairne is not None:
            ws.cell(row=r, column=3, value=nairne).number_format = MONEY
        if raj is not None:
            ws.cell(row=r, column=4, value=raj).number_format = MONEY
        if phil is not None:
            ws.cell(row=r, column=5, value=phil).number_format = MONEY
        if bold:
            for c in range(1, 6):
                ws.cell(row=r, column=c).font = Font(bold=True)
        if warn:
            for c in range(1, 6):
                ws.cell(row=r, column=c).fill = WARN_FILL
        elif fill:
            for c in range(1, 6):
                ws.cell(row=r, column=c).fill = fill
        r += 1

    line("1", "Balance at beginning of year", nairne=0, raj=0, phil=0)
    line("2a", "Capital contributed during year — cash", None, None, None,
         warn=True)
    line("2b", "Capital contributed during year — property", None, None, None, warn=True)
    line("3", "Net income (loss) — partner's share",
         nairne=T["nairne_k1_income"], raj=T["raj_k1_income"], phil=T["phil_k1_income"], bold=True)
    line("4", "Other increases (statement)", nairne=0, raj=0, phil=0)
    line("5", "Sum of lines 1–4",
         nairne=T["nairne_k1_income"], raj=T["raj_k1_income"], phil=T["phil_k1_income"],
         bold=True, fill=SUBTOTAL_FILL)
    line("6a", "Distributions of cash",
         nairne=-T["nairne_cash"], raj=-T["raj_cash"], phil=-T["phil_cash"])
    line("6b", "Distributions of property", nairne=0, raj=0, phil=0)
    line("7", "Other decreases (statement)", nairne=0, raj=0, phil=0)
    line("8", "Subtract sum of lines 6 and 7 from line 5",
         nairne=T["nairne_k1_income"] - T["nairne_cash"],
         raj=T["raj_k1_income"] - T["raj_cash"],
         phil=T["phil_k1_income"] - T["phil_cash"],
         bold=True, fill=TOTAL_FILL)

    r += 2
    r = section(ws, r, "ANALYSIS — Negative Capital Account Risk", 5)
    notes = [
        f"Nairne ending capital: ${T['nairne_k1_income'] - T['nairne_cash']:,.2f} (BEFORE any 2025 capital contributions)",
        f"Raj ending capital: ${T['raj_k1_income'] - T['raj_cash']:,.2f} (BEFORE any 2025 capital contributions)",
        f"Phil ending capital: ${T['phil_k1_income'] - T['phil_cash']:,.2f} (BEFORE any 2025 capital contributions)",
        "If a partner's capital account goes negative, the cash distributions exceeded their share of income + capital. This typically means the partner needs to make additional capital contributions OR the distributions are recharacterized.",
        "For Nairne: her cash distributions ($85,996.83 if combined Fund Mgmt + direct) exceed her K-1 income share. Either she has substantial 2025 capital contributions to add to Line 2a, OR the operating agreement may treat her Fund Mgmt portion as a guaranteed payment (which would add it back to her income).",
        "Recommend: confirm with Monily whether any portion of the $85K Nairne received should be recharacterized as a guaranteed payment (Form 1065 Line 10) to avoid negative capital.",
    ]
    for note in notes:
        ws.cell(row=r, column=1, value=note).alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        ws.row_dimensions[r].height = 35
        r += 1

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18


# ---------------------------------------------------------------------------
# Tab 10: 1099 Contractor Detail
# ---------------------------------------------------------------------------

def build_1099_detail(wb, T):
    ws = wb.create_sheet("10. 1099 Detail")
    r = title(ws, "1099-NEC Contractor Detail (Supporting Form 1099 Issuance)", ncols=8)

    months = ["2025-08", "2025-09", "2025-10", "2025-11", "2025-12"]
    headers = ["Recipient"] + [PERIOD_LABELS[p] for p in months] + ["2025 Total ($)", "1099 Required?"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=r, column=i, value=h)
    hdr_row(ws, r, len(headers))
    r += 1

    contractors = [
        ("Alec Atkinson", "Alec Atkinson", "Yes (>$600)"),
        ("Jake Gordon", "Jake Gordon", "Yes (>$600)"),
        ("AJ Affleck", "AJ Affleck", "Yes (>$600)"),
        ("Issac Morris", "Issac", "Yes (>$600)"),
        ("Luke Affleck", "Luke", "No (<$600)"),
        ("Nikki", "Nikki", "No (<$600 — single Nov payment)"),
    ]
    grand = 0
    for display, key, required in contractors:
        ws.cell(row=r, column=1, value=display)
        row_total = 0
        for j, period in enumerate(months):
            v = ACTUAL_PAID.get(period, {}).get(key, 0)
            ws.cell(row=r, column=2+j, value=v if v else None)
            if v:
                ws.cell(row=r, column=2+j).number_format = MONEY
            row_total += v
        ws.cell(row=r, column=2+len(months), value=row_total).number_format = MONEY
        ws.cell(row=r, column=2+len(months)).font = Font(bold=True)
        ws.cell(row=r, column=3+len(months), value=required)
        if "No" in required:
            for c in range(1, 4+len(months)):
                ws.cell(row=r, column=c).fill = WARN_FILL
        grand += row_total
        r += 1

    # Chris (operating contractor)
    ws.cell(row=r, column=1, value="Chris (operating contractor)")
    chris_amounts = {"2025-08": 0, "2025-09": 0, "2025-10": 0, "2025-11": 7500, "2025-12": 4000}
    for j, period in enumerate(months):
        v = chris_amounts[period]
        ws.cell(row=r, column=2+j, value=v if v else None)
        if v:
            ws.cell(row=r, column=2+j).number_format = MONEY
    ws.cell(row=r, column=2+len(months), value=11500).number_format = MONEY
    ws.cell(row=r, column=2+len(months)).font = Font(bold=True)
    ws.cell(row=r, column=3+len(months), value="Yes (>$600)")
    grand += 11500
    r += 1

    # Total
    ws.cell(row=r, column=1, value="TOTAL")
    ws.cell(row=r, column=2+len(months), value=grand).number_format = MONEY
    for c in range(1, 4+len(months)):
        ws.cell(row=r, column=c).fill = TOTAL_FILL
        ws.cell(row=r, column=c).font = Font(bold=True)
    r += 2

    r = section(ws, r, "NOTES", len(headers))
    notes = [
        "The IRS 1099-NEC threshold is $600. Recipients receiving less than $600 in a calendar year do not require a 1099-NEC.",
        "Phil is NOT on this list — Phil is a K-1 partner (per Nairne 2026-05-05). See Tab 6.",
        "Each recipient's SSN/EIN and address must be obtained (W-9) before issuing 1099-NEC.",
        "1099-NECs must be furnished to recipients by January 31, 2026, and filed with IRS by January 31, 2026.",
        "Vendors (PVD, Ad Spend providers, Website builder, Alpha Verification) may also require 1099-NEC if individuals/sole proprietors and >$600. Verify their W-9 status.",
    ]
    for note in notes:
        ws.cell(row=r, column=1, value=note).alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(headers))
        ws.row_dimensions[r].height = 30
        r += 1

    ws.column_dimensions["A"].width = 28
    for j in range(len(months)):
        ws.column_dimensions[get_column_letter(2+j)].width = 12
    ws.column_dimensions[get_column_letter(2+len(months))].width = 14
    ws.column_dimensions[get_column_letter(3+len(months))].width = 28


# ---------------------------------------------------------------------------
# Tab 11: Op Expenses Detail
# ---------------------------------------------------------------------------

def build_op_detail(wb, T):
    ws = wb.create_sheet("11. Op Expenses Detail")
    r = title(ws, "Operating Expenses — Vendor × Month Detail", ncols=8)

    months = ["2025-08", "2025-09", "2025-10", "2025-11", "2025-12"]
    headers = ["Vendor / Item", "Form 1065 Line"] + [PERIOD_LABELS[p] for p in months] + ["2025 Total ($)"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=r, column=i, value=h)
    hdr_row(ws, r, len(headers))
    r += 1

    matrix = {}
    for period, items in GP_OP_EXPENSES.items():
        for vendor, amount in items:
            matrix.setdefault(vendor, {})[period] = amount

    line_map = {
        "Insurance": "Line 20 (Other) — Insurance D&O",
        "Chris": "Line 20 (Other) — Contractor labor",
        "PVD": "Line 20 (Other) — Professional fees",
        "Website": "Line 20 (Other) — Marketing",
        "Ad Spend": "Line 20 (Other) — Marketing",
        "Alpha Verification": "Line 20 (Other) — Compliance",
        "Formidium (TPA)": "Line 20 (Other) — Admin/TPA",
        "TPA (Formidium)": "Line 20 (Other) — Admin/TPA",
        "TPA (second line)": "Line 20 (Other) — Admin/TPA",
        "506c SPV Loan": "*** RECLASSIFIED to Schedule L (Asset)",
    }

    grand = 0
    spv_total = 0
    for vendor in sorted(matrix.keys(), key=lambda v: -sum(matrix[v].values())):
        ws.cell(row=r, column=1, value=vendor)
        ws.cell(row=r, column=2, value=line_map.get(vendor, ""))
        row_total = 0
        for j, period in enumerate(months):
            v = matrix[vendor].get(period, 0)
            ws.cell(row=r, column=3+j, value=v if v else None)
            if v:
                ws.cell(row=r, column=3+j).number_format = MONEY
            row_total += v
        ws.cell(row=r, column=3+len(months), value=row_total).number_format = MONEY
        ws.cell(row=r, column=3+len(months)).font = Font(bold=True)
        if "RECLASSIFIED" in line_map.get(vendor, ""):
            for c in range(1, len(headers)+1):
                ws.cell(row=r, column=c).fill = WARN_FILL
            spv_total += row_total
        else:
            grand += row_total
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="P&L OP EXPENSES (Form 1065 Line 20)")
    ws.cell(row=r, column=3+len(months), value=grand).number_format = MONEY
    for c in range(1, len(headers)+1):
        ws.cell(row=r, column=c).fill = TOTAL_FILL
        ws.cell(row=r, column=c).font = Font(bold=True)
    r += 1

    ws.cell(row=r, column=1, value="RECLASSIFIED TO SCHEDULE L (Balance Sheet)")
    ws.cell(row=r, column=3+len(months), value=spv_total).number_format = MONEY
    for c in range(1, len(headers)+1):
        ws.cell(row=r, column=c).fill = WARN_FILL
        ws.cell(row=r, column=c).font = Font(bold=True)
    r += 1

    ws.cell(row=r, column=1, value="GRAND TOTAL CASH OUTFLOWS")
    ws.cell(row=r, column=3+len(months), value=grand + spv_total).number_format = MONEY
    for c in range(1, len(headers)+1):
        ws.cell(row=r, column=c).fill = SUBTOTAL_FILL
        ws.cell(row=r, column=c).font = Font(bold=True)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 38
    for j in range(len(months)):
        ws.column_dimensions[get_column_letter(3+j)].width = 12
    ws.column_dimensions[get_column_letter(3+len(months))].width = 14


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    T = compute_totals()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    build_cover(wb, T)
    build_page1(wb, T)
    build_schedule_k(wb, T)
    build_k1(wb, "Nairne", T["nairne_pct"], T["nairne_k1_income"], T["nairne_cash"], 4)
    build_k1(wb, "Raj Duggal", T["raj_pct"], T["raj_k1_income"], T["raj_cash"], 5)
    build_k1(wb, "Phil", T["phil_pct"], T["phil_k1_income"], T["phil_cash"], 6)
    build_schedule_l(wb, T)
    build_schedule_m1(wb, T)
    build_schedule_m2(wb, T)
    build_1099_detail(wb, T)
    build_op_detail(wb, T)

    wb.save(OUT_XLSX)
    print(f"Wrote {OUT_XLSX}")
    print(f"\nKEY NUMBERS (Accrual Basis, Tax-Optimized):")
    print(f"  Revenue:              ${T['revenue']:>12,.2f}")
    print(f"  Total Deductions:     ${T['total_deductions']:>12,.2f}")
    print(f"  Ordinary Bus Income:  ${T['ordinary_business_income']:>12,.2f}")
    print(f"  Nairne K-1 (98.36%):  ${T['nairne_k1_income']:>12,.2f}")
    print(f"  Raj K-1 (0.82%):      ${T['raj_k1_income']:>12,.2f}")
    print(f"  Phil K-1 (0.82%):     ${T['phil_k1_income']:>12,.2f}")
